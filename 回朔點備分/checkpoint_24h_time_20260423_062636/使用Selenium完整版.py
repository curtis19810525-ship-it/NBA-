"""
使用 Selenium 爬取 Playsport MLB 賽事資料並寫入「每日數據」分頁
（國際盤讓分 + 比例；allianceid=1）
支援單日爬取（接受日期參數 YYYYMMDD，語意為台灣日）
"""

import os
import sys
import time
import re
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, Border
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
import requests

# Selenium 導入
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("錯誤: 未安裝 Selenium")
    print("請執行: pip install selenium")
    sys.exit(1)

# 引入設定檔
try:
    from config import NBA_XLSX_FILE
except ImportError:
    NBA_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"

# 球隊名稱對應表
TEAM_NAME_MAPPING = {
    '塞爾提': '塞爾提克',
    '塞爾蒂克': '塞爾提克',
    '塞爾提克': '塞爾提克',
}


def extract_score_and_team(text):
    """
    從文字中分離分數和球隊名稱
    例如：'110塞爾提' -> (110, '塞爾提克')
          '公鹿' -> (None, '公鹿')
          '76人' -> (None, '76人')  # 76人為隊名，非分數+單字
    
    Returns:
        tuple: (分數, 球隊名稱) 或 (None, 球隊名稱)
    """
    if not text:
        return None, ''
    text = text.strip()
    # 特殊：76人為隊名（可能為「76人」或「76人 客 43%...」），不當作「分數+人」
    m76 = re.match(r'^(\d+人)(?:\s|$|%|，|、)', text) or re.match(r'^(\d+人)$', text)
    if m76:
        return None, m76.group(1)
    # 匹配「數字+中文」格式，例如：110塞爾提、94魔術國
    match = re.match(r'^(\d+)([\u4e00-\u9fff]+)$', text)
    if match:
        score = match.group(1)
        team_name = match.group(2)
        # 標準化球隊名稱
        normalized_team = TEAM_NAME_MAPPING.get(team_name, team_name)
        return score, normalized_team
    
    # 如果沒有數字，直接返回球隊名稱
    # 移除可能的數字前綴（如果有的話）
    team_name = re.sub(r'^\d+', '', text.strip())
    if team_name:
        normalized_team = TEAM_NAME_MAPPING.get(team_name, team_name)
        return None, normalized_team
    
    return None, ''


def setup_driver(headless=False):
    """設定 Chrome 瀏覽器驅動程式"""
    if not SELENIUM_AVAILABLE:
        return None
        
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception as e:
        print(f"無法啟動 Chrome 瀏覽器: {e}")
        print("\n請確保：")
        print("1. 已安裝 Chrome 瀏覽器")
        print("2. 已安裝 selenium: pip install selenium")
        print("3. ChromeDriver 會自動下載（Selenium 4.x）")
        return None


def get_page_with_requests(url):
    """
    使用 requests 直接抓取網頁（快速方法）
    如果內容是靜態的，這個方法會比 Selenium 快很多（1-2秒 vs 15-20秒）
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        
        print("嘗試使用 requests 直接抓取（快速方法）...")
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # 檢查是否有關鍵內容（對戰資訊連結）
            battle_links = soup.find_all('a', href=re.compile(r'/gamesData/battle'))
            if battle_links:
                print(f"✓ 成功使用 requests 抓取（找到 {len(battle_links)} 個對戰連結）")
                return soup
            else:
                print("⚠ requests 抓取成功，但未找到對戰資訊連結（可能需要 JavaScript）")
                return None
        else:
            print(f"⚠ HTTP 錯誤: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"⚠ requests 方法失敗: {e}")
        return None


def get_page_with_selenium(url, wait_time=15):
    """使用 Selenium 載入網頁（備用方法，用於需要 JavaScript 的情況）"""
    driver = setup_driver(headless=True)  # 改為 headless 模式，更快
    if not driver:
        return None
    
    try:
        print("使用 Selenium 載入網頁（備用方法）...")
        print(f"網址: {url}")
        driver.get(url)
        
        print(f"等待 {wait_time} 秒讓 JavaScript 執行...")
        time.sleep(wait_time)
        
        # 等待表格出現
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            print("✓ 表格已載入")
        except:
            print("⚠ 未找到表格，繼續解析...")
        
        # 獲取頁面 HTML
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        return soup
        
    except Exception as e:
        print(f"✗ Selenium 載入失敗: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        print("關閉瀏覽器...")
        driver.quit()


def fetch_playsport_data(date_str):
    """
    從 Playsport 爬取指定日期的賽事資料
    
    Args:
        date_str: 日期字串，格式為 YYYYMMDD
        
    Returns:
        list: 賽事資料列表，每個元素為一個字典
    """
    url = f"https://www.playsport.cc/predict/scale?allianceid=1&gametime={date_str}&sid=0"
    
    print(f"正在爬取資料...")
    print(f"網址: {url}")
    print()
    
    # 優先使用 requests 直接抓取（快速方法）
    soup = get_page_with_requests(url)
    
    # 如果 requests 失敗或找不到資料，改用 Selenium（備用方法）
    if not soup:
        print("\n改用 Selenium 方法（備用方案）...")
        soup = get_page_with_selenium(url, wait_time=15)
    
    if not soup:
        print("❌ 無法載入網頁（兩種方法都失敗）")
        return []
    
    try:
        games = []
        
        # 尋找包含對戰資訊的連結
        battle_links = soup.find_all('a', href=re.compile(r'/gamesData/battle'))
        print(f"找到 {len(battle_links)} 個對戰資訊連結")
        
        if not battle_links:
            print("⚠ 警告: 未找到對戰資訊連結")
            return []
        
        # 解析每場比賽
        for link_idx, link in enumerate(battle_links):
            try:
                row = link.find_parent('tr')
                if not row:
                    print(f"    ⚠ 第 {link_idx + 1} 個連結：找不到包含連結的表格行")
                    continue
                
                game_data = {
                    'time': '',
                    'guest_team': '',
                    'guest_score': '',  # 客隊分數
                    'guest_spread': '',
                    'guest_percentage': '',
                    'home_team': '',
                    'home_score': '',  # 主隊分數
                    'home_spread': '',
                    'home_percentage': '',
                }
                
                # 獲取該行的所有儲存格
                cells = row.find_all(['td', 'th'])
                cell_texts = [cell.get_text(separator=' ', strip=True) for cell in cells]
                row_text = ' | '.join(cell_texts)
                
                # 解析時間
                time_match = re.search(r'(AM|PM)\s+(\d{2}:\d{2})', row_text, re.IGNORECASE)
                if time_match:
                    game_data['time'] = time_match.group(2)
                else:
                    time_match = re.search(r'(\d{3})\s+(AM|PM)\s+(\d{2}:\d{2})', row_text, re.IGNORECASE)
                    if time_match:
                        game_data['time'] = time_match.group(3)
                    else:
                        time_match = re.search(r'(\d{2}:\d{2})', row_text)
                        if time_match:
                            game_data['time'] = time_match.group(1)
                
                # 方法1：直接從 td.td-teaminfo 中提取球隊資訊和分數（最準確的方法）
                team_info_cell = row.find('td', class_='td-teaminfo')
                if team_info_cell:
                    # 在 td.td-teaminfo 中找到嵌套的 table
                    nested_table = team_info_cell.find('table')
                    if nested_table:
                        # 提取分數：從 td.scores ul li 中
                        scores_cell = nested_table.find('td', class_='scores')
                        if scores_cell:
                            score_items = scores_cell.find_all('li')
                            if len(score_items) >= 3:
                                # 第一個 li 是第一個分數，第三個 li 是第二個分數
                                score1_text = score_items[0].get_text(strip=True)
                                score2_text = score_items[2].get_text(strip=True)
                                
                                # 判斷哪個是獲勝隊伍的分數（有 winnerscores class）
                                if 'winnerscores' in score_items[0].get('class', []):
                                    # 第一個分數是獲勝隊伍
                                    winner_score = score1_text
                                    loser_score = score2_text
                                elif 'winnerscores' in score_items[2].get('class', []):
                                    # 第二個分數是獲勝隊伍
                                    winner_score = score2_text
                                    loser_score = score1_text
                                else:
                                    # 如果沒有 winnerscores，假設第一個是客隊，第二個是主隊
                                    winner_score = None
                                    loser_score = None
                                
                                # 根據 HTML 結構判斷主客隊
                                # 第一個 tr 中的 td 是上方隊伍（通常是客隊）
                                # 第三個 tr 中的 td 是下方隊伍（通常是主隊）
                                
                                # 找到所有 tr，確定球隊位置
                                all_trs = nested_table.find_all('tr')
                                if len(all_trs) >= 3:
                                    # 第一個 tr 包含上方隊伍
                                    first_tr_team_td = all_trs[0].find('td', class_=re.compile(r'(winnerteam|secondteam)'))
                                    third_tr_team_td = all_trs[2].find('td', class_=re.compile(r'(winnerteam|secondteam)'))
                                    
                                    if first_tr_team_td and third_tr_team_td:
                                        first_tr_team_link = first_tr_team_td.find('a')
                                        third_tr_team_link = third_tr_team_td.find('a')
                                        
                                        if first_tr_team_link and third_tr_team_link:
                                            # 上方隊伍是客隊，下方隊伍是主隊
                                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(first_tr_team_link.get_text(strip=True), first_tr_team_link.get_text(strip=True))
                                            game_data['home_team'] = TEAM_NAME_MAPPING.get(third_tr_team_link.get_text(strip=True), third_tr_team_link.get_text(strip=True))
                                            
                                            # 分配分數：第一個 li 對應上方隊伍（客隊），第三個 li 對應下方隊伍（主隊）
                                            game_data['guest_score'] = score1_text
                                            game_data['home_score'] = score2_text
                                else:
                                    # 如果找不到完整的 tr 結構，使用備用方法：直接從 winnerteam 和 secondteam 提取
                                    winner_team_cell = nested_table.find('td', class_='winnerteam')
                                    second_team_cell = nested_table.find('td', class_='secondteam')
                                    
                                    if winner_team_cell and second_team_cell:
                                        winner_link = winner_team_cell.find('a')
                                        second_link = second_team_cell.find('a')
                                        
                                        if winner_link and second_link:
                                            # 假設 winnerteam 在上方（客隊），secondteam 在下方（主隊）
                                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(winner_link.get_text(strip=True), winner_link.get_text(strip=True))
                                            game_data['home_team'] = TEAM_NAME_MAPPING.get(second_link.get_text(strip=True), second_link.get_text(strip=True))
                                            game_data['guest_score'] = score1_text
                                            game_data['home_score'] = score2_text
                
                # 方法2：如果方法1沒找到，從下一行提取主隊和客隊（備用方法）
                next_row = row.find_next_sibling('tr')
                home_team_cell0 = ''
                
                if not game_data['guest_team'] or not game_data['home_team']:
                    next_cells = next_row.find_all(['td', 'th']) if next_row else []
                    # 提取主隊（儲存格0）
                    if len(next_cells) > 0:
                        cell0_text = next_cells[0].get_text(strip=True)
                        if cell0_text and '對戰' not in cell0_text and '資訊' not in cell0_text:
                            # 檢查是否包含讓分信息（包含"分"和"贏"/"輸"、或"E"/"主"開頭的讓分格式）
                            if not (re.search(r'\d+分(贏|輸)', cell0_text) or '%' in cell0_text or 
                                    re.match(r'^(E|主)[+-]?\d+\.?\d*', cell0_text) or 
                                    ',' in cell0_text or ('.' in cell0_text and re.search(r'\d+\.\d+', cell0_text))):
                                # 使用新函數分離分數和球隊名稱
                                score, team_name = extract_score_and_team(cell0_text)
                                # 加強檢查：確保不是讓分信息
                                if (team_name and '預測' not in team_name and '比例' not in team_name and 
                                    '讓分' not in team_name and '分' not in team_name and 
                                    '贏' not in team_name and '輸' not in team_name and '%' not in team_name and
                                    len(team_name) >= 2):
                                    home_team_cell0 = team_name
                                    game_data['home_team'] = team_name
                                    if score:
                                        game_data['home_score'] = score
                                else:
                                    # 如果新函數沒找到，使用舊方法
                                    number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell0_text)
                                    if number_match:
                                        home_team_cell0 = number_match.group(0)
                                        if '預測' not in home_team_cell0 and '比例' not in home_team_cell0 and '讓分' not in home_team_cell0:
                                            score, team_name = extract_score_and_team(home_team_cell0)
                                            game_data['home_team'] = team_name
                                            if score:
                                                game_data['home_score'] = score
                                            home_team_cell0 = team_name
                                    else:
                                        chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell0_text)
                                        common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                                        valid_home = [name for name in chinese_names if name not in common_words and 2 <= len(name) <= 4 and '對戰' not in name and '資訊' not in name and '預測' not in name]
                                        if valid_home:
                                            home_team_cell0 = valid_home[0]
                                            game_data['home_team'] = TEAM_NAME_MAPPING.get(home_team_cell0, home_team_cell0)
                    
                    # 從下一行其他儲存格提取客隊
                    for cell_idx in range(1, len(next_cells)):
                        cell = next_cells[cell_idx]
                        cell_text = cell.get_text(strip=True)
                        if not cell_text or '對戰' in cell_text or '資訊' in cell_text:
                            continue
                        
                        # 檢查是否包含讓分信息（包含"分"和"贏"/"輸"、或"E"/"主"開頭的讓分格式）
                        if (re.search(r'\d+分(贏|輸)', cell_text) or '%' in cell_text or 
                            re.match(r'^(E|主)[+-]?\d+\.?\d*', cell_text) or 
                            ',' in cell_text or '.' in cell_text and re.search(r'\d+\.\d+', cell_text)):
                            # 這是讓分信息，不是球隊名稱，跳過
                            continue
                        
                        # 使用新函數分離分數和球隊名稱
                        score, team_name = extract_score_and_team(cell_text)
                        # 加強檢查：確保不是讓分信息
                        if (team_name and '預測' not in team_name and '比例' not in team_name and 
                            '讓分' not in team_name and '分' not in team_name and 
                            '贏' not in team_name and '輸' not in team_name and '%' not in team_name and
                            team_name != home_team_cell0 and len(team_name) >= 2):
                            game_data['guest_team'] = team_name
                            if score:
                                game_data['guest_score'] = score
                            break
                        
                        # 如果新函數沒找到，使用舊方法
                        number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell_text)
                        if number_match:
                            team_name_raw = number_match.group(0)
                            if '預測' not in team_name_raw and '比例' not in team_name_raw and '讓分' not in team_name_raw and team_name_raw != home_team_cell0:
                                score, team_name = extract_score_and_team(team_name_raw)
                                game_data['guest_team'] = team_name
                                if score:
                                    game_data['guest_score'] = score
                                break
                        
                        chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell_text)
                        common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                        valid_teams = [name for name in chinese_names if name not in common_words and 2 <= len(name) <= 4 and '對戰' not in name and '資訊' not in name and '預測' not in name and name != home_team_cell0]
                        if valid_teams:
                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(valid_teams[0], valid_teams[0])
                            break
                    
                    # 如果下一行沒找到客隊，從當前行提取客隊
                    if not game_data['guest_team']:
                        current_cells = row.find_all(['td', 'th'])
                        for cell in current_cells:
                            cell_text = cell.get_text(strip=True)
                            if not cell_text or '對戰' in cell_text or '資訊' in cell_text:
                                continue
                            
                            # 檢查是否包含讓分信息（包含"分"和"贏"/"輸"、或"E"/"主"開頭的讓分格式）
                            if (re.search(r'\d+分(贏|輸)', cell_text) or '%' in cell_text or 
                                re.match(r'^(E|主)[+-]?\d+\.?\d*', cell_text) or 
                                ',' in cell_text or ('.' in cell_text and re.search(r'\d+\.\d+', cell_text))):
                                # 這是讓分信息，不是球隊名稱，跳過
                                continue
                            
                            # 使用新函數分離分數和球隊名稱
                            score, team_name = extract_score_and_team(cell_text)
                            # 加強檢查：確保不是讓分信息
                            if (team_name and '預測' not in team_name and '比例' not in team_name and 
                                '讓分' not in team_name and '分' not in team_name and 
                                '贏' not in team_name and '輸' not in team_name and '%' not in team_name and
                                team_name != home_team_cell0 and len(team_name) >= 2):
                                game_data['guest_team'] = team_name
                                if score:
                                    game_data['guest_score'] = score
                                break
                            
                            # 如果新函數沒找到，使用舊方法
                            number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell_text)
                            if number_match:
                                team_name_raw = number_match.group(0)
                                if '預測' not in team_name_raw and '比例' not in team_name_raw and '讓分' not in team_name_raw and team_name_raw != home_team_cell0:
                                    score, team_name = extract_score_and_team(team_name_raw)
                                    game_data['guest_team'] = team_name
                                    if score:
                                        game_data['guest_score'] = score
                                    break
                            
                            chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell_text)
                            common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                            valid_teams = [name for name in chinese_names if name not in common_words and 2 <= len(name) <= 4 and '對戰' not in name and '資訊' not in name and '預測' not in name and name != home_team_cell0]
                            if valid_teams:
                                game_data['guest_team'] = TEAM_NAME_MAPPING.get(valid_teams[0], valid_teams[0])
                                break
                
                # 方法2：從連結提取球隊（row + next_row）
                if not game_data['guest_team'] or not game_data['home_team']:
                    team_links = row.find_all('a', href=re.compile(r'/gamesData/teams'))
                    if next_row:
                        team_links = team_links + next_row.find_all('a', href=re.compile(r'/gamesData/teams'))
                    if len(team_links) >= 2:
                        guest_team_raw = TEAM_NAME_MAPPING.get(team_links[0].get_text(strip=True), team_links[0].get_text(strip=True))
                        home_team_raw = TEAM_NAME_MAPPING.get(team_links[1].get_text(strip=True), team_links[1].get_text(strip=True))
                        if not game_data['guest_team']:
                            game_data['guest_team'] = guest_team_raw
                        if not game_data['home_team']:
                            game_data['home_team'] = home_team_raw
                    elif len(team_links) == 1:
                        link_team = TEAM_NAME_MAPPING.get(team_links[0].get_text(strip=True), team_links[0].get_text(strip=True))
                        if not game_data['guest_team'] and link_team != game_data.get('home_team', ''):
                            game_data['guest_team'] = link_team
                        elif not game_data['home_team'] and link_team != game_data.get('guest_team', ''):
                            game_data['home_team'] = link_team
                
                # 方法3：如果還是沒找到，從當前行提取球隊名稱（含 76人 等）
                if not game_data['guest_team'] or not game_data['home_team']:
                    all_chinese = re.findall(r'[\u4e00-\u9fff]{2,4}', row_text)
                    all_chinese += re.findall(r'\d+人', row_text)
                    # 補齊 common_words，包含所有可能的非球隊名稱關鍵字
                    common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客', '小', '大', '輸', '贏', '分', '對', '戰', '資', '訊'}
                    # 過濾掉單字和常見關鍵字
                    all_valid_teams = [name for name in all_chinese if name not in common_words and len(name) >= 2]
                    all_valid_teams = [name for name in all_valid_teams if '對戰' not in name and '資訊' not in name and '預測' not in name]
                    
                    # 進一步過濾：排除明顯不是球隊名稱的單字
                    invalid_single_chars = {'主', '客', '小', '大', '輸', '贏', '分', '對', '戰', '資', '訊', '預', '測', '比', '例', '讓', '場', '隊', '賽', '時', '間', '國', '際', '運', '彩', '不', '讓', '會', '員', '推', '近', '日', '過'}
                    all_valid_teams = [name for name in all_valid_teams if name not in invalid_single_chars and len(name) >= 2]
                    
                    # 去重但保持順序
                    seen = set()
                    unique_teams = []
                    for team in all_valid_teams:
                        if team not in seen:
                            seen.add(team)
                            unique_teams.append(team)
                    
                    # 只在使用方法3時，確保提取到的球隊名稱不是單字
                    if len(unique_teams) >= 2:
                        # 檢查提取到的球隊名稱是否有效（不是單字或常見關鍵字）
                        valid_guest = unique_teams[0] if unique_teams[0] not in invalid_single_chars and len(unique_teams[0]) >= 2 else None
                        valid_home = unique_teams[1] if unique_teams[1] not in invalid_single_chars and len(unique_teams[1]) >= 2 else None
                        
                        if valid_guest and valid_home:
                            if not game_data['guest_team']:
                                game_data['guest_team'] = TEAM_NAME_MAPPING.get(valid_guest, valid_guest)
                            if not game_data['home_team']:
                                game_data['home_team'] = TEAM_NAME_MAPPING.get(valid_home, valid_home)
                
                # 最終檢查：確保沒有"對戰資訊"、"預測"等被當作球隊名稱
                invalid_keywords = ['對戰', '資訊', '對戰資訊', '預測', '人預測', '比例', '讓分']
                # 單字關鍵字（不應該被當作球隊名稱）
                invalid_single_chars = {'主', '客', '小', '大', '輸', '贏', '分', '對', '戰', '資', '訊', '預', '測', '比', '例', '讓', '場', '隊', '賽', '時', '間', '國', '際', '運', '彩', '不', '讓', '會', '員', '推', '近', '日', '過'}
                
                if game_data['guest_team']:
                    # 檢查是否包含讓分信息（包含"分"和"贏"/"輸"、或"E"/"主"開頭的讓分格式）
                    if (re.search(r'\d+分(贏|輸)', game_data['guest_team']) or '%' in game_data['guest_team'] or
                        re.match(r'^(E|主)[+-]?\d+\.?\d*', game_data['guest_team']) or 
                        ',' in game_data['guest_team'] or ('.' in game_data['guest_team'] and re.search(r'\d+\.\d+', game_data['guest_team']))):
                        game_data['guest_team'] = ''
                    # 檢查是否為單字或包含無效關鍵字
                    elif game_data['guest_team'] in invalid_single_chars or len(game_data['guest_team']) < 2:
                        game_data['guest_team'] = ''
                    else:
                        for keyword in invalid_keywords:
                            if keyword in game_data['guest_team']:
                                game_data['guest_team'] = ''
                                break
                
                if game_data['home_team']:
                    # 檢查是否包含讓分信息（包含"分"和"贏"/"輸"、或"E"/"主"開頭的讓分格式）
                    if (re.search(r'\d+分(贏|輸)', game_data['home_team']) or '%' in game_data['home_team'] or
                        re.match(r'^(E|主)[+-]?\d+\.?\d*', game_data['home_team']) or 
                        ',' in game_data['home_team'] or ('.' in game_data['home_team'] and re.search(r'\d+\.\d+', game_data['home_team']))):
                        game_data['home_team'] = ''
                    # 檢查是否為單字或包含無效關鍵字
                    elif game_data['home_team'] in invalid_single_chars or len(game_data['home_team']) < 2:
                        game_data['home_team'] = ''
                    else:
                        for keyword in invalid_keywords:
                            if keyword in game_data['home_team']:
                                game_data['home_team'] = ''
                                break
                
                # 提取讓分和比例
                combined_text = row_text
                next_row_text = ''
                if next_row:
                    next_row_text = next_row.get_text(separator=' ', strip=True)
                    combined_text = row_text + ' | ' + next_row_text
                
                # 提取客隊讓分（僅寫入頁面有出現的「客」讓分文字；支援小數分）
                guest_spread_match = re.search(r'客\s*(\d+(?:\.\d+)?)\s*分\s*(輸|贏)', combined_text)
                if guest_spread_match:
                    spread_points = guest_spread_match.group(1)
                    spread_result = guest_spread_match.group(2)
                    game_data['guest_spread'] = f"{spread_points}分{spread_result}"
                
                # 提取主隊讓分（僅寫入頁面有出現的「主」讓分文字）
                home_spread_match = re.search(r'主\s*(\d+(?:\.\d+)?)\s*分\s*(輸|贏)', combined_text)
                if home_spread_match:
                    spread_points = home_spread_match.group(1)
                    spread_result = home_spread_match.group(2)
                    game_data['home_spread'] = f"{spread_points}分{spread_result}"
                
                # 提取比例（國際盤優先）
                # 客隊只在上列(row_text)抓，主隊只在下列(next_row_text)抓，避免誤抓到運彩盤讓分比例。
                guest_percentage_match = re.search(
                    r'客\s*(?:\d+(?:\.\d+)?\s*分\s*(?:輸|贏))?\s*50%[^0-9%]*(\d+)%\s*\d+\s*人預測',
                    row_text
                )
                if not guest_percentage_match:
                    guest_percentage_match = re.search(r'客[^0-9%]*(\d+)%\s*\d+\s*人預測', row_text)
                if guest_percentage_match:
                    game_data['guest_percentage'] = guest_percentage_match.group(1) + '%'

                home_percentage_match = re.search(
                    r'主\s*(?:\d+(?:\.\d+)?\s*分\s*(?:輸|贏))?\s*50%[^0-9%]*(\d+)%\s*\d+\s*人預測',
                    next_row_text
                )
                if not home_percentage_match:
                    home_percentage_match = re.search(r'主[^0-9%]*(\d+)%\s*\d+\s*人預測', next_row_text)
                if home_percentage_match:
                    game_data['home_percentage'] = home_percentage_match.group(1) + '%'
                
                # 如果找到客隊和主隊，則加入列表
                if game_data['guest_team'] and game_data['home_team']:
                    games.append(game_data)
                    print(f"    ✓ 解析第 {len(games)} 場: {game_data['time']} {game_data['guest_team']} vs {game_data['home_team']}")
                else:
                    print(f"    ⚠ 跳過（缺少球隊資訊）:")
                    print(f"       時間: '{game_data['time']}'")
                    print(f"       客隊: '{game_data['guest_team']}'")
                    print(f"       主隊: '{game_data['home_team']}'")
                    
            except Exception as e:
                print(f"    ⚠ 解析比賽時發生錯誤: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        if not games:
            print("⚠ 警告：未找到任何賽事資料")
        else:
            print(f"\n✓ 成功解析 {len(games)} 場賽事")
            
        return games
        
    except Exception as e:
        print(f"❌ 解析錯誤: {e}")
        import traceback
        traceback.print_exc()
        return []


def sort_daily_data_by_date(ws):
    """
    對「每日數據」分頁按日期排序（舊到新，一列一場）
    A-J 欄：日期、比賽時間、客隊分數、客隊、讓分、比例、主隊分數、主隊、讓分、比例
    跨日期時插入 1 個空行
    注意：此函數會重新寫入整個分頁，請確保檔案已備份
    """
    try:
        max_rows_to_process = 5000
        data_rows = []  # [(date_str, row_data), ...]，row_data 為 10 欄的 list
        
        for row in range(2, min(ws.max_row + 1, max_rows_to_process + 1)):
            date_value = ws[f'A{row}'].value
            if not date_value:
                continue
            if isinstance(date_value, datetime):
                date_str = date_value.strftime("%Y%m%d")
            else:
                date_str = str(date_value).strip()
            if not date_str or len(date_str) != 8:
                continue
            
            row_data = []
            for col in range(1, 11):  # A 到 J
                cell = ws.cell(row=row, column=col)
                alignment_info = None
                if cell.alignment:
                    try:
                        alignment_info = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    except:
                        alignment_info = Alignment(horizontal="center", vertical="center")
                font_info = None
                if cell.font:
                    try:
                        font_info = Font(
                            name=cell.font.name, size=cell.font.size,
                            bold=cell.font.bold, italic=cell.font.italic,
                            underline=cell.font.underline, color=cell.font.color
                        )
                    except:
                        font_info = None
                row_data.append({
                    'value': cell.value,
                    'alignment': alignment_info,
                    'number_format': cell.number_format,
                    'font': font_info
                })
            data_rows.append((date_str, row_data))
        
        if not data_rows:
            print("⚠ 未找到任何資料列，跳過排序")
            return
        
        # 按日期排序：舊到新（ascending）
        data_rows.sort(key=lambda x: x[0] if x[0] else '')
        
        # 清空從第 2 行開始，A 到 J
        for row in range(2, min(ws.max_row + 1, 10000)):
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.value = None
                cell.border = Border()
        
        current_row = 2
        prev_date = None
        for date_str, row_data in data_rows:
            if prev_date and prev_date != date_str:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).border = Border()
                current_row += 1
            for col_idx, cell_data in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                if cell_data.get('value') is not None:
                    cell.value = cell_data['value']
                if cell_data.get('alignment'):
                    try:
                        cell.alignment = cell_data['alignment']
                    except:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell_data.get('number_format'):
                    cell.number_format = cell_data['number_format']
                if cell_data.get('font'):
                    try:
                        cell.font = cell_data['font']
                    except:
                        pass
                cell.border = Border()
            current_row += 1
            prev_date = date_str
        
        print(f"✓ 已按日期排序（舊到新）")
        
    except Exception as e:
        print(f"⚠ 排序時發生錯誤（不影響資料寫入）：{e}")
        import traceback
        traceback.print_exc()
        # 重要：如果排序失敗，資料已經被清空，但無法重新寫入
        # 這是一個嚴重的問題，需要從備份還原
        print("⚠ 警告：排序失敗可能導致資料遺失，請檢查檔案並從備份還原")


def write_to_daily_data_sheet(excel_file, date_str, games):
    """
    將爬取的資料寫入「每日數據」分頁
    
    Args:
        excel_file: Excel 檔案路徑
        date_str: 日期字串（YYYYMMDD格式）
        games: 比賽資料列表
    """
    if not os.path.exists(excel_file):
        print(f"✗ 檔案不存在：{excel_file}")
        return False
    
    # 檢查檔案是否被鎖定（正在使用中）
    try:
        # 嘗試以讀取模式開啟檔案，檢查是否被鎖定
        test_file = open(excel_file, 'r+b')
        test_file.close()
    except PermissionError:
        print(f"✗ 檔案正在被其他程式使用：{excel_file}")
        print(f"  請關閉 Excel 或其他正在使用此檔案的程式後再試")
        return False
    except Exception as e:
        print(f"⚠ 檢查檔案時發生錯誤：{e}")
    
    try:
        wb = load_workbook(excel_file, read_only=False)
        
        # 檢查「每日數據」分頁
        if "每日數據" not in wb.sheetnames:
            print("✗ 找不到「每日數據」分頁")
            wb.close()
            return False
        
        ws = wb["每日數據"]
        # 確保第 1 列為新格式表頭（圖2）
        header = ["日期", "比賽時間", "客隊分數", "客隊", "讓分(客)", "比例(客)", "主隊分數", "主隊", "讓分(主)", "比例(主)"]
        for col, title in enumerate(header, start=1):
            ws.cell(row=1, column=col).value = title
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=col).border = Border()

        # 預檢：該日無賽事則不寫入、不刪除
        if len(games) == 0:
            print(f"  {date_str} 無賽事資料，跳過寫入")
            wb.close()
            return True

        # 方案1：找出該日既有列（A 欄 = date_str 的連續區塊）
        def _cell_date_str(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.strftime("%Y%m%d")
            return str(val).strip()

        rows_same_date = []
        for row in range(2, ws.max_row + 1):
            v = ws[f'A{row}'].value
            if v is None:
                if rows_same_date:
                    break  # 已找到該日區塊，遇空白即停止
                continue
            if _cell_date_str(v) == date_str:
                rows_same_date.append(row)
            elif rows_same_date:
                break  # 遇到不同日期即停止

        old_count = len(rows_same_date)
        if old_count > 0 and len(games) < old_count:
            print(f"  ⚠ {date_str} 新 {len(games)} 場 < 原 {old_count} 場，可能延賽/取消，仍覆蓋")

        # 由下往上刪除該日既有列，避免列號錯位
        for r in reversed(rows_same_date):
            ws.delete_rows(r, 1)

        # 一列一場：找到最後一筆資料列（A 欄有日期的最後一行）
        last_row = 1
        for row in range(2, ws.max_row + 1):
            if ws[f'A{row}'].value:
                last_row = row

        last_row_date = None
        if last_row > 1:
            v = ws[f'A{last_row}'].value
            if v:
                last_row_date = _cell_date_str(v)

        if last_row_date and last_row_date != date_str:
            start_row = last_row + 2  # 跨日期：空一列後開始
        else:
            start_row = last_row + 1

        games_sorted = sorted(games, key=lambda x: x.get('time', '00:00'))
        
        for game_idx, game in enumerate(games_sorted):
            r = start_row + game_idx
            ws[f'A{r}'] = date_str
            ws[f'B{r}'] = game.get('time', '')
            guest_score = game.get('guest_score', '')
            home_score = game.get('home_score', '')
            if guest_score:
                try:
                    ws[f'C{r}'] = int(guest_score)
                except:
                    ws[f'C{r}'] = guest_score
            ws[f'D{r}'] = game.get('guest_team', '')
            ws[f'E{r}'] = game.get('guest_spread', '')
            guest_percentage = game.get('guest_percentage', '')
            if guest_percentage:
                try:
                    pct = float(guest_percentage.replace('%', '')) / 100
                    ws[f'F{r}'] = pct
                    ws[f'F{r}'].number_format = '0%'
                except:
                    ws[f'F{r}'] = guest_percentage
            if home_score:
                try:
                    ws[f'G{r}'] = int(home_score)
                except:
                    ws[f'G{r}'] = home_score
            ws[f'H{r}'] = game.get('home_team', '')
            ws[f'I{r}'] = game.get('home_spread', '')
            home_percentage = game.get('home_percentage', '')
            if home_percentage:
                try:
                    pct = float(home_percentage.replace('%', '')) / 100
                    ws[f'J{r}'] = pct
                    ws[f'J{r}'].number_format = '0%'
                except:
                    ws[f'J{r}'] = home_percentage
            elif guest_percentage:
                try:
                    pct = float(guest_percentage.replace('%', '')) / 100
                    ws[f'J{r}'] = 1 - pct if pct <= 1 else (100 - pct) / 100
                    ws[f'J{r}'].number_format = '0%'
                except:
                    pass
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                c = ws[f'{col}{r}']
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border()
        
        # 對整個「每日數據」分頁按日期排序（舊到新）
        # 注意：排序功能已修復，但建議先備份檔案
        try:
            sort_daily_data_by_date(ws)
        except Exception as e:
            print(f"⚠ 排序時發生錯誤，已跳過排序：{e}")
            # 即使排序失敗，也不影響資料寫入
            import traceback
            traceback.print_exc()
        
        # 儲存檔案
        try:
            wb.save(excel_file)
            print(f"\n✓ 已將 {len(games_sorted)} 場比賽寫入「每日數據」分頁")
            print(f"  起始行：{start_row}")
            return True
        except PermissionError:
            print(f"  ✗ 無法儲存檔案：檔案可能正在被其他程式使用")
            print(f"  提示：請關閉 Excel 後再試")
            wb.close()
            return False
        
    except KeyError as e:
        error_msg = str(e)
        if "[Content_Types].xml" in error_msg or "archive" in error_msg.lower():
            print(f"✗ Excel 檔案損壞或格式異常：{excel_file}")
            print(f"  錯誤訊息：{error_msg}")
            print(f"\n  可能的解決方案：")
            print(f"  1. 檢查檔案是否為有效的 .xlsx 格式")
            print(f"  2. 嘗試在 Excel 中開啟檔案，確認檔案是否損壞")
            print(f"  3. 如果檔案損壞，請從備份還原")
            print(f"  4. 確認檔案路徑正確：{excel_file}")
        else:
            print(f"✗ 讀取 Excel 檔案時發生錯誤：{error_msg}")
        return False
    except PermissionError:
        print(f"✗ 無法開啟檔案：檔案可能正在被其他程式使用")
        print(f"  請關閉 Excel 後再試")
        return False
    except Exception as e:
        error_msg = str(e)
        print(f"✗ 寫入時發生錯誤：{error_msg}")
        print(f"\n  錯誤類型：{type(e).__name__}")
        if "archive" in error_msg.lower() or "zip" in error_msg.lower():
            print(f"  這可能是 Excel 檔案損壞的徵兆")
            print(f"  建議：嘗試在 Excel 中開啟檔案檢查")
        import traceback
        traceback.print_exc()
        return False
    finally:
        try:
            wb.close()
        except:
            pass


def main():
    """主函數"""
    print("=" * 60)
    print("使用 Selenium 爬取 MLB 賽事資料（Playsport 國際盤）")
    print("=" * 60)
    
    # 檢查參數
    if len(sys.argv) >= 2:
        target_date = sys.argv[1].strip()
    else:
        # 如果沒有參數，使用今天的日期
        target_date = datetime.now().strftime("%Y%m%d")
    
    # 驗證日期格式
    if not re.match(r'^\d{8}$', target_date):
        print("✗ 日期格式錯誤！")
        print("  請使用格式：YYYYMMDD（例如：20251022）")
        sys.exit(1)
    
    print(f"\n目標日期：{target_date}")
    print(f"Excel 檔案：{NBA_XLSX_FILE}")
    print("=" * 60)
    
    # 爬取資料
    games = fetch_playsport_data(target_date)

    if not games:
        print("\n⚠ 未找到任何賽事資料（該日可能無比賽、尚未開盤或頁面結構變更）")

    # 寫入 Excel（無賽事時仍會安全結束，不覆寫其它日期）
    if write_to_daily_data_sheet(NBA_XLSX_FILE, target_date, games):
        print("\n" + "=" * 60)
        print("✓ 執行完成！")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("✗ 執行失敗！")
        print("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    main()
