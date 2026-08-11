# -*- coding: utf-8 -*-
"""
匯出 MLB「紀錄」為 NotebookLM 資料來源用條列文字。

定案：
  - 格式：條列
  - 空值：_
  - 金流：72%
  - 賽果日／關注日各存一份；剪貼簿預設複製關注日
  - 檔案位置：專案目錄下 exports\\NotebookLM_紀錄_YYYYMMDD.txt
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime

from openpyxl import load_workbook

from analysis_helpers import _norm_date, _norm_time, team_nickname
from config import MLB_XLSX_FILE
from fill_observation import (
    SHEET_RECORD,
    _cell_str,
    parse_team_venue,
    record_block_cols,
)

EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
EMPTY = "_"
GAME_LABEL_RE = re.compile(r"第\s*\d+\s*場")


def _is_error_value(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s.startswith("#") or s in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!"}


def _fmt_empty(value) -> str:
    if value is None:
        return EMPTY
    if _is_error_value(value):
        return EMPTY
    s = str(value).strip()
    return s if s else EMPTY


def _fmt_flow(value) -> str:
    if value is None or _is_error_value(value) or str(value).strip() == "":
        return EMPTY
    try:
        v = float(value)
    except (TypeError, ValueError):
        s = str(value).strip()
        if s.endswith("%"):
            return s
        return s or EMPTY
    pct = round(v * 100) if 0 < abs(v) <= 1 else round(v)
    return f"{pct}%"


def _fmt_number(value) -> str:
    if value is None or _is_error_value(value) or str(value).strip() == "":
        return EMPTY
    try:
        v = float(value)
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.3f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return _fmt_empty(value)


def read_record_export_games(ws, focus_date: str) -> list[dict]:
    """讀取紀錄某日場次（含合理讓分／讓分／金流／得分／結果）。"""
    for dc in record_block_cols(ws):
        date_row = None
        for r in range(1, (ws.max_row or 1) + 1):
            if _norm_date(ws.cell(r, dc).value) == focus_date:
                date_row = r
                break
        if date_row is None:
            continue

        games: list[dict] = []
        r = date_row + 1
        blanks = 0
        while r <= (ws.max_row or 1):
            a = ws.cell(r, dc).value
            cell_date = _norm_date(a)
            if cell_date and cell_date != focus_date:
                break
            label = _cell_str(a)
            if GAME_LABEL_RE.match(label):
                team_venue = _cell_str(ws.cell(r, dc + 1).value)
                time_str = _norm_time(ws.cell(r + 1, dc).value)
                fair = ws.cell(r + 1, dc + 2).value
                handicap = ws.cell(r + 1, dc + 3).value
                flow = ws.cell(r + 1, dc + 4).value
                quant = ws.cell(r + 1, dc + 5).value
                score_fav = ws.cell(r + 1, dc + 6).value
                score_dog = ws.cell(r + 1, dc + 7).value
                result = ws.cell(r + 1, dc + 8).value
                team, side = parse_team_venue(team_venue)
                if team:
                    games.append(
                        {
                            "game_label": re.sub(r"\s+", "", label),
                            "time": time_str or EMPTY,
                            "fav_team": team_nickname(team) or team,
                            "side": side or EMPTY,
                            "fair": fair,
                            "handicap": handicap,
                            "flow": flow,
                            "quant": quant,
                            "score_fav": score_fav,
                            "score_dog": score_dog,
                            "result": result,
                        }
                    )
                blanks = 0
                r += 2
                continue
            if label == "" and _cell_str(ws.cell(r + 1, dc).value) == "":
                blanks += 1
                if blanks >= 3:
                    break
            r += 1
        if games:
            return games
    return []


def format_notebooklm_text(focus_date: str, games: list[dict], *, exported_at: datetime | None = None) -> str:
    exported_at = exported_at or datetime.now()
    lines = [
        "【MLB 紀錄匯出｜資料來源】",
        f"日期：{focus_date}",
        "來源檔：MLB26-27數據.xlsx → 紀錄",
        "匯出範圍：A1:I33（有比賽區；實際依日期區塊解析）",
        f"匯出時間：{exported_at:%Y-%m-%d %H:%M:%S}",
        "",
        "說明：",
        "- 本檔供 NotebookLM「資料來源」上傳／貼上用",
        "- 一場包含：第N場、時間、讓分球隊、主客、合理讓分、讓分、金流、合理性、得分、結果",
        "",
        "========================================",
    ]
    if not games:
        lines.append("（此日期在「紀錄」查無場次）")
    else:
        for g in games:
            lines.extend(
                [
                    g.get("game_label") or "第?場",
                    f"時間：{_fmt_empty(g.get('time'))}",
                    f"讓分球隊：{_fmt_empty(g.get('fav_team'))}",
                    f"主客：{_fmt_empty(g.get('side'))}",
                    f"合理讓分：{_fmt_number(g.get('fair'))}",
                    f"讓分：{_fmt_empty(g.get('handicap'))}",
                    f"金流：{_fmt_flow(g.get('flow'))}",
                    f"合理性：{_fmt_empty(g.get('quant'))}",
                    f"讓分方得分：{_fmt_number(g.get('score_fav'))}",
                    f"受讓方得分：{_fmt_number(g.get('score_dog'))}",
                    f"結果：{_fmt_empty(g.get('result'))}",
                    "",
                ]
            )
    lines.append("========================================")
    lines.append(f"合計場次：{len(games)}")
    lines.append("")
    return "\n".join(lines)


def copy_text_to_clipboard(text: str) -> None:
    """Windows 優先用 pywin32；失敗則嘗試 clip。"""
    if sys.platform == "win32":
        try:
            import win32clipboard

            win32clipboard.OpenClipboard()
            try:
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, text)
            finally:
                win32clipboard.CloseClipboard()
            return
        except Exception:
            pass
        try:
            import subprocess

            subprocess.run(
                ["clip"],
                input=text.encode("utf-16"),
                check=True,
            )
            return
        except Exception as e:
            raise SystemExit(f"無法寫入剪貼簿：{e}") from e
    raise SystemExit("目前僅支援在 Windows 複製到剪貼簿。檔案已儲存，請手動開啟複製。")


def export_record_for_notebooklm(
    focus_date: str,
    *,
    mlb_xlsx: str | None = None,
    export_dir: str | None = None,
    copy_to_clipboard: bool = False,
) -> tuple[str, int]:
    """
    匯出單一日期。回傳 (檔案路徑, 場次數)。
    """
    mlb_xlsx = mlb_xlsx or MLB_XLSX_FILE
    export_dir = export_dir or EXPORT_DIR
    os.makedirs(export_dir, exist_ok=True)

    if not os.path.exists(mlb_xlsx):
        raise SystemExit(f"找不到檔案：{mlb_xlsx}")

    wb = load_workbook(mlb_xlsx, data_only=True)
    try:
        if SHEET_RECORD not in wb.sheetnames:
            raise SystemExit("MLB26-27 找不到「紀錄」分頁")
        games = read_record_export_games(wb[SHEET_RECORD], focus_date)
    finally:
        wb.close()

    text = format_notebooklm_text(focus_date, games)
    out_path = os.path.join(export_dir, f"NotebookLM_紀錄_{focus_date}.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)

    if copy_to_clipboard:
        copy_text_to_clipboard(text)

    return out_path, len(games)


def export_results_and_focus(
    results_day: str,
    focus_day: str,
    *,
    mlb_xlsx: str | None = None,
) -> dict:
    """匯出賽果日＋關注日；剪貼簿只放關注日。"""
    r_path, r_n = export_record_for_notebooklm(
        results_day, mlb_xlsx=mlb_xlsx, copy_to_clipboard=False
    )
    f_path, f_n = export_record_for_notebooklm(
        focus_day, mlb_xlsx=mlb_xlsx, copy_to_clipboard=True
    )
    export_dir = os.path.dirname(os.path.abspath(f_path))
    return {
        "export_dir": export_dir,
        "results_path": r_path,
        "results_games": r_n,
        "focus_path": f_path,
        "focus_games": f_n,
        # 舊鍵別名（相容）
        "yesterday_path": r_path,
        "yesterday_games": r_n,
        "today_path": f_path,
        "today_games": f_n,
    }


def export_yesterday_and_today(
    yesterday: str,
    today: str,
    *,
    mlb_xlsx: str | None = None,
) -> dict:
    """相容舊名稱：yesterday＝賽果日、today＝關注日。"""
    return export_results_and_focus(yesterday, today, mlb_xlsx=mlb_xlsx)
