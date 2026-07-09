"""
將「每日數據」分頁從舊格式（2列一場）轉成新格式（1列一場 A-J）
並寫入新表頭、排序：日期舊→新、同一天依比賽時間、跨日空一列
執行前請先關閉 Excel 並建議備份檔案。
"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border

try:
    from config import NBA_XLSX_FILE
except ImportError:
    NBA_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"

# 新格式表頭（圖2）
HEADER_ROW = [
    "日期",
    "比賽時間",
    "客隊分數",
    "客隊",
    "讓分(客)",
    "比例(客)",
    "主隊分數",
    "主隊",
    "讓分(主)",
    "比例(主)",
]


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y%m%d")
    s = str(val).strip()
    if not s:
        return None
    return s


def _parse_time(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    return str(val).strip()


def _parse_percentage(val):
    """將儲存格轉成 0~1 的數字（寫入時用 0% 格式）。若無法解析回傳 None。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val <= 1:
            return float(val)
        return float(val) / 100
    s = str(val).strip().replace("%", "")
    if not s:
        return None
    try:
        x = float(s)
        return x / 100 if x > 1 else x
    except ValueError:
        return None


def _is_numeric_score(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        int(str(val).strip())
        return True
    except ValueError:
        return False


def read_games_old_format(ws):
    """
    讀取舊格式：每場 2 列。
    第 1 列：A=日期, B=時間, C=客隊, D=客讓分, E=客比例, F=主隊, G=主讓分, H=主比例
    第 2 列：C=客隊分數, F=主隊分數（若為數字）；若 F 為文字則當主隊名，主隊分數可能無
    """
    games = []
    processed = set()
    max_row = getattr(ws, "max_row", 0) or 0

    row = 2
    while row <= max_row:
        if row in processed:
            row += 1
            continue
        a1 = ws[f"A{row}"].value
        if not a1:
            row += 1
            continue

        date_str = _parse_date(a1)
        if not date_str:
            row += 1
            continue

        time_val = ws[f"B{row}"].value
        time_str = _parse_time(time_val)

        c1 = ws[f"C{row}"].value
        d1 = ws[f"D{row}"].value
        e1 = ws[f"E{row}"].value
        f1 = ws[f"F{row}"].value
        g1 = ws[f"G{row}"].value
        h1 = ws[f"H{row}"].value

        row2 = row + 1
        if row2 > max_row:
            row += 1
            continue

        c2 = ws[f"C{row2}"].value
        f2 = ws[f"F{row2}"].value
        g2 = ws[f"G{row2}"].value
        h2 = ws[f"H{row2}"].value

        # 客隊、客讓分、客比例：以第 1 列為準
        guest_team = (c1 if c1 is not None else "") if not _is_numeric_score(c1) else ""
        guest_spread = d1
        guest_pct = _parse_percentage(e1)

        # 主隊、主讓分、主比例：第 1 列有則用，否則用第 2 列
        if f1 is not None and not _is_numeric_score(f1):
            home_team = str(f1).strip()
            home_spread = g1
            home_pct = _parse_percentage(h1)
        else:
            home_team = (str(f2).strip() if f2 is not None and not _is_numeric_score(f2) else "") or ""
            home_spread = g2 if g2 is not None else g1
            home_pct = _parse_percentage(h2) if h2 is not None else _parse_percentage(h1)

        # 客隊分數、主隊分數：第 2 列 C、F 若為數字則為分數
        guest_score = None
        home_score = None
        if _is_numeric_score(c2):
            try:
                guest_score = int(float(str(c2).strip()))
            except (ValueError, TypeError):
                guest_score = c2
        if _is_numeric_score(f2):
            try:
                home_score = int(float(str(f2).strip()))
            except (ValueError, TypeError):
                home_score = f2

        # 若第 1 列 C 為數字（少數舊檔），當客隊分數
        if guest_score is None and _is_numeric_score(c1):
            try:
                guest_score = int(float(str(c1).strip()))
            except (ValueError, TypeError):
                guest_score = c1
            guest_team = ""

        games.append({
            "date": date_str,
            "time": time_str,
            "guest_score": guest_score,
            "guest_team": guest_team or (str(c1).strip() if c1 else ""),
            "guest_spread": guest_spread,
            "guest_pct": guest_pct,
            "home_score": home_score,
            "home_team": home_team,
            "home_spread": home_spread,
            "home_pct": home_pct,
        })
        processed.add(row)
        processed.add(row2)
        row += 2

    return games


def convert_daily_data_to_new_format(excel_file):
    """
    將「每日數據」由舊格式轉成新格式並排序。
    - 表頭：日期、比賽時間、客隊分數、客隊、讓分(客)、比例(客)、主隊分數、主隊、讓分(主)、比例(主)
    - 一列一場 A-J，日期舊→新，同一天依比賽時間，跨日空一列
    """
    if not os.path.exists(excel_file):
        print(f"✗ 檔案不存在：{excel_file}")
        return False

    try:
        f = open(excel_file, "r+b")
        f.close()
    except PermissionError:
        print(f"✗ 檔案使用中：{excel_file}")
        print("  請關閉 Excel 後再執行")
        return False

    try:
        wb = load_workbook(excel_file, read_only=False)
        if "每日數據" not in wb.sheetnames:
            print("✗ 找不到「每日數據」分頁")
            wb.close()
            return False

        ws = wb["每日數據"]
        games = read_games_old_format(ws)
        if not games:
            print("⚠ 未讀到任何一場比賽（請確認是否為舊格式 2 列一場）")
            wb.close()
            return False

        # 排序：日期舊→新，同一天依比賽時間
        games.sort(key=lambda g: (g["date"], g["time"] or "00:00"))

        # 清空從第 2 列開始的內容（保留第 1 列待會重寫表頭），清到足夠多列
        max_row = max(ws.max_row, 2 + len(games) * 2 + 100)
        for r in range(2, max_row + 1):
            for c in range(1, 11):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = Border()

        # 寫入新表頭（第 1 列）
        for col, title in enumerate(HEADER_ROW, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = title
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border()

        # 寫入資料：一列一場，跨日空一列
        current_row = 2
        prev_date = None
        for g in games:
            if prev_date is not None and prev_date != g["date"]:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).border = Border()
                current_row += 1
            prev_date = g["date"]

            ws.cell(row=current_row, column=1).value = g["date"]
            ws.cell(row=current_row, column=2).value = g["time"]
            # C 客隊分數
            if g["guest_score"] is not None:
                ws.cell(row=current_row, column=3).value = g["guest_score"]
            ws.cell(row=current_row, column=4).value = g["guest_team"]
            ws.cell(row=current_row, column=5).value = g["guest_spread"]
            if g["guest_pct"] is not None:
                cell = ws.cell(row=current_row, column=6)
                cell.value = g["guest_pct"]
                cell.number_format = "0%"
            # G 主隊分數
            if g["home_score"] is not None:
                ws.cell(row=current_row, column=7).value = g["home_score"]
            ws.cell(row=current_row, column=8).value = g["home_team"]
            ws.cell(row=current_row, column=9).value = g["home_spread"]
            if g["home_pct"] is not None:
                cell = ws.cell(row=current_row, column=10)
                cell.value = g["home_pct"]
                cell.number_format = "0%"
            elif g["guest_pct"] is not None:
                cell = ws.cell(row=current_row, column=10)
                cell.value = 1 - g["guest_pct"]
                cell.number_format = "0%"

            for col in range(1, 11):
                c = ws.cell(row=current_row, column=col)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border()
            current_row += 1

        wb.save(excel_file)
        wb.close()
        print(f"✓ 已轉換 {len(games)} 場比賽為新格式（1 列一場），並已排序（日期舊→新、跨日空一列）")
        return True

    except PermissionError:
        print("✗ 無法儲存，請關閉 Excel 後再試")
        return False
    except Exception as e:
        print(f"✗ 轉換失敗：{e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    import sys
    excel_file = NBA_XLSX_FILE
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    print("=" * 60)
    print("每日數據：舊格式（2 列一場）→ 新格式（1 列一場 + 新表頭）")
    print("=" * 60)
    print(f"檔案：{excel_file}")
    print("建議先備份後再執行。")
    print()
    ok = convert_daily_data_to_new_format(excel_file)
    print()
    if ok:
        print("執行完成。")
    else:
        print("未完成轉換。")
    return 0 if ok else 1


if __name__ == "__main__":
    exit(main())
