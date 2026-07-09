"""
修正「每日數據」分頁中 20260220~20260223 的分數位置（一列一場 A-J）
若 D/H 欄為「分數+球隊」則拆成 C/D、G/H
"""

import os
import sys
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 引入設定檔
try:
    from config import NBA_XLSX_FILE
except ImportError:
    NBA_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"

# 球隊名稱對應表
TEAM_NAME_MAPPING = {
    '塞爾提': '塞爾提克',
    '塞爾蒂克': '塞爾提克',
    '塞爾提克': '塞爾提克',
}


def extract_score_and_team(text):
    """
    從文字中分離分數和球隊名稱
    例如：'110塞爾提' -> (110, '塞爾提克')
          '公鹿' -> (None, '公鹿')
    
    Returns:
        tuple: (分數, 球隊名稱) 或 (None, 球隊名稱)
    """
    if not text:
        return None, ''
    
    # 匹配「數字+中文」格式，例如：110塞爾提、94魔術國
    match = re.match(r'^(\d+)([\u4e00-\u9fff]+)$', str(text).strip())
    if match:
        score = match.group(1)
        team_name = match.group(2)
        # 標準化球隊名稱
        normalized_team = TEAM_NAME_MAPPING.get(team_name, team_name)
        return score, normalized_team
    
    # 如果沒有數字，直接返回球隊名稱
    team_name = re.sub(r'^\d+', '', str(text).strip())
    if team_name:
        normalized_team = TEAM_NAME_MAPPING.get(team_name, team_name)
        return None, normalized_team
    
    return None, ''


def fix_scores_for_dates(excel_file, date_list):
    """
    修正指定日期的分數位置
    
    Args:
        excel_file: Excel 檔案路徑
        date_list: 日期列表，格式為 ['20260220', '20260221', ...]
    """
    if not os.path.exists(excel_file):
        print(f"✗ 檔案不存在：{excel_file}")
        return False
    
    # 檢查檔案是否被鎖定
    try:
        test_file = open(excel_file, 'r+b')
        test_file.close()
    except PermissionError:
        print(f"✗ 檔案正在被其他程式使用：{excel_file}")
        print(f"  請關閉 Excel 或其他正在使用此檔案的程式後再試")
        return False
    
    try:
        print(f"\n正在讀取檔案：{excel_file}")
        wb = load_workbook(excel_file, read_only=False)
        
        # 檢查「每日數據」分頁
        if "每日數據" not in wb.sheetnames:
            print("✗ 找不到「每日數據」分頁")
            wb.close()
            return False
        
        ws = wb["每日數據"]
        print(f"✓ 找到「每日數據」分頁")
        
        # 轉換日期列表為字串格式（確保格式一致）
        date_str_list = [str(date) for date in date_list]
        
        fixed_count = 0
        
        # 一列一場：A=日期, B=時間, C=客分, D=客隊, E=客讓分, F=客比例, G=主分, H=主隊, I=主讓分, J=主比例
        for row in range(2, ws.max_row + 1):
            date_value = ws[f'A{row}'].value
            if not date_value:
                continue
            if isinstance(date_value, datetime):
                date_str = date_value.strftime("%Y%m%d")
            else:
                date_str = str(date_value).strip()
            if date_str not in date_str_list:
                continue

            guest_team_value = ws[f'D{row}'].value
            home_team_value = ws[f'H{row}'].value
            guest_score = None
            home_score = None
            guest_team_clean = None
            home_team_clean = None

            if guest_team_value:
                score, team_name = extract_score_and_team(guest_team_value)
                if score:
                    guest_score = score
                    guest_team_clean = team_name
                    print(f"  Row {row}: 客隊 '{guest_team_value}' -> 分數: {guest_score}, 球隊: {guest_team_clean}")
            if home_team_value:
                score, team_name = extract_score_and_team(home_team_value)
                if score:
                    home_score = score
                    home_team_clean = team_name
                    print(f"  Row {row}: 主隊 '{home_team_value}' -> 分數: {home_score}, 球隊: {home_team_clean}")

            if guest_score or home_score:
                if guest_team_clean:
                    ws[f'D{row}'] = guest_team_clean
                if home_team_clean:
                    ws[f'H{row}'] = home_team_clean
                if guest_score:
                    try:
                        ws[f'C{row}'] = int(guest_score)
                    except Exception:
                        ws[f'C{row}'] = guest_score
                if home_score:
                    try:
                        ws[f'G{row}'] = int(home_score)
                    except Exception:
                        ws[f'G{row}'] = home_score
                if guest_score:
                    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")
                if home_score:
                    ws[f'G{row}'].alignment = Alignment(horizontal="center", vertical="center")
                fixed_count += 1
        
        if fixed_count > 0:
            print(f"\n✓ 共修正 {fixed_count} 場比賽的分數位置")
            
            # 儲存檔案
            print("\n正在儲存檔案...")
            wb.save(excel_file)
            print(f"✓ 修正完成並已儲存")
        else:
            print("\n⚠ 未找到需要修正的資料")
        
        wb.close()
        return True
        
    except KeyError as e:
        error_msg = str(e)
        if "[Content_Types].xml" in error_msg or "archive" in error_msg.lower():
            print(f"✗ Excel 檔案損壞或格式異常：{excel_file}")
            print(f"  錯誤訊息：{error_msg}")
            print(f"\n  建議：請從備份還原檔案")
        else:
            print(f"✗ 讀取 Excel 檔案時發生錯誤：{error_msg}")
        return False
    except PermissionError:
        print(f"✗ 無法儲存檔案：檔案可能正在被其他程式使用")
        print(f"  請關閉 Excel 後再試")
        return False
    except Exception as e:
        error_msg = str(e)
        print(f"✗ 執行時發生錯誤：{error_msg}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主函數"""
    print("=" * 60)
    print("修正「每日數據」分頁中 20260220~20260223 的分數位置")
    print("=" * 60)
    
    excel_file = NBA_XLSX_FILE
    
    # 需要修正的日期列表
    date_list = ['20260220', '20260221', '20260222', '20260223']
    
    print(f"\n將修正以下日期的資料：")
    for date in date_list:
        print(f"  - {date}")
    
    if fix_scores_for_dates(excel_file, date_list):
        print("\n" + "=" * 60)
        print("✓ 執行完成！")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("✗ 執行失敗！")
        print("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    main()
