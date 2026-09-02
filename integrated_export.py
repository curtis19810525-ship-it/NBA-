# -*- coding: utf-8 -*-
"""
整合匯出：紀錄 ＋ 盤口觀察 → 每日兩份 NotebookLM 資料來源。

定案：
  - {關注日}賽前分析.txt ＝ 紀錄（賽前）＋ 盤口觀察（賽前），剪貼簿＝此文
  - {賽果日}賽後結果.txt ＝ 紀錄（賽後）＋ 盤口結果（賽後）
  - 一場兩段：【模型｜紀錄】＋【盤口走勢｜觀察】；以讓分球隊對齊
  - 不再產出舊四份獨立 txt
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import load_workbook

from config import MLB_XLSX_FILE, OBSERVATION_XLSX_FILE
from fill_observation import SHEET_RECORD
from notebooklm_export import (
    EMPTY,
    EXPORT_DIR,
    _fmt_empty,
    _fmt_flow,
    _fmt_number,
    copy_text_to_clipboard,
    export_filename,
    read_record_export_games,
)
from observation_export import (
    _fmt_cell,
    _fmt_flow_export,
    read_observation_rows,
    summarize_trend,
)

KIND_PREGAME_ANALYSIS = "賽前分析"
KIND_POSTGAME_RESULTS = "賽後結果"


def _team_key(name: str) -> str:
    return str(name or "").strip()


def merge_games_by_team(
    record_games: list[dict],
    obs_rows: list[dict],
) -> list[dict]:
    """以讓分球隊合併；順序：紀錄場次序優先，再補盤口獨有場次。"""
    record_by_team = {_team_key(g.get("fav_team")): g for g in record_games}
    obs_by_team = {_team_key(r.get("team")): r for r in obs_rows}

    order: list[str] = []
    seen: set[str] = set()
    for g in record_games:
        key = _team_key(g.get("fav_team"))
        if key and key not in seen:
            order.append(key)
            seen.add(key)
    for r in obs_rows:
        key = _team_key(r.get("team"))
        if key and key not in seen:
            order.append(key)
            seen.add(key)

    merged: list[dict] = []
    for i, team in enumerate(order, 1):
        merged.append(
            {
                "game_label": f"第{i}場",
                "team": team,
                "record": record_by_team.get(team),
                "observation": obs_by_team.get(team),
            }
        )
    return merged


def _format_record_block(rec: dict | None, *, postgame: bool) -> list[str]:
    if not rec:
        return [
            "【模型｜紀錄】",
            "（此場紀錄查無資料）",
        ]
    lines = [
        "【模型｜紀錄】",
        f"時間：{_fmt_empty(rec.get('time'))}",
        f"讓分球隊：{_fmt_empty(rec.get('fav_team'))}",
        f"主客：{_fmt_empty(rec.get('side'))}",
        f"合理讓分：{_fmt_number(rec.get('fair'))}",
        f"讓分：{_fmt_empty(rec.get('handicap'))}",
        f"金流：{_fmt_flow(rec.get('flow'))}",
        f"合理性：{_fmt_empty(rec.get('quant'))}",
    ]
    if postgame:
        lines.extend(
            [
                f"讓分方得分：{_fmt_number(rec.get('score_fav'))}",
                f"受讓方得分：{_fmt_number(rec.get('score_dog'))}",
                f"結果：{_fmt_empty(rec.get('result'))}",
            ]
        )
    return lines


def _format_observation_block(obs: dict | None) -> list[str]:
    if not obs:
        return [
            "【盤口走勢｜觀察】",
            "（此場盤口觀察查無資料）",
        ]
    return [
        "【盤口走勢｜觀察】",
        f"讓分(頭)：{_fmt_cell(obs.get('sp_head'))}",
        f"讓分(尾)：{_fmt_cell(obs.get('sp_tail'))}",
        f"讓分走勢：{summarize_trend(obs.get('sp_head'), obs.get('sp_tail'), market='spread')}",
        f"讓分獨贏(頭/尾)：{_fmt_cell(obs.get('favml_head'))} / {_fmt_cell(obs.get('favml_tail'))}",
        f"讓分獨贏走勢：{summarize_trend(obs.get('favml_head'), obs.get('favml_tail'))}",
        f"受讓獨贏(頭/尾)：{_fmt_cell(obs.get('recml_head'))} / {_fmt_cell(obs.get('recml_tail'))}",
        f"受讓獨贏走勢：{summarize_trend(obs.get('recml_head'), obs.get('recml_tail'))}",
        f"讓2分(頭/尾)：{_fmt_cell(obs.get('fav2_head'))} / {_fmt_cell(obs.get('fav2_tail'))}",
        f"讓2分走勢：{summarize_trend(obs.get('fav2_head'), obs.get('fav2_tail'))}",
        f"受讓2分(頭/尾)：{_fmt_cell(obs.get('rec2_head'))} / {_fmt_cell(obs.get('rec2_tail'))}",
        f"受讓2分走勢：{summarize_trend(obs.get('rec2_head'), obs.get('rec2_tail'))}",
        f"讓分結果：{_fmt_cell(obs.get('sp_result'))}",
        f"獨贏結果：{_fmt_cell(obs.get('ml_result'))}",
        f"2分結果：{_fmt_cell(obs.get('two_result'))}",
    ]


def format_integrated_text(
    focus_date: str,
    merged_games: list[dict],
    *,
    kind: str,
    exported_at: datetime | None = None,
) -> str:
    exported_at = exported_at or datetime.now()
    postgame = kind == KIND_POSTGAME_RESULTS
    title = "賽後整合結果" if postgame else "賽前整合分析"
    lines = [
        f"【MLB {title}｜資料來源】",
        f"類型：{kind}",
        f"日期：{focus_date}",
        "來源：MLB26-27數據.xlsx → 紀錄 ＋ 盤口觀察.xlsx → 工作表1",
        f"匯出時間：{exported_at:%Y-%m-%d %H:%M:%S}",
        "",
        "說明：",
        "- 每場含【模型｜紀錄】與【盤口走勢｜觀察】，以讓分球隊對齊",
        "- 頭盤＝玖九第1筆快照；尾盤＝開賽前最後快照",
        "- H/M/R：讓分結果／獨贏結果／2分結果",
        "- 空值：_",
        "",
        "========================================",
    ]
    if not merged_games:
        lines.append("（此日期查無可合併場次）")
    else:
        for game in merged_games:
            lines.append(game["game_label"])
            lines.append(f"讓分球隊：{_fmt_empty(game.get('team'))}")
            lines.extend(_format_record_block(game.get("record"), postgame=postgame))
            lines.extend(_format_observation_block(game.get("observation")))
            lines.append("")
    lines.append("========================================")
    lines.append(f"合計場次：{len(merged_games)}")
    lines.append("")
    return "\n".join(lines)


def _load_record_games(
    focus_date: str,
    *,
    mlb_xlsx: str | None = None,
) -> list[dict]:
    mlb_xlsx = mlb_xlsx or MLB_XLSX_FILE
    if not os.path.exists(mlb_xlsx):
        raise SystemExit(f"找不到檔案：{mlb_xlsx}")
    wb = load_workbook(mlb_xlsx, data_only=True)
    try:
        if SHEET_RECORD not in wb.sheetnames:
            raise SystemExit("MLB26-27 找不到「紀錄」分頁")
        return read_record_export_games(wb[SHEET_RECORD], focus_date)
    finally:
        wb.close()


def _load_observation_rows(
    focus_date: str,
    *,
    observation_xlsx: str | None = None,
) -> list[dict]:
    observation_xlsx = observation_xlsx or OBSERVATION_XLSX_FILE
    if not os.path.exists(observation_xlsx):
        raise SystemExit(f"找不到檔案：{observation_xlsx}")
    wb = load_workbook(observation_xlsx, data_only=True)
    try:
        return read_observation_rows(wb.worksheets[0], focus_date)
    finally:
        wb.close()


def export_integrated_day(
    focus_date: str,
    *,
    kind: str,
    mlb_xlsx: str | None = None,
    observation_xlsx: str | None = None,
    export_dir: str | None = None,
    copy_to_clipboard: bool = False,
) -> tuple[str, int]:
    """匯出單日整合檔。回傳 (檔案路徑, 場次數)。"""
    export_dir = export_dir or EXPORT_DIR
    os.makedirs(export_dir, exist_ok=True)

    record_games = _load_record_games(focus_date, mlb_xlsx=mlb_xlsx)
    obs_rows = _load_observation_rows(focus_date, observation_xlsx=observation_xlsx)
    merged = merge_games_by_team(record_games, obs_rows)

    text = format_integrated_text(focus_date, merged, kind=kind)
    out_path = os.path.join(export_dir, export_filename(focus_date, kind))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)

    if copy_to_clipboard:
        copy_text_to_clipboard(text)

    return out_path, len(merged)


def export_integrated_for_pipeline(
    results_day: str,
    focus_day: str,
    *,
    mlb_xlsx: str | None = None,
    observation_xlsx: str | None = None,
) -> dict:
    """賽果日→賽後結果；關注日→賽前分析（剪貼簿）。"""
    results_path, results_n = export_integrated_day(
        results_day,
        kind=KIND_POSTGAME_RESULTS,
        mlb_xlsx=mlb_xlsx,
        observation_xlsx=observation_xlsx,
        copy_to_clipboard=False,
    )
    pregame_path, pregame_n = export_integrated_day(
        focus_day,
        kind=KIND_PREGAME_ANALYSIS,
        mlb_xlsx=mlb_xlsx,
        observation_xlsx=observation_xlsx,
        copy_to_clipboard=True,
    )
    export_dir = os.path.dirname(os.path.abspath(pregame_path))
    return {
        "export_dir": export_dir,
        "results_path": results_path,
        "results_games": results_n,
        "pregame_path": pregame_path,
        "pregame_games": pregame_n,
        # 相容舊鍵
        "focus_path": pregame_path,
        "focus_games": pregame_n,
    }
