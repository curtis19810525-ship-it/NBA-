# -*- coding: utf-8 -*-
"""
姓名對照表增量更新（不覆蓋既有資料）

規則：
1) A(隊名)、B(球員狀況全名) 完全不改動
2) 只在 C/D 為空時補值
3) C 已有值時，不覆蓋
"""

import os
import re
import sys
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook

try:
    from config import NBA_STATS_FILE
except ImportError:
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"


def _norm(s):
    t = unicodedata.normalize("NFKD", str(s or "").strip())
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    t = re.sub(r"[^a-z0-9'\-\. ]+", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _split_name(s):
    p = [x for x in _norm(s).split(" ") if x]
    return (p[0], p[-1]) if p else ("", "")


def _short_matches_full(short_name, full_name):
    ns = _norm(short_name)
    first, last = _split_name(full_name)
    if not last:
        return False

    # 例：J. Aranda -> Jonathan Aranda
    m = re.match(r"^([a-z])\.\s*(.+)$", ns)
    if m:
        initial = m.group(1)
        short_last = m.group(2)
        return last == short_last and first.startswith(initial)

    # 若先發名單本身就是全名，直接允許
    sf, sl = _split_name(short_name)
    if sf and sl:
        return last == sl and (first == sf or first.startswith(sf) or sf.startswith(first))
    return False


def _get_header_map(ws):
    return {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}


def _collect_lineup_names_by_team(ws_lineups):
    header = _get_header_map(ws_lineups)
    col_away = header.get("客隊", 3)
    col_home = header.get("主隊", 4)

    away_cols = []
    home_cols = []
    for c in range(1, ws_lineups.max_column + 1):
        h = str(ws_lineups.cell(1, c).value or "").strip()
        if "客隊先發" in h and "棒" in h:
            away_cols.append(c)
        if "主隊先發" in h and "棒" in h:
            home_cols.append(c)

    names_by_team = defaultdict(set)
    for r in range(2, ws_lineups.max_row + 1):
        away_team = str(ws_lineups.cell(r, col_away).value or "").strip()
        home_team = str(ws_lineups.cell(r, col_home).value or "").strip()

        for c in away_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if away_team and n:
                names_by_team[away_team].add(n)
        for c in home_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if home_team and n:
                names_by_team[home_team].add(n)
    return names_by_team


def update_mapping_sheet(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 檔案不存在：{path}")

    wb = load_workbook(path)
    if "姓名對照表" not in wb.sheetnames:
        wb.create_sheet("姓名對照表")
    if "先發名單" not in wb.sheetnames:
        raise KeyError("找不到分頁：先發名單")

    ws_map = wb["姓名對照表"]
    ws_line = wb["先發名單"]

    # 保留使用者現有結構，僅補齊標頭空白
    if not ws_map["A1"].value:
        ws_map["A1"] = "隊名"
    if not ws_map["B1"].value:
        ws_map["B1"] = "球員狀況"
    if not ws_map["C1"].value:
        ws_map["C1"] = "先發名單"
    if not ws_map["D1"].value:
        ws_map["D1"] = "核對"

    lineup_by_team = _collect_lineup_names_by_team(ws_line)

    processed = 0
    filled = 0
    unresolved = 0
    multi = 0
    preserved = 0

    for r in range(2, ws_map.max_row + 1):
        team = str(ws_map.cell(r, 1).value or "").strip()
        full = str(ws_map.cell(r, 2).value or "").strip()
        cur_c = str(ws_map.cell(r, 3).value or "").strip()
        cur_d = str(ws_map.cell(r, 4).value or "").strip()

        if not team and not full:
            continue

        processed += 1

        # 不覆蓋既有 C/D
        if cur_c:
            preserved += 1
            continue

        candidates = []
        for short in lineup_by_team.get(team, set()):
            if _short_matches_full(short, full):
                candidates.append(short)
        candidates = sorted(set(candidates))

        if len(candidates) == 1:
            ws_map.cell(r, 3).value = candidates[0]
            if not cur_d:
                ws_map.cell(r, 4).value = "已匹配"
            filled += 1
        elif len(candidates) > 1:
            # C 留空避免誤判，只寫核對狀態
            if not cur_d:
                ws_map.cell(r, 4).value = "多候選"
            multi += 1
        else:
            if not cur_d:
                ws_map.cell(r, 4).value = "未匹配"
            unresolved += 1

    wb.save(path)
    wb.close()
    return processed, filled, multi, unresolved, preserved


def main():
    print("=" * 60)
    print("姓名對照表增量更新（只補 C/D，不改 A/B）")
    print("=" * 60)
    print(f"Excel: {NBA_STATS_FILE}")

    try:
        processed, filled, multi, unresolved, preserved = update_mapping_sheet(NBA_STATS_FILE)
    except Exception as e:
        print(f"錯誤：{e}")
        return 1

    print("-" * 60)
    print(f"處理列數: {processed}")
    print(f"已保留既有 C 欄: {preserved}")
    print(f"本次新填 C 欄: {filled}")
    print(f"多候選: {multi}")
    print(f"未匹配: {unresolved}")
    print("-" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
