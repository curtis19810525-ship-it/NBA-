# -*- coding: utf-8 -*-
"""
姓名對照表增量更新（Step 8 only）

兩個對照區塊（同一張「姓名對照表」分頁）：
  A 區塊：A=隊名、B=球員狀況（球員狀態全名）、C=先發名單、D=核對
          → 比對來源：球員狀態 vs 先發名單（客隊／主隊先發1～9棒）
  F 區塊：F=隊名、G=各隊投手、H=先發名單、I=核對
          → 比對來源：各隊投手 vs 先發名單（客隊先發投手／主隊先發投手）

執行範圍：
- 僅 Step 8 會讀取／寫入「姓名對照表」。

兩區塊的列「不強制同列對齊」（互相獨立）：
- A 區塊用 A 欄找該隊最末列；F 區塊用 F 欄找該隊最末列。
- 插入時 ws.insert_rows 會推動所有欄一起下移，這是接受的後果。

各區塊規則（皆相同）：
1) 拼法以來源（球員狀態／各隊投手）為準；命中時必要時改寫全名欄。
2) 球員不在來源 → 全名欄塗黃底，核對欄寫「已離隊」（覆寫原值）。
3) 核對欄已是「已離隊」整列跳過（不自動還原；含球員回到來源亦同）。
4) 在隊球員：C(H) 已有值整列跳過；C(H) 空白才用先發名單比對補 C/D（H/I）。
5) 來源有、姓名對照表沒有：於該隊區塊最末新增一列；該隊整段不存在則附在 sheet 末尾。
   F 區塊首次填入（整欄空白）：從 F2 起依來源順序連續寫入，不使用 insert_rows。
"""

import os
import re
import sys
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from config import NBA_STATS_FILE
except ImportError:
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _norm(s):
    t = unicodedata.normalize("NFKD", str(s or "").strip())
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    t = re.sub(r"[^a-z0-9'\-\. ]+", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _split_name(s):
    p = [x for x in _norm(s).split(" ") if x]
    return (p[0], p[-1]) if p else ("", "")


def _short_matches_full(short_name, full_name):
    ns = _norm(short_name)
    first, last = _split_name(full_name)
    if not last:
        return False

    m = re.match(r"^([a-z])\.\s*(.+)$", ns)
    if m:
        initial = m.group(1)
        short_last = m.group(2)
        return last == short_last and first.startswith(initial)

    sf, sl = _split_name(short_name)
    if sf and sl:
        return last == sl and (first == sf or first.startswith(sf) or sf.startswith(first))
    return False


def _get_header_map(ws):
    return {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}


def _collect_lineup_batters_by_team(ws_lineups):
    """先發名單『客／主隊先發1～9棒』縮寫，依隊伍歸戶。"""
    header = _get_header_map(ws_lineups)
    col_away = header.get("客隊", 3)
    col_home = header.get("主隊", 4)

    away_cols = []
    home_cols = []
    for c in range(1, ws_lineups.max_column + 1):
        h = str(ws_lineups.cell(1, c).value or "").strip()
        if "客隊先發" in h and "棒" in h:
            away_cols.append(c)
        if "主隊先發" in h and "棒" in h:
            home_cols.append(c)

    names_by_team = defaultdict(set)
    for r in range(2, ws_lineups.max_row + 1):
        away_team = str(ws_lineups.cell(r, col_away).value or "").strip()
        home_team = str(ws_lineups.cell(r, col_home).value or "").strip()

        for c in away_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if away_team and n:
                names_by_team[away_team].add(n)
        for c in home_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if home_team and n:
                names_by_team[home_team].add(n)
    return names_by_team


def _collect_lineup_pitchers_by_team(ws_lineups):
    """先發名單『客隊先發投手／主隊先發投手』縮寫，依隊伍歸戶。"""
    header = _get_header_map(ws_lineups)
    col_away = header.get("客隊", 3)
    col_home = header.get("主隊", 4)
    col_away_p = header.get("客隊先發投手")
    col_home_p = header.get("主隊先發投手")

    names_by_team = defaultdict(set)
    if not col_away_p and not col_home_p:
        return names_by_team

    for r in range(2, ws_lineups.max_row + 1):
        away_team = str(ws_lineups.cell(r, col_away).value or "").strip()
        home_team = str(ws_lineups.cell(r, col_home).value or "").strip()

        if col_away_p:
            n = str(ws_lineups.cell(r, col_away_p).value or "").strip()
            if away_team and n:
                names_by_team[away_team].add(n)
        if col_home_p:
            n = str(ws_lineups.cell(r, col_home_p).value or "").strip()
            if home_team and n:
                names_by_team[home_team].add(n)
    return names_by_team


def _load_roster_from_sheet(wb, sheet_name):
    """讀指定分頁的 (A=隊名, B=全名)，回傳 {team: {norm: full_spelling}}（同隊同正規化以首見為準）。"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rosters = {}
    for r in range(2, ws.max_row + 1):
        team = str(ws.cell(r, 1).value or "").strip()
        full = str(ws.cell(r, 2).value or "").strip()
        if not team or not full:
            continue
        rosters.setdefault(team, {}).setdefault(_norm(full), full)
    return rosters


def _try_fill_match(ws_map, r, team, full, lineup_short_by_team, c_col, d_col):
    """在(r,c_col/d_col) 用先發名單縮寫補 C/D（C 已有值則保留）。
    回傳 (filled, multi, unresolved, preserved)。"""
    cur_c = str(ws_map.cell(r, c_col).value or "").strip()
    cur_d = str(ws_map.cell(r, d_col).value or "").strip()
    if cur_c:
        return (0, 0, 0, 1)

    candidates = sorted(
        {s for s in lineup_short_by_team.get(team, set()) if _short_matches_full(s, full)}
    )
    if len(candidates) == 1:
        ws_map.cell(r, c_col).value = candidates[0]
        if not cur_d:
            ws_map.cell(r, d_col).value = "已匹配"
        return (1, 0, 0, 0)
    if len(candidates) > 1:
        if not cur_d:
            ws_map.cell(r, d_col).value = "多候選"
        return (0, 1, 0, 0)
    if not cur_d:
        ws_map.cell(r, d_col).value = "未匹配"
    return (0, 0, 1, 0)


def _process_block(
    ws_map,
    roster_by_team,
    lineup_short_by_team,
    *,
    team_col,
    full_col,
    c_col,
    d_col,
    allow_first_run_bulk=False,
):
    """通用區塊處理（A 區塊 / F 區塊共用）。"""
    stats = {
        "processed": 0,
        "preserved": 0,
        "filled": 0,
        "multi": 0,
        "unresolved": 0,
        "departed_new": 0,
        "departed_skipped": 0,
        "spelling_updated": 0,
        "new_added": 0,
    }

    if allow_first_run_bulk:
        has_data = False
        for r in range(2, ws_map.max_row + 1):
            if str(ws_map.cell(r, team_col).value or "").strip():
                has_data = True
                break
        if not has_data:
            next_row = 2
            for team, lookup in roster_by_team.items():
                for full in lookup.values():
                    ws_map.cell(next_row, team_col).value = team
                    ws_map.cell(next_row, full_col).value = full
                    ws_map.cell(next_row, c_col).value = None
                    ws_map.cell(next_row, d_col).value = None
                    f, m, u, p = _try_fill_match(
                        ws_map, next_row, team, full, lineup_short_by_team, c_col, d_col
                    )
                    stats["filled"] += f
                    stats["multi"] += m
                    stats["unresolved"] += u
                    stats["preserved"] += p
                    stats["new_added"] += 1
                    stats["processed"] += 1
                    next_row += 1
            return stats

    matched = defaultdict(set)
    last_row_per_team = {}

    for r in range(2, ws_map.max_row + 1):
        team = str(ws_map.cell(r, team_col).value or "").strip()
        full = str(ws_map.cell(r, full_col).value or "").strip()
        cur_d = str(ws_map.cell(r, d_col).value or "").strip()

        if team:
            last_row_per_team[team] = max(last_row_per_team.get(team, 0), r)
        if not team and not full:
            continue
        stats["processed"] += 1

        if cur_d == "已離隊":
            stats["departed_skipped"] += 1
            norm_full = _norm(full)
            if team in roster_by_team and norm_full in roster_by_team[team]:
                matched[team].add(norm_full)
            continue

        team_lookup = roster_by_team.get(team, {})
        norm_full = _norm(full)

        if norm_full in team_lookup:
            standard_full = team_lookup[norm_full]
            if standard_full != full:
                ws_map.cell(r, full_col).value = standard_full
                full = standard_full
                stats["spelling_updated"] += 1
            matched[team].add(norm_full)
            f, m, u, p = _try_fill_match(
                ws_map, r, team, full, lineup_short_by_team, c_col, d_col
            )
            stats["filled"] += f
            stats["multi"] += m
            stats["unresolved"] += u
            stats["preserved"] += p
        else:
            ws_map.cell(r, d_col).value = "已離隊"
            ws_map.cell(r, full_col).fill = YELLOW_FILL
            stats["departed_new"] += 1

    insert_targets = []
    append_targets = []
    for team, lookup in roster_by_team.items():
        new_fulls = [full for norm, full in lookup.items() if norm not in matched[team]]
        if not new_fulls:
            continue
        if team in last_row_per_team:
            insert_targets.append((last_row_per_team[team] + 1, team, new_fulls))
        else:
            append_targets.append((team, new_fulls))

    insert_targets.sort(key=lambda x: x[0], reverse=True)

    def _write_block(start_row, team, fulls):
        for i, full in enumerate(fulls):
            r = start_row + i
            ws_map.cell(r, team_col).value = team
            ws_map.cell(r, full_col).value = full
            ws_map.cell(r, c_col).value = None
            ws_map.cell(r, d_col).value = None
            f, m, u, p = _try_fill_match(
                ws_map, r, team, full, lineup_short_by_team, c_col, d_col
            )
            stats["filled"] += f
            stats["multi"] += m
            stats["unresolved"] += u
            stats["preserved"] += p
            stats["new_added"] += 1
            stats["processed"] += 1

    for ins_row, team, fulls in insert_targets:
        ws_map.insert_rows(ins_row, amount=len(fulls))
        _write_block(ins_row, team, fulls)

    for team, fulls in append_targets:
        base = ws_map.max_row + 1
        _write_block(base, team, fulls)

    return stats


def update_mapping_sheet(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 檔案不存在：{path}")

    wb = load_workbook(path)
    if "姓名對照表" not in wb.sheetnames:
        wb.create_sheet("姓名對照表")
    if "先發名單" not in wb.sheetnames:
        raise KeyError("找不到分頁：先發名單")

    ws_map = wb["姓名對照表"]
    ws_line = wb["先發名單"]

    if not ws_map["A1"].value:
        ws_map["A1"] = "隊名"
    if not ws_map["B1"].value:
        ws_map["B1"] = "球員狀況"
    if not ws_map["C1"].value:
        ws_map["C1"] = "先發名單"
    if not ws_map["D1"].value:
        ws_map["D1"] = "核對"
    if not ws_map["F1"].value:
        ws_map["F1"] = "隊名"
    if not ws_map["G1"].value:
        ws_map["G1"] = "各隊投手"
    if not ws_map["H1"].value:
        ws_map["H1"] = "先發名單"
    if not ws_map["I1"].value:
        ws_map["I1"] = "核對"

    batters = _load_roster_from_sheet(wb, "球員狀態")
    pitchers = _load_roster_from_sheet(wb, "各隊投手")
    lineup_batters = _collect_lineup_batters_by_team(ws_line)
    lineup_pitchers = _collect_lineup_pitchers_by_team(ws_line)

    a_stats = _process_block(
        ws_map, batters, lineup_batters,
        team_col=1, full_col=2, c_col=3, d_col=4,
        allow_first_run_bulk=False,
    )

    f_stats = _process_block(
        ws_map, pitchers, lineup_pitchers,
        team_col=6, full_col=7, c_col=8, d_col=9,
        allow_first_run_bulk=True,
    )

    wb.save(path)
    wb.close()
    return a_stats, f_stats


def _print_stats(label, s):
    print(f"[{label}]")
    print(f"  處理列數              : {s['processed']}")
    print(f"  新增列                : {s['new_added']}")
    print(f"  本次新標『已離隊』    : {s['departed_new']}")
    print(f"  原已『已離隊』整列跳過: {s['departed_skipped']}")
    print(f"  全名拼法已校正        : {s['spelling_updated']}")
    print(f"  已保留既有 C/H 欄     : {s['preserved']}")
    print(f"  本次新填 C/H 欄       : {s['filled']}")
    print(f"  多候選                : {s['multi']}")
    print(f"  未匹配                : {s['unresolved']}")


def main():
    print("=" * 60)
    print("姓名對照表增量更新（Step 8 only）")
    print("=" * 60)
    print(f"Excel: {NBA_STATS_FILE}")
    print("規則：")
    print("  - A 區塊：球員狀態 ↔ 先發名單（1～9棒）")
    print("  - F 區塊：各隊投手 ↔ 先發名單（先發投手欄）")
    print("  - 兩區塊獨立處理；同列不強制對齊")
    print("  - 拼法以來源為準；不在來源 → B/G 黃底 + D/I 寫『已離隊』")
    print("  - 已標『已離隊』整列跳過（不自動還原）")
    print("  - C/H 已有值整列跳過；空白才用先發名單補")

    try:
        a_stats, f_stats = update_mapping_sheet(NBA_STATS_FILE)
    except Exception as e:
        print(f"錯誤：{e}")
        return 1

    print("-" * 60)
    _print_stats("A 區塊（球員狀態 vs 先發名單1~9棒）", a_stats)
    print()
    _print_stats("F 區塊（各隊投手 vs 先發名單先發投手）", f_stats)
    print("-" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
