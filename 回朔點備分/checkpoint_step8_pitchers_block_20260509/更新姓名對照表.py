# -*- coding: utf-8 -*-
"""
姓名對照表增量更新（Step 8 only）

執行範圍與保護原則：
- Step 4（球員狀態）與 Step 8（姓名對照）完全分開。
- 只有 Step 8 會讀取／寫入「姓名對照表」。

更新規則（依使用者定稿）：
1) 從「球員狀態」A=隊名、B=全名 取得每隊現役名單。
2) 既有列以「正規化姓名」比對：
   - 若該球員仍在球員狀態 → 拼法以球員狀態為準（必要時改寫 B 欄）；
     C 已有值整列跳過；C 空白才用「先發名單」補 C/D。
   - 若該球員已不在球員狀態 → 在 B 欄塗黃色底色，D 寫「已離隊」（覆寫原 D）。
3) 不自動「還原」：D 已是「已離隊」的列整列跳過，
   含「球員又回到球員狀態」的情況亦不還原（保留黃底與「已離隊」由人工處理）。
4) 球員狀態存在、但姓名對照表沒有的球員，於該隊區塊最末新增一列（A、B 寫入；C/D 自動嘗試補）。
   若該隊在姓名對照表完全不存在，則整段附加在表尾。
"""

import os
import re
import sys
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from config import NBA_STATS_FILE
except ImportError:
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _norm(s):
    t = unicodedata.normalize("NFKD", str(s or "").strip())
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    t = re.sub(r"[^a-z0-9'\-\. ]+", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _split_name(s):
    p = [x for x in _norm(s).split(" ") if x]
    return (p[0], p[-1]) if p else ("", "")


def _short_matches_full(short_name, full_name):
    ns = _norm(short_name)
    first, last = _split_name(full_name)
    if not last:
        return False

    m = re.match(r"^([a-z])\.\s*(.+)$", ns)
    if m:
        initial = m.group(1)
        short_last = m.group(2)
        return last == short_last and first.startswith(initial)

    sf, sl = _split_name(short_name)
    if sf and sl:
        return last == sl and (first == sf or first.startswith(sf) or sf.startswith(first))
    return False


def _get_header_map(ws):
    return {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}


def _collect_lineup_names_by_team(ws_lineups):
    header = _get_header_map(ws_lineups)
    col_away = header.get("客隊", 3)
    col_home = header.get("主隊", 4)

    away_cols = []
    home_cols = []
    for c in range(1, ws_lineups.max_column + 1):
        h = str(ws_lineups.cell(1, c).value or "").strip()
        if "客隊先發" in h and "棒" in h:
            away_cols.append(c)
        if "主隊先發" in h and "棒" in h:
            home_cols.append(c)

    names_by_team = defaultdict(set)
    for r in range(2, ws_lineups.max_row + 1):
        away_team = str(ws_lineups.cell(r, col_away).value or "").strip()
        home_team = str(ws_lineups.cell(r, col_home).value or "").strip()

        for c in away_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if away_team and n:
                names_by_team[away_team].add(n)
        for c in home_cols:
            n = str(ws_lineups.cell(r, c).value or "").strip()
            if home_team and n:
                names_by_team[home_team].add(n)
    return names_by_team


def _load_roster_by_team(wb):
    """從「球員狀態」讀 (A=隊名, B=全名)，回傳 {team: {norm: full_spelling}}（同隊同正規化以首見為準）。"""
    if "球員狀態" not in wb.sheetnames:
        return {}
    ws = wb["球員狀態"]
    rosters = {}
    for r in range(2, ws.max_row + 1):
        team = str(ws.cell(r, 1).value or "").strip()
        full = str(ws.cell(r, 2).value or "").strip()
        if not team or not full:
            continue
        rosters.setdefault(team, {}).setdefault(_norm(full), full)
    return rosters


def _try_fill_c_d(ws_map, r, team, full, lineup_by_team):
    """C 已有值則保留；否則用先發名單補 C/D。回傳 (filled, multi, unresolved, preserved)。"""
    cur_c = str(ws_map.cell(r, 3).value or "").strip()
    cur_d = str(ws_map.cell(r, 4).value or "").strip()
    if cur_c:
        return (0, 0, 0, 1)

    candidates = sorted(
        {s for s in lineup_by_team.get(team, set()) if _short_matches_full(s, full)}
    )
    if len(candidates) == 1:
        ws_map.cell(r, 3).value = candidates[0]
        if not cur_d:
            ws_map.cell(r, 4).value = "已匹配"
        return (1, 0, 0, 0)
    if len(candidates) > 1:
        if not cur_d:
            ws_map.cell(r, 4).value = "多候選"
        return (0, 1, 0, 0)
    if not cur_d:
        ws_map.cell(r, 4).value = "未匹配"
    return (0, 0, 1, 0)


def update_mapping_sheet(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 檔案不存在：{path}")

    wb = load_workbook(path)
    if "姓名對照表" not in wb.sheetnames:
        wb.create_sheet("姓名對照表")
    if "先發名單" not in wb.sheetnames:
        raise KeyError("找不到分頁：先發名單")

    ws_map = wb["姓名對照表"]
    ws_line = wb["先發名單"]

    if not ws_map["A1"].value:
        ws_map["A1"] = "隊名"
    if not ws_map["B1"].value:
        ws_map["B1"] = "球員狀況"
    if not ws_map["C1"].value:
        ws_map["C1"] = "先發名單"
    if not ws_map["D1"].value:
        ws_map["D1"] = "核對"

    rosters = _load_roster_by_team(wb)
    lineup_by_team = _collect_lineup_names_by_team(ws_line)

    matched = defaultdict(set)
    last_row_per_team = {}

    processed = 0
    preserved = 0
    filled = 0
    multi = 0
    unresolved = 0
    departed = 0
    already_departed_skipped = 0
    spelling_updated = 0

    for r in range(2, ws_map.max_row + 1):
        team = str(ws_map.cell(r, 1).value or "").strip()
        full = str(ws_map.cell(r, 2).value or "").strip()
        cur_d = str(ws_map.cell(r, 4).value or "").strip()

        if team:
            last_row_per_team[team] = max(last_row_per_team.get(team, 0), r)
        if not team and not full:
            continue
        processed += 1

        if cur_d == "已離隊":
            already_departed_skipped += 1
            norm_full = _norm(full)
            if team in rosters and norm_full in rosters[team]:
                matched[team].add(norm_full)
            continue

        team_lookup = rosters.get(team, {})
        norm_full = _norm(full)

        if norm_full in team_lookup:
            standard_full = team_lookup[norm_full]
            if standard_full != full:
                ws_map.cell(r, 2).value = standard_full
                full = standard_full
                spelling_updated += 1
            matched[team].add(norm_full)
            f, m, u, p = _try_fill_c_d(ws_map, r, team, full, lineup_by_team)
            filled += f
            multi += m
            unresolved += u
            preserved += p
        else:
            ws_map.cell(r, 4).value = "已離隊"
            ws_map.cell(r, 2).fill = YELLOW_FILL
            departed += 1

    insert_plans = []
    for team, lookup in rosters.items():
        new_fulls = [full for norm, full in lookup.items() if norm not in matched[team]]
        if not new_fulls:
            continue
        if team in last_row_per_team:
            insert_plans.append((last_row_per_team[team] + 1, team, new_fulls, "insert"))
        else:
            insert_plans.append((ws_map.max_row + 1, team, new_fulls, "append"))

    insert_plans.sort(key=lambda x: x[0], reverse=True)

    new_added = 0
    for ins_row, team, fulls, mode in insert_plans:
        n = len(fulls)
        if mode == "insert":
            ws_map.insert_rows(ins_row, amount=n)
        for i, full in enumerate(fulls):
            r = ins_row + i
            ws_map.cell(r, 1).value = team
            ws_map.cell(r, 2).value = full
            ws_map.cell(r, 3).value = None
            ws_map.cell(r, 4).value = None
            f, m, u, p = _try_fill_c_d(ws_map, r, team, full, lineup_by_team)
            filled += f
            multi += m
            unresolved += u
            preserved += p
            new_added += 1
            processed += 1

    wb.save(path)
    wb.close()
    return {
        "processed": processed,
        "preserved": preserved,
        "filled": filled,
        "multi": multi,
        "unresolved": unresolved,
        "departed_new": departed,
        "departed_skipped": already_departed_skipped,
        "spelling_updated": spelling_updated,
        "new_added": new_added,
    }


def main():
    print("=" * 60)
    print("姓名對照表增量更新（Step 8 only）")
    print("=" * 60)
    print(f"Excel: {NBA_STATS_FILE}")
    print("規則：")
    print("  - 拼法以「球員狀態」為準（必要時改寫 B 欄）")
    print("  - 球員狀態有、姓名對照表沒有 → 於該隊區塊最末新增一列")
    print("  - 球員不在球員狀態 → B 欄黃色底色 + D 寫「已離隊」（覆寫）")
    print("  - 已標「已離隊」的列整列跳過（不自動還原）")
    print("  - 在隊球員：C 有值整列跳過；C 空才用先發名單補 C/D")

    try:
        stats = update_mapping_sheet(NBA_STATS_FILE)
    except Exception as e:
        print(f"錯誤：{e}")
        return 1

    print("-" * 60)
    print(f"處理列數              : {stats['processed']}")
    print(f"新增列（從球員狀態）  : {stats['new_added']}")
    print(f"本次新標『已離隊』    : {stats['departed_new']}")
    print(f"原已『已離隊』整列跳過: {stats['departed_skipped']}")
    print(f"B 拼法已校正          : {stats['spelling_updated']}")
    print(f"已保留既有 C 欄       : {stats['preserved']}")
    print(f"本次新填 C 欄         : {stats['filled']}")
    print(f"多候選                : {stats['multi']}")
    print(f"未匹配                : {stats['unresolved']}")
    print("-" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
