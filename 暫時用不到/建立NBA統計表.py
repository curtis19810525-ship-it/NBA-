"""
建立 NBA 25-26 賽季統計表
按照用戶需求逐步建立表格結構
包含總表 + 30支球隊各自的分頁
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# 引入設定檔
try:
    from config import NBA_STATS_FILE
except ImportError:
    # 如果 config.py 不存在，使用預設路徑（向後相容）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    NBA_STATS_FILE = os.path.join(script_dir, "NBA25-26.xlsx")

# NBA 30支球隊列表（按字母順序排列，方便查找）
NBA_TEAMS = [
    '76人', '公牛', '公鹿', '勇士', '快艇',
    '太陽', '巫師', '尼克', '活塞', '湖人',
    '溜馬', '灰熊', '灰狼', '火箭', '熱火',
    '爵士', '獨行俠', '老鷹', '騎士', '金塊',
    '雷霆', '馬刺', '魔術', '鵜鶘', '黃蜂',
    '塞爾提克', '拓荒者', '國王', '暴龍', '籃網'
]


def create_main_sheet_structure(ws, sheet_name, start_date, end_date):
    """
    建立總表的結構（包含15場比賽 + 統計列）
    
    Args:
        ws: 工作表物件
        sheet_name: 工作表名稱
        start_date: 起始日期（datetime物件）
        end_date: 結束日期（datetime物件）
    """
    ws.title = sheet_name
    
    # ==================== 樣式設定 - 柔和馬卡龍配色 ====================
    # 標題欄配色（左側欄位）：柔和淺藍紫色背景，深色字
    label_fill = PatternFill(start_color="AAB6FB", end_color="AAB6FB", fill_type="solid")
    label_font = Font(bold=True, color="4A4A4A", size=11)
    
    # 數據欄配色（天數、日期、星期）：極淺粉色背景，深色字
    header_fill = PatternFill(start_color="FED7DD", end_color="FED7DD", fill_type="solid")
    header_font = Font(bold=True, color="4A4A4A", size=10)
    
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 計算總天數
    total_days = (end_date - start_date).days + 1
    
    # 中文星期對照
    weekdays = ['一', '二', '三', '四', '五', '六', '日']
    
    # ==================== 設定欄寬 ====================
    ws.column_dimensions['A'].width = 15
    
    # ==================== 建立標題列（A3~A5）====================
    # A3：天數標籤
    ws['A3'] = '天數'
    ws['A3'].fill = label_fill
    ws['A3'].font = label_font
    ws['A3'].alignment = center_align
    
    # A4：日期標籤
    ws['A4'] = '日期'
    ws['A4'].fill = label_fill
    ws['A4'].font = label_font
    ws['A4'].alignment = center_align
    
    # A5：星期標籤
    ws['A5'] = '星期'
    ws['A5'].fill = label_fill
    ws['A5'].font = label_font
    ws['A5'].alignment = center_align
    
    # ==================== 填入日期資料（B3~, B4~, B5~）====================
    # 第3列：天數（1, 2, 3, ...）
    current_col = 2  # 從 B 欄開始（A欄留給標籤）
    for day_num in range(1, total_days + 1):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}3']
        cell.value = day_num
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        # 設定欄寬
        ws.column_dimensions[col_letter].width = 8
        
        current_col += 1
    
    # 第4列：日期（10/22, 10/23, ...）
    current_col = 2
    current_date = start_date
    for day_num in range(total_days):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}4']
        cell.value = current_date.strftime('%m/%d')
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        current_date += timedelta(days=1)
        current_col += 1
    
    # 第5列：星期（三, 四, 五, ...）
    current_col = 2
    current_date = start_date
    for day_num in range(total_days):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}5']
        weekday = weekdays[current_date.weekday()]
        cell.value = weekday
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        current_date += timedelta(days=1)
        current_col += 1
    
    # ==================== 建立比賽列標籤（A6~A20）====================
    chinese_numbers = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十', 
                       '十一', '十二', '十三', '十四', '十五']
    for idx, cn_num in enumerate(chinese_numbers, start=6):
        cell = ws[f'A{idx}']
        cell.value = cn_num
        cell.fill = label_fill
        cell.font = label_font
        cell.alignment = center_align
    
    # 設定列高
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22
    
    for row_num in range(6, 21):
        ws.row_dimensions[row_num].height = 22
    
    # ==================== 建立統計列標籤（A21~A25）====================
    # A21：首場結果
    ws['A21'] = '首場結果'
    ws['A21'].fill = label_fill
    ws['A21'].font = label_font
    ws['A21'].alignment = center_align
    ws.row_dimensions[21].height = 22
    
    # A22：加總
    ws['A22'] = '加總'
    ws['A22'].fill = label_fill
    ws['A22'].font = label_font
    ws['A22'].alignment = center_align
    ws.row_dimensions[22].height = 22
    
    # A23：一周累加
    ws['A23'] = '一周累加'
    ws['A23'].fill = label_fill
    ws['A23'].font = label_font
    ws['A23'].alignment = center_align
    ws.row_dimensions[23].height = 22
    
    # A24：賽季累加
    ws['A24'] = '賽季累加'
    ws['A24'].fill = label_fill
    ws['A24'].font = label_font
    ws['A24'].alignment = center_align
    ws.row_dimensions[24].height = 22
    
    # A25：尾場結果
    ws['A25'] = '尾場結果'
    ws['A25'].fill = label_fill
    ws['A25'].font = label_font
    ws['A25'].alignment = center_align
    ws.row_dimensions[25].height = 22
    
    # ==================== 為統計列設定公式 ====================
    current_date = start_date
    for day_num in range(total_days):
        col_letter = get_column_letter(day_num + 2)  # +2 因為從B欄開始
        weekday_idx = current_date.weekday()  # 0=週一, 1=週二, ..., 6=週日
        
        # 第21列：首場結果（=第6列的值）
        cell = ws[f'{col_letter}21']
        cell.value = f'={col_letter}6'
        cell.alignment = center_align
        
        # 第22列：加總（O=1, X=-1）
        cell = ws[f'{col_letter}22']
        cell.value = f'=SUMPRODUCT(({col_letter}6:{col_letter}20="O")*1+({col_letter}6:{col_letter}20="X")*-1)'
        cell.alignment = center_align
        
        # 第23列：一周累加（週二重新開始）
        cell = ws[f'{col_letter}23']
        if day_num == 0:
            cell.value = f'={col_letter}22'
        else:
            prev_col = get_column_letter(day_num + 1)
            if weekday_idx == 1:  # 週二：重新開始
                cell.value = f'={col_letter}22'
            else:  # 週三~週一：累加
                cell.value = f'={prev_col}23+{col_letter}22'
        cell.alignment = center_align
        
        # 第24列：賽季累加
        cell = ws[f'{col_letter}24']
        if day_num == 0:
            cell.value = f'={col_letter}22'
        else:
            prev_col = get_column_letter(day_num + 1)
            cell.value = f'={prev_col}24+{col_letter}22'
        cell.alignment = center_align
        
        # 第25列：尾場結果（最後一場有數據的比賽）
        cell = ws[f'{col_letter}25']
        cell.value = f'=IFERROR(LOOKUP(2,1/({col_letter}6:{col_letter}20<>""),{col_letter}6:{col_letter}20),"")'
        cell.alignment = center_align
        
        current_date += timedelta(days=1)


def create_team_sheet_structure(ws, team_name, start_date, end_date):
    """
    建立球隊分頁的結構（追蹤 W/L、O/X、大/小）
    
    Args:
        ws: 工作表物件
        team_name: 球隊名稱
        start_date: 起始日期（datetime物件）
        end_date: 結束日期（datetime物件）
    """
    ws.title = team_name
    
    # ==================== 樣式設定 - 柔和馬卡龍配色 ====================
    label_fill = PatternFill(start_color="AAB6FB", end_color="AAB6FB", fill_type="solid")
    label_font = Font(bold=True, color="4A4A4A", size=11)
    
    header_fill = PatternFill(start_color="FED7DD", end_color="FED7DD", fill_type="solid")
    header_font = Font(bold=True, color="4A4A4A", size=10)
    
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 計算總天數
    total_days = (end_date - start_date).days + 1
    
    # 中文星期對照
    weekdays = ['一', '二', '三', '四', '五', '六', '日']
    
    # ==================== 設定欄寬 ====================
    ws.column_dimensions['A'].width = 15
    
    # ==================== A2：顯示球隊名稱 ====================
    ws['A2'] = team_name
    ws['A2'].fill = label_fill
    ws['A2'].font = Font(bold=True, color="4A4A4A", size=14)
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 25
    
    # ==================== 建立標題列（A3~A5）====================
    ws['A3'] = '天數'
    ws['A3'].fill = label_fill
    ws['A3'].font = label_font
    ws['A3'].alignment = center_align
    
    ws['A4'] = '日期'
    ws['A4'].fill = label_fill
    ws['A4'].font = label_font
    ws['A4'].alignment = center_align
    
    ws['A5'] = '星期'
    ws['A5'].fill = label_fill
    ws['A5'].font = label_font
    ws['A5'].alignment = center_align
    
    # ==================== 填入日期資料（B3~, B4~, B5~）====================
    # 第3列：天數
    current_col = 2
    for day_num in range(1, total_days + 1):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}3']
        cell.value = day_num
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[col_letter].width = 8
        current_col += 1
    
    # 第4列：日期
    current_col = 2
    current_date = start_date
    for day_num in range(total_days):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}4']
        cell.value = current_date.strftime('%m/%d')
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        current_date += timedelta(days=1)
        current_col += 1
    
    # 第5列：星期
    current_col = 2
    current_date = start_date
    for day_num in range(total_days):
        col_letter = get_column_letter(current_col)
        cell = ws[f'{col_letter}5']
        weekday = weekdays[current_date.weekday()]
        cell.value = weekday
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        current_date += timedelta(days=1)
        current_col += 1
    
    # 設定標題列的列高
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22
    
    # ==================== 建立統計列標籤（按照76人格式）====================
    # A6: W/L（第一場）
    ws['A6'] = 'W/L'
    ws['A6'].fill = label_fill
    ws['A6'].font = label_font
    ws['A6'].alignment = center_align
    ws.row_dimensions[6].height = 22
    
    # A7: W/L（第二場）
    ws['A7'] = 'W/L'
    ws['A7'].fill = label_fill
    ws['A7'].font = label_font
    ws['A7'].alignment = center_align
    ws.row_dimensions[7].height = 22
    
    # A8: W/L加總
    ws['A8'] = 'W/L加總'
    ws['A8'].fill = label_fill
    ws['A8'].font = label_font
    ws['A8'].alignment = center_align
    ws.row_dimensions[8].height = 22
    
    # A9: O/X（第一場）
    ws['A9'] = 'O/X'
    ws['A9'].fill = label_fill
    ws['A9'].font = label_font
    ws['A9'].alignment = center_align
    ws.row_dimensions[9].height = 22
    
    # A10: O/X（第二場）
    ws['A10'] = 'O/X'
    ws['A10'].fill = label_fill
    ws['A10'].font = label_font
    ws['A10'].alignment = center_align
    ws.row_dimensions[10].height = 22
    
    # A11: O/X加總
    ws['A11'] = 'O/X加總'
    ws['A11'].fill = label_fill
    ws['A11'].font = label_font
    ws['A11'].alignment = center_align
    ws.row_dimensions[11].height = 22
    
    # A12: 大/小（第一場）
    ws['A12'] = '大/小'
    ws['A12'].fill = label_fill
    ws['A12'].font = label_font
    ws['A12'].alignment = center_align
    ws.row_dimensions[12].height = 22
    
    # A13: 大/小（第二場）
    ws['A13'] = '大/小'
    ws['A13'].fill = label_fill
    ws['A13'].font = label_font
    ws['A13'].alignment = center_align
    ws.row_dimensions[13].height = 22
    
    # A14: 大/小加總
    ws['A14'] = '大/小加總'
    ws['A14'].fill = label_fill
    ws['A14'].font = label_font
    ws['A14'].alignment = center_align
    ws.row_dimensions[14].height = 22
    
    # ==================== 為統計列設定公式（累積加總）- 按照76人格式 ====================
    # 加總公式：累積計算第一場和第二場的結果
    # W=1, L=-1; O=1, X=-1; 大=1, 小=-1
    for day_num in range(total_days):
        col_letter = get_column_letter(day_num + 2)
        
        if day_num == 0:
            # 第一天（B欄）：直接根據當天結果計算
            # B8: W/L加總（累積B6和B7）
            cell = ws[f'{col_letter}8']
            cell.value = f'=SUM(IF({col_letter}6="W",1,IF({col_letter}6="L",-1,0)),IF({col_letter}7="W",1,IF({col_letter}7="L",-1,0)))'
            cell.alignment = center_align
            
            # B11: O/X加總（累積B9和B10）
            cell = ws[f'{col_letter}11']
            cell.value = f'=SUM(IF({col_letter}9="O",1,IF({col_letter}9="X",-1,0)),IF({col_letter}10="O",1,IF({col_letter}10="X",-1,0)))'
            cell.alignment = center_align
            
            # B14: 大/小加總（累積B12和B13）
            cell = ws[f'{col_letter}14']
            cell.value = f'=SUM(IF({col_letter}12="大",1,IF({col_letter}12="小",-1,0)),IF({col_letter}13="大",1,IF({col_letter}13="小",-1,0)))'
            cell.alignment = center_align
        else:
            # 第二天之後（C欄開始）：累積加總
            prev_col = get_column_letter(day_num + 1)  # 前一天的欄位
            
            # C8~: W/L累積加總 = 昨天的累積值 + 今天當天的加總（C6和C7）
            # 如果今天沒比賽（C6和C7都為空），則顯示昨天的累積值
            cell = ws[f'{col_letter}8']
            cell.value = f'=IF(AND({col_letter}6="",{col_letter}7=""),{prev_col}8,{prev_col}8+SUM(IF({col_letter}6="W",1,IF({col_letter}6="L",-1,0)),IF({col_letter}7="W",1,IF({col_letter}7="L",-1,0))))'
            cell.alignment = center_align
            
            # C11~: O/X累積加總 = 昨天的累積值 + 今天當天的加總（C9和C10）
            cell = ws[f'{col_letter}11']
            cell.value = f'=IF(AND({col_letter}9="",{col_letter}10=""),{prev_col}11,{prev_col}11+SUM(IF({col_letter}9="O",1,IF({col_letter}9="X",-1,0)),IF({col_letter}10="O",1,IF({col_letter}10="X",-1,0))))'
            cell.alignment = center_align
            
            # C14~: 大/小累積加總 = 昨天的累積值 + 今天當天的加總（C12和C13）
            cell = ws[f'{col_letter}14']
            cell.value = f'=IF(AND({col_letter}12="",{col_letter}13=""),{prev_col}14,{prev_col}14+SUM(IF({col_letter}12="大",1,IF({col_letter}12="小",-1,0)),IF({col_letter}13="大",1,IF({col_letter}13="小",-1,0))))'
            cell.alignment = center_align


def create_nba_statistics_table():
    """
    建立 NBA 25-26 賽季統計表
    包含：總表 + 30支球隊各自的分頁
    """
    # 設定日期範圍
    start_date = datetime(2025, 10, 22)
    end_date = datetime(2026, 5, 31)
    total_days = (end_date - start_date).days + 1
    
    print("=" * 60)
    print("建立 NBA 25-26 賽季統計表")
    print("=" * 60)
    print(f"起始日期：{start_date.strftime('%Y/%m/%d')}")
    print(f"結束日期：{end_date.strftime('%Y/%m/%d')}")
    print(f"總天數：{total_days} 天")
    print(f"總共建立：1 個總表 + {len(NBA_TEAMS)} 支球隊分頁 = {len(NBA_TEAMS) + 1} 個工作表")
    print()
    
    # 建立工作簿
    wb = Workbook()
    
    # ==================== 第一個工作表：總表 ====================
    print("[ 1/31] 建立總表...")
    ws_main = wb.active
    create_main_sheet_structure(ws_main, "NBA 25-26 賽季", start_date, end_date)
    
    # ==================== 為每支球隊建立分頁 ====================
    for idx, team_name in enumerate(NBA_TEAMS, start=2):
        print(f"[{idx:2d}/31] 建立球隊分頁：{team_name}")
        ws_team = wb.create_sheet(title=team_name)
        create_team_sheet_structure(ws_team, team_name, start_date, end_date)
    
    # 儲存檔案（使用設定檔中的路徑）
    filename = NBA_STATS_FILE
    wb.save(filename)
    
    print()
    print("=" * 60)
    print("[完成] 已建立檔案：")
    print(f"   {filename}")
    print()
    print("[工作表結構]")
    print(f"   - 第 1 個工作表：NBA 25-26 賽季（總表）")
    print(f"   - 第 2~31 個工作表：30支球隊分頁")
    print()
    print("[總表結構]")
    print(f"   - A3~A5: 天數、日期、星期")
    print(f"   - A6~A20: 一 ~ 十五（比賽場次）")
    print(f"   - A21~A25: 首場結果、加總、一周累加、賽季累加、尾場結果")
    print()
    print("[球隊分頁結構（76人格式）]")
    print(f"   - A2: 球隊名稱")
    print(f"   - A3~A5: 天數、日期、星期")
    print(f"   - A6: W/L（第一場）")
    print(f"   - A7: W/L（第二場）")
    print(f"   - A8: W/L加總（累積計算A6和A7）")
    print(f"   - A9: O/X（第一場）")
    print(f"   - A10: O/X（第二場）")
    print(f"   - A11: O/X加總（累積計算A9和A10）")
    print(f"   - A12: 大/小（第一場）")
    print(f"   - A13: 大/小（第二場）")
    print(f"   - A14: 大/小加總（累積計算A12和A13）")
    print()
    print("[配色] 柔和馬卡龍風格")
    print(f"   - 標籤欄（A欄）：淺藍紫色 #AAB6FB")
    print(f"   - 數據欄（B欄~）：極淺粉色 #FED7DD")
    print("=" * 60)


if __name__ == "__main__":
    create_nba_statistics_table()
