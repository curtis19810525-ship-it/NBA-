# -*- coding: utf-8 -*-
"""
整合 MLB 總表 + 每日數據 + 玖九盤口 → 分析母版。

用法:
  python build_analysis_workbook.py [YYYYMMDD]
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis_helpers import (
    RECENT_CAPTURES,
    check_file_unlocked,
    game_key,
    merge_games,
    read_daily_games,
    read_result_rows,
    read_snapshot_rows,
    read_zongbiao_games,
)
from config import ANALYSIS_XLSX_FILE, JIUJIU_XLSX_FILE, MLB_XLSX_FILE

TZ = ZoneInfo("Asia/Taipei")
SHEET_SETTINGS = "0_當日設定"
SHEET_PREGAME = "1_賽前整合"
SHEET_POST = "2_賽後檢討"
SHEET_HELP = "3_說明"

FONT_BOLD = Font(bold=True)
FILL_HDR = PatternFill("solid", fgColor="D9E1F2")
FILL_GAME_A = PatternFill("solid", fgColor="E2F0D9")
FILL_GAME_B = PatternFill("solid", fgColor="DDEBF7")
FILL_SPREAD = PatternFill("solid", fgColor="E2F0D9")
FILL_ML = PatternFill("solid", fgColor="C5E0B4")
FILL_TWO = PatternFill("solid", fgColor="A8CF96")

PREGAME_HEADERS = [
    "場次",
    "時間",
    "客@主",
    "合理讓分(客)",
    "合理讓分(主)",
    "讓分",
    "金流(客)",
    "金流(主)",
    "合理性",
    "讓分方",
    "受讓方",
    "賽前備註",
]

# D~J = 玖九時間軸；每場 10 列（模型+表頭+6盤口+檢討+空列）
COL_D, COL_E, COL_F, COL_G, COL_H, COL_I, COL_J = 4, 5, 6, 7, 8, 9, 10
COL_K, COL_L = 11, 12
COLS_E_I = (COL_E, COL_F, COL_G, COL_H, COL_I)
JIUJIU_DATA_ROWS = 6
ROWS_PER_GAME = 10

POST_HEADERS = [
    "場次",
    "時間",
    "客@主",
    "比分",
    "玖九讓分方",
    "玖九讓分結算",
    "玖九獨贏",
    "玖九2分贏",
    "每日OX",
    "是否過盤",
    "檢討",
]

MANUAL_MODEL = {"合理性", "讓分方", "受讓方", "賽前備註"}
MANUAL_REVIEW = "檢討"
MANUAL_POST = {"是否過盤"}


def today_yyyymmdd() -> str:
    return datetime.now(TZ).strftime("%Y%m%d")


def validate_date(s: str) -> str:
    s = s.strip()
    if not re.fullmatch(r"\d{8}", s):
        raise SystemExit("日期格式須為 YYYYMMDD")
    datetime.strptime(s, "%Y%m%d")
    return s


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _coalesce(*vals: Any) -> Any:
    for v in vals:
        if v is not None and str(v).strip() != "":
            return v
    return ""


def read_manual_maps(path: str) -> tuple[dict, dict, dict]:
    if not os.path.exists(path):
        return {}, {}, {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"警告：無法讀取舊分析母版的手動欄位，將略過保留（{e}）")
        return {}, {}, {}
    pre: dict[tuple[str, str, str], dict[str, str]] = {}
    review: dict[tuple[str, str, str], str] = {}
    post: dict[tuple[str, str, str], dict[str, str]] = {}
    if SHEET_PREGAME in wb.sheetnames:
        ws = wb[SHEET_PREGAME]
        headers = [_cell_str(ws.cell(1, c).value) for c in range(1, 13)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}
        row = 2
        while row <= ws.max_row:
            label = _cell_str(ws.cell(row, 1).value)
            if label.startswith("第") and "場" in label:
                time_v = _cell_str(ws.cell(row, 2).value)
                matchup = _cell_str(ws.cell(row, 3).value)
                parts = re.split(r"\s*@\s*", matchup)
                away = parts[0] if parts else ""
                home = re.sub(r"\(主\)$", "", parts[1]).strip() if len(parts) > 1 else ""
                key = game_key(time_v, away, home)
                saved = {}
                for name in MANUAL_MODEL:
                    col = idx.get(name)
                    if col:
                        val = _cell_str(ws.cell(row, col).value)
                        if val:
                            saved[name] = val
                if saved:
                    pre[key] = saved
                rev = _cell_str(ws.cell(row + 8, 2).value)
                if rev:
                    review[key] = rev
                row += ROWS_PER_GAME
                continue
            row += 1
    if SHEET_POST in wb.sheetnames:
        ws = wb[SHEET_POST]
        headers = [_cell_str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}
        for row in range(2, ws.max_row + 1):
            if not _cell_str(ws.cell(row, 1).value):
                continue
            time_v = _cell_str(ws.cell(row, idx.get("時間", 2)).value)
            matchup = _cell_str(ws.cell(row, idx.get("客@主", 3)).value)
            parts = re.split(r"\s*@\s*", matchup)
            away = parts[0] if parts else ""
            home = re.sub(r"\(主\)$", "", parts[1]).strip() if len(parts) > 1 else ""
            key = game_key(time_v, away, home)
            saved = {}
            for name in MANUAL_POST:
                col = idx.get(name)
                if col:
                    val = _cell_str(ws.cell(row, col).value)
                    if val:
                        saved[name] = val
            col = idx.get("檢討")
            if col:
                val = _cell_str(ws.cell(row, col).value)
                if val and key not in review:
                    review[key] = val
            if saved:
                post[key] = saved
    wb.close()
    return pre, review, post


def style_header_row(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = FONT_BOLD
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_workbook() -> Workbook:
    """每次重建活頁簿，避免 delete_rows 留下損壞的合併儲存格。"""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(SHEET_SETTINGS)
    wb.create_sheet(SHEET_PREGAME)
    wb.create_sheet(SHEET_POST)
    wb.create_sheet(SHEET_HELP)
    return wb


def write_settings(ws, focus_date: str, game_count: int) -> None:
    rows = [
        ("關注日", focus_date),
        ("資料更新時間", datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")),
        ("場次數", game_count),
        ("MLB來源", MLB_XLSX_FILE),
        ("玖九來源", JIUJIU_XLSX_FILE),
        ("每日另存", r"D:\龍今\MLB26每日分析\分析_yyyyMMdd.xlsx（請手動另存）"),
    ]
    ws["A1"] = "項目"
    ws["B1"] = "值"
    ws["A1"].font = FONT_BOLD
    ws["B1"].font = FONT_BOLD
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72


def write_help(ws) -> None:
    lines = [
        "【更新方式】",
        "1. 關閉本檔與 OneDrive 上的 MLB26-27數據.xlsx、玖九盤口變化.xlsx",
        "2. 執行「更新分析母版.bat」，輸入關注日",
        "3. 腳本更新 1_賽前整合（含賽後欄）與 2_賽後檢討；L2 備註、第10列檢討 有字保留",
        "4. 分析完成後手動另存至 D:\\龍今\\MLB26每日分析\\",
        "",
        "【1_賽前整合 每場10列】",
        "列2=A~L模型 | 列3=D~J抓取時間 | 列4~9=玖九盤口 | 列10=檢討 | 列11=空列",
        "D=第1次 E~I=最近5次 J=最後1次（不足5次 E起依序填）",
        "有賽果時：C3=比分 K/L=玖九結算",
        "",
        "【2_賽後檢討】一場一列精簡總表，方便掃描當日結果。",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 100


def write_pregame_sheet(
    ws,
    games: list[dict],
    manual: dict,
    review_notes: dict,
) -> None:
    for c, h in enumerate(PREGAME_HEADERS, start=1):
        ws.cell(1, c, h)
    style_header_row(ws, COL_L)

    bet_blocks = (
        ("讓分", 0, 1, FILL_SPREAD),
        ("獨贏", 2, 3, FILL_ML),
        ("2分贏", 4, 5, FILL_TWO),
    )

    row = 2
    for gi, g in enumerate(games):
        base_fill = FILL_GAME_A if gi % 2 == 0 else FILL_GAME_B
        key = game_key(g["time"], g["away"], g["home"])
        saved = manual.get(key, {})
        jj = g.get("jiujiu") or {}
        settle = g.get("settlement") or {}
        has_result = bool(g.get("score_display"))

        # 列2：模型 A~L
        model = [
            g.get("game_label"),
            g.get("time"),
            g.get("matchup"),
            g.get("fair_away"),
            g.get("fair_home"),
            g.get("handicap_line"),
            g.get("flow_away"),
            g.get("flow_home"),
            _coalesce(saved.get("合理性"), g.get("reasonableness_src")),
            _coalesce(saved.get("讓分方"), g.get("handicap_side_src")),
            _coalesce(saved.get("受讓方"), g.get("receiver_side_src")),
            saved.get("賽前備註") or "",
        ]
        for c, val in enumerate(model, start=1):
            cell = ws.cell(row, c, val)
            cell.fill = base_fill
            cell.alignment = Alignment(vertical="center")

        # 列3：時間表頭 D~J
        hdr_row = row + 1
        ws.cell(hdr_row, COL_D, "第一次")
        ws.cell(hdr_row, COL_D).font = FONT_BOLD
        ws.cell(hdr_row, COL_D).alignment = Alignment(horizontal="center")
        for i, col in enumerate(COLS_E_I):
            ws.cell(hdr_row, col, jj.get("headers_e_i", [""] * RECENT_CAPTURES)[i])
            ws.cell(hdr_row, col).font = FONT_BOLD
            ws.cell(hdr_row, col).alignment = Alignment(
                horizontal="center", wrap_text=True, vertical="center"
            )
        ws.cell(hdr_row, COL_J, "最後一次")
        ws.cell(hdr_row, COL_J).font = FONT_BOLD
        ws.cell(hdr_row, COL_J).alignment = Alignment(horizontal="center")
        if has_result:
            ws.cell(hdr_row, 3, g.get("score_display", ""))
            ws.cell(hdr_row, 3).font = FONT_BOLD
        for c in range(1, COL_L + 1):
            ws.cell(hdr_row, c).fill = base_fill

        away_name = str(g.get("away") or "")
        home_name = str(g.get("home") or "")
        if home_name and "(主)" not in home_name:
            home_name = f"{home_name}(主)"

        col_d = jj.get("col_d") or [""] * 6
        col_j = jj.get("col_j") or [""] * 6
        cols_e_i = jj.get("cols_e_i") or [[""] * 6 for _ in range(RECENT_CAPTURES)]

        for label, r0, r1, block_fill in bet_blocks:
            r_top = row + 2 + r0
            r_bot = row + 2 + r1
            ws.merge_cells(start_row=r_top, start_column=3, end_row=r_bot, end_column=3)
            ws.cell(r_top, 3, label)
            ws.cell(r_top, 3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r_top, 2, away_name)
            ws.cell(r_bot, 2, home_name)
            ws.cell(r_top, COL_D, col_d[r0])
            ws.cell(r_bot, COL_D, col_d[r1])
            ws.cell(r_top, COL_J, col_j[r0])
            ws.cell(r_bot, COL_J, col_j[r1])
            for i, col in enumerate(COLS_E_I):
                ws.cell(r_top, col, cols_e_i[i][r0])
                ws.cell(r_bot, col, cols_e_i[i][r1])
            for rr in (r_top, r_bot):
                for cc in range(1, COL_L + 1):
                    c = ws.cell(rr, cc)
                    c.fill = block_fill
                    c.alignment = Alignment(horizontal="center", vertical="center")

            if has_result:
                if label == "讓分":
                    ws.cell(r_top, COL_K, settle.get("spread_k_away", ""))
                    ws.cell(r_bot, COL_K, settle.get("spread_k_home", ""))
                    ws.merge_cells(start_row=r_top, start_column=COL_L, end_row=r_bot, end_column=COL_L)
                    ws.cell(r_top, COL_L, settle.get("spread_l", ""))
                    ws.cell(r_top, COL_L).alignment = Alignment(horizontal="center", vertical="center")
                elif label == "獨贏":
                    ws.merge_cells(start_row=r_top, start_column=COL_L, end_row=r_bot, end_column=COL_L)
                    ws.cell(r_top, COL_L, settle.get("ml_l", ""))
                    ws.cell(r_top, COL_L).alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws.merge_cells(start_row=r_top, start_column=COL_L, end_row=r_bot, end_column=COL_L)
                    ws.cell(r_top, COL_L, settle.get("two_l", ""))
                    ws.cell(r_top, COL_L).alignment = Alignment(horizontal="center", vertical="center")

        # 列10：檢討
        rev_row = row + 8
        ws.cell(rev_row, 1, "檢討")
        ws.cell(rev_row, 1).font = FONT_BOLD
        ws.merge_cells(start_row=rev_row, start_column=2, end_row=rev_row, end_column=COL_L)
        ws.cell(rev_row, 2, review_notes.get(key, ""))
        ws.cell(rev_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, COL_L + 1):
            ws.cell(rev_row, c).fill = base_fill

        row += ROWS_PER_GAME

    widths = {
        "A": 10,
        "B": 14,
        "C": 10,
        "D": 11,
        "E": 11,
        "F": 11,
        "G": 11,
        "H": 11,
        "I": 11,
        "J": 11,
        "K": 12,
        "L": 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_post_sheet(
    ws,
    games: list[dict],
    manual_post: dict,
    review_notes: dict,
) -> None:
    for c, h in enumerate(POST_HEADERS, start=1):
        ws.cell(1, c, h)
    style_header_row(ws, len(POST_HEADERS))
    for i, g in enumerate(games, start=2):
        key = game_key(g["time"], g["away"], g["home"])
        saved = manual_post.get(key, {})
        res = g.get("result") or {}
        ws.cell(i, 1, g.get("game_label"))
        ws.cell(i, 2, g.get("time"))
        ws.cell(i, 3, g.get("matchup"))
        ws.cell(i, 4, g.get("score_display") or res.get("比分", ""))
        ws.cell(i, 5, res.get("讓分方", ""))
        ws.cell(i, 6, res.get("讓分結算", ""))
        ws.cell(i, 7, res.get("獨贏結果", ""))
        ws.cell(i, 8, res.get("兩分贏結果", ""))
        ws.cell(i, 9, g.get("daily_ox", ""))
        ws.cell(i, 10, _coalesce(saved.get("是否過盤"), g.get("pass_label_src")))
        ws.cell(i, 11, review_notes.get(key, ""))
    for col, w in zip("ABCDEFGHIJK", [10, 8, 24, 8, 16, 10, 16, 16, 8, 10, 30]):
        ws.column_dimensions[col].width = w


def save_workbook(wb: Workbook, path: str) -> None:
    """先寫暫存檔、驗證可讀後再覆蓋，避免 OneDrive/Excel 開到半套 XML。"""
    out_dir = os.path.dirname(path) or "."
    os.makedirs(out_dir, exist_ok=True)
    tmp = path + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    wb.save(tmp)
    try:
        test = load_workbook(tmp, read_only=True)
        test.close()
    except Exception as e:
        os.remove(tmp)
        raise SystemExit(f"產生的 Excel 無法通過驗證，未覆蓋原檔：{e}") from e
    if os.path.exists(path):
        os.remove(path)
    shutil.move(tmp, path)


def remove_invalid_xlsm() -> None:
    legacy = os.path.join(os.path.dirname(ANALYSIS_XLSX_FILE), "MLB26-分析母版.xlsm")
    if not os.path.exists(legacy):
        return
    try:
        os.remove(legacy)
        print(f"已刪除無法開啟的舊檔：{legacy}")
    except OSError as e:
        print(f"警告：無法刪除舊 xlsm：{legacy} ({e})")


def build(focus_date: str) -> None:
    check_file_unlocked(MLB_XLSX_FILE)
    check_file_unlocked(JIUJIU_XLSX_FILE)
    check_file_unlocked(ANALYSIS_XLSX_FILE)

    remove_invalid_xlsm()
    manual_pre, review_notes, manual_post = read_manual_maps(ANALYSIS_XLSX_FILE)

    print("讀取 MLB 主檔 …")
    mlb_wb = load_workbook(MLB_XLSX_FILE, read_only=True, data_only=True)
    daily = read_daily_games(mlb_wb, focus_date)
    zongbiao = read_zongbiao_games(mlb_wb, focus_date)
    mlb_wb.close()
    if not daily:
        raise SystemExit(f"每日數據找不到 {focus_date} 的場次")

    print("讀取玖九檔 …")
    jj_wb = load_workbook(JIUJIU_XLSX_FILE, read_only=True, data_only=True)
    snapshots = read_snapshot_rows(jj_wb)
    results = read_result_rows(jj_wb)
    jj_wb.close()

    games = merge_games(daily, zongbiao, snapshots, results, focus_date)
    print(f"關注日 {focus_date}：{len(games)} 場")

    wb = create_workbook()
    write_settings(wb[SHEET_SETTINGS], focus_date, len(games))
    write_pregame_sheet(wb[SHEET_PREGAME], games, manual_pre, review_notes)
    write_post_sheet(wb[SHEET_POST], games, manual_post, review_notes)
    write_help(wb[SHEET_HELP])

    save_workbook(wb, ANALYSIS_XLSX_FILE)
    print(f"✓ 已更新：{ANALYSIS_XLSX_FILE}")


def main() -> None:
    parser = argparse.ArgumentParser(description="更新 MLB 分析母版")
    parser.add_argument("date", nargs="?", default=today_yyyymmdd(), help="關注日 YYYYMMDD")
    args = parser.parse_args()
    focus_date = validate_date(args.date)
    print("=" * 60)
    print("MLB 分析母版更新")
    print("=" * 60)
    print(f"關注日：{focus_date}")
    print("請確認已關閉 Excel 中的相關檔案。")
    print()
    build(focus_date)


if __name__ == "__main__":
    main()
