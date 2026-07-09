# -*- coding: utf-8 -*-
"""
爬取 Rotowire MLB 先發名單（日期範圍）
資料來源: https://www.rotowire.com/baseball/daily-lineups.php
寫入策略: 每次覆蓋整個「先發名單」分頁

日期語意（規則 3）：
  輸入的 YYYYMMDD 為「台灣日曆日」。
  向 Rotowire 請求時，使用「該台灣日 Asia/Taipei 00:00」換算到
  America/New_York 當下的日曆日，作為美東查詢日（與網站分頁一致）。
  Excel「日期」欄寫入台灣日，與其他步驟之台灣日一致。
"""

import os
import re
import sys
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo
    USE_ZONEINFO = True
except ImportError:
    USE_ZONEINFO = False

try:
    from config import MLB_XLSX_FILE
except ImportError:
    MLB_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"


BASE_URL = "https://www.rotowire.com/baseball/daily-lineups.php"
SHEET_NAME = "先發名單"
HEADERS = [
    "日期",
    "比賽時間",
    "客隊",
    "主隊",
    "客隊先發投手",
    "主隊先發投手",
    "客隊先發1棒",
    "客隊先發2棒",
    "客隊先發3棒",
    "客隊先發4棒",
    "客隊先發5棒",
    "客隊先發6棒",
    "客隊先發7棒",
    "客隊先發8棒",
    "客隊先發9棒",
    "主隊先發1棒",
    "主隊先發2棒",
    "主隊先發3棒",
    "主隊先發4棒",
    "主隊先發5棒",
    "主隊先發6棒",
    "主隊先發7棒",
    "主隊先發8棒",
    "主隊先發9棒",
]
HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}
MONTH_MAP = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}
TEAM_ABBR_TO_ZH = {
    "BAL": "金鶯",
    "BOS": "紅襪",
    "NYY": "洋基",
    "TB": "光芒",
    "TOR": "藍鳥",
    "CHW": "白襪",
    "CWS": "白襪",
    "CLE": "守護者",
    "DET": "老虎",
    "KC": "皇家",
    "MIN": "雙城",
    "HOU": "太空人",
    "LAA": "天使",
    "ATH": "運動家",
    "OAK": "運動家",
    "SEA": "水手",
    "TEX": "遊騎兵",
    "ATL": "勇士",
    "MIA": "馬林魚",
    "NYM": "大都會",
    "PHI": "費城人",
    "WSH": "國民",
    "WAS": "國民",
    "CHC": "小熊",
    "CIN": "紅人",
    "MIL": "釀酒人",
    "PIT": "海盜",
    "STL": "紅雀",
    "ARI": "響尾蛇",
    "AZ": "響尾蛇",
    "COL": "落磯",
    "LAD": "道奇",
    "SD": "教士",
    "SDP": "教士",
    "SF": "巨人",
    "SFG": "巨人",
}


def _et_today():
    if USE_ZONEINFO:
        return datetime.now(ZoneInfo("America/New_York")).date()
    return datetime.utcnow().date()


def _taiwan_date_to_et_fetch_date(tw_date):
    """台灣日曆日 → Rotowire 用的美東日曆日（該台灣日台北 00:00 換算）。"""
    if USE_ZONEINFO:
        dt_tw = datetime.combine(tw_date, datetime.min.time()).replace(
            tzinfo=ZoneInfo("Asia/Taipei")
        )
        return dt_tw.astimezone(ZoneInfo("America/New_York")).date()
    return tw_date - timedelta(days=1)


def _parse_date(value):
    return datetime.strptime(value, "%Y%m%d").date()


def _prompt_date_range():
    while True:
        start = input("Enter Start Date (YYYYMMDD, 台灣日): ").strip()
        end = input("Enter End Date (YYYYMMDD, 台灣日): ").strip()
        try:
            start_date = _parse_date(start)
            end_date = _parse_date(end)
        except ValueError:
            print("日期格式錯誤，請使用 YYYYMMDD。")
            continue
        if start_date > end_date:
            print("錯誤：起始日期不能晚於結束日期。")
            continue
        return start_date, end_date


def _parse_args():
    args = sys.argv[1:]
    if len(args) == 2:
        return _parse_date(args[0]), _parse_date(args[1])
    if len(args) == 1 and "~" in args[0]:
        s, e = [x.strip() for x in args[0].split("~", 1)]
        return _parse_date(s), _parse_date(e)
    if len(args) == 0:
        return _prompt_date_range()
    raise ValueError(
        "參數格式錯誤。請使用：YYYYMMDD YYYYMMDD 或 YYYYMMDD~YYYYMMDD（皆為台灣日）"
    )


def _iter_dates(start_date, end_date):
    cur = start_date
    while cur <= end_date:
        yield cur
        cur += timedelta(days=1)


def _build_url(target_date):
    today = _et_today()
    if target_date == today:
        return BASE_URL, "today"
    if target_date == today + timedelta(days=1):
        return f"{BASE_URL}?date=tomorrow", "tomorrow"
    return f"{BASE_URL}?date={target_date.strftime('%Y-%m-%d')}", "explicit"


def _extract_page_month_day(soup):
    for node in soup.select('a[href*="slateID="]'):
        text = node.get_text(" ", strip=True)
        m = re.search(r"\b([A-Z][a-z]{2})\s+(\d{1,2})\b", text)
        if not m:
            continue
        month = MONTH_MAP.get(m.group(1))
        day = int(m.group(2))
        if month:
            return month, day
    return None


def _page_matches_target(soup, target_date, mode):
    if mode in {"today", "tomorrow"}:
        return True
    page_md = _extract_page_month_day(soup)
    if not page_md:
        return True
    return page_md == (target_date.month, target_date.day)


def _team_abbr_to_zh(raw):
    key = (raw or "").strip().upper()
    return TEAM_ABBR_TO_ZH.get(key, key)


def _clean_pitcher_name(text):
    s = (text or "").strip()
    s = re.sub(r"\s+[RL]$", "", s)
    return s


def _extract_batters(card, side_css):
    names = []
    for a in card.select(f"{side_css} .lineup__player a"):
        name = a.get_text(" ", strip=True)
        if name:
            names.append(name)
    names = names[:9]
    while len(names) < 9:
        names.append("")
    return names


def _extract_games_from_soup(soup, display_date):
    """display_date：寫入 Excel「日期」欄之用（台灣日）。"""
    rows = []
    cards = soup.select(".lineup.is-mlb")
    date_text = display_date.strftime("%Y/%m/%d")
    for card in cards:
        away_raw = ""
        home_raw = ""
        away_node = card.select_one(".lineup__team.is-visit .lineup__abbr")
        if away_node:
            away_raw = away_node.get_text(" ", strip=True)
        home_node = card.select_one(".lineup__team.is-home .lineup__abbr")
        if home_node:
            home_raw = home_node.get_text(" ", strip=True)
        if not away_raw:
            away_fallback = card.select_one(".lineup__team.is-visit")
            away_raw = away_fallback.get_text(" ", strip=True) if away_fallback else ""
        if not home_raw:
            home_fallback = card.select_one(".lineup__team.is-home")
            home_raw = home_fallback.get_text(" ", strip=True) if home_fallback else ""

        game_time = ""
        time_node = card.select_one(".lineup__time")
        if time_node:
            game_time = time_node.get_text(" ", strip=True)

        away_pitcher = ""
        home_pitcher = ""
        away_pitcher_node = card.select_one(".lineup__list.is-visit .lineup__player-highlight-name")
        if away_pitcher_node:
            away_pitcher = _clean_pitcher_name(away_pitcher_node.get_text(" ", strip=True))
        home_pitcher_node = card.select_one(".lineup__list.is-home .lineup__player-highlight-name")
        if home_pitcher_node:
            home_pitcher = _clean_pitcher_name(home_pitcher_node.get_text(" ", strip=True))

        away_batters = _extract_batters(card, ".lineup__list.is-visit")
        home_batters = _extract_batters(card, ".lineup__list.is-home")

        row = [
            date_text,
            game_time,
            _team_abbr_to_zh(away_raw),
            _team_abbr_to_zh(home_raw),
            away_pitcher,
            home_pitcher,
        ] + away_batters + home_batters
        rows.append(row)
    return rows


def _fetch_date_rows(et_fetch_date, tw_display_date):
    """et_fetch_date：請求 Rotowire 之美東日；tw_display_date：Excel 日期欄（台灣日）。"""
    url, mode = _build_url(et_fetch_date)
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=25)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    if not _page_matches_target(soup, et_fetch_date, mode):
        return None, (
            f"來源頁面沒有美東 {et_fetch_date.strftime('%Y-%m-%d')} 可用內容"
        )
    rows = _extract_games_from_soup(soup, tw_display_date)
    return rows, None


def _ensure_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def _write_sheet(xlsx_file, rows):
    wb = load_workbook(xlsx_file)
    ws = _ensure_sheet(wb, SHEET_NAME)

    for idx, name in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx, value=name)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(xlsx_file)


def main():
    print("=" * 60)
    print("爬取 Rotowire MLB 先發名單（日期範圍，輸入為台灣日）")
    print("=" * 60)

    try:
        start_date, end_date = _parse_args()
    except ValueError as exc:
        print(f"錯誤：{exc}")
        return 1

    print(
        f"台灣日範圍：{start_date.strftime('%Y%m%d')}~{end_date.strftime('%Y%m%d')} "
        "（Rotowire 依美東日曆換算後抓取）"
    )
    print(f"Excel: {MLB_XLSX_FILE}")
    if not os.path.exists(MLB_XLSX_FILE):
        print("錯誤：Excel 檔案不存在。")
        return 1

    all_rows = []
    unavailable_dates = []
    for tw_day in _iter_dates(start_date, end_date):
        d_text = tw_day.strftime("%Y%m%d")
        et_day = _taiwan_date_to_et_fetch_date(tw_day)
        et_text = et_day.strftime("%Y-%m-%d")
        print(f"\n[{d_text}] 台灣日 → 美東查詢日 {et_text}，抓取中...")
        try:
            rows, err = _fetch_date_rows(et_day, tw_day)
        except Exception as exc:
            print(f"  失敗：{exc}")
            unavailable_dates.append(d_text)
            continue
        if err:
            print(f"  跳過：{err}")
            unavailable_dates.append(d_text)
            continue
        print(f"  成功：{len(rows)} 場")
        all_rows.extend(rows)

    _write_sheet(MLB_XLSX_FILE, all_rows)
    print("\n" + "-" * 60)
    print(f"已覆蓋「{SHEET_NAME}」分頁，共寫入 {len(all_rows)} 場")
    if unavailable_dates:
        print("以下日期無可用資料或抓取失敗：")
        print(", ".join(unavailable_dates))
    print("-" * 60)

    if unavailable_dates and not all_rows:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
