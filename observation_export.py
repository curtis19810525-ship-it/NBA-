# -*- coding: utf-8 -*-
"""
匯出「盤口觀察.xlsx」讀取與走勢摘要（供 integrated_export 合併用）。

獨立四份 txt 已停用；請使用 integrated_export.export_integrated_for_pipeline。
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import load_workbook

from analysis_helpers import _norm_date, team_nickname
from config import OBSERVATION_XLSX_FILE
from fill_observation import (
    COL_DATE,
    COL_FAV2_HEAD,
    COL_FAV2_TAIL,
    COL_FAVML_HEAD,
    COL_FAVML_TAIL,
    COL_FLOW,
    COL_ML_RESULT,
    COL_QUANT,
    COL_REC2_HEAD,
    COL_REC2_TAIL,
    COL_RECML_HEAD,
    COL_RECML_TAIL,
    COL_SIDE,
    COL_SP_HEAD,
    COL_SP_RESULT,
    COL_SP_TAIL,
    COL_TEAM,
    COL_TWO_RESULT,
    _cell_str,
    fmt_flow,
)
from notebooklm_export import EMPTY, EXPORT_DIR, _fmt_empty

KIND_OBS_PREGAME = "盤口觀察（賽前）"
KIND_OBS_RESULTS = "盤口結果（賽後）"


def observation_filename(yyyymmdd: str, kind: str) -> str:
    """例：20260814盤口觀察（賽前）.txt"""
    return f"{yyyymmdd}{kind}.txt"


def _fmt_cell(value) -> str:
    s = _cell_str(value)
    return s if s else EMPTY


def _fmt_flow_export(value) -> str:
    s = fmt_flow(value)
    return s if s else EMPTY


def summarize_trend(head: str, tail: str, *, market: str = "odds") -> str:
    """頭→尾走勢摘要。"""
    h, t = _fmt_cell(head), _fmt_cell(tail)
    if h == EMPTY or t == EMPTY:
        return EMPTY
    if h == t:
        return "持平"
    if market == "spread":
        return f"變動（{h}→{t}）"
    try:
        fh, ft = float(h), float(t)
        diff = ft - fh
        if abs(diff) < 0.02:
            return "持平"
        if diff < 0:
            return "降（熱）" if diff <= -0.05 else "微降"
        return "升（冷）" if diff >= 0.05 else "微升"
    except ValueError:
        return f"變動（{h}→{t}）"


def read_observation_rows(ws, focus_date: str) -> list[dict]:
    """讀取盤口觀察某日所有列。"""
    rows: list[dict] = []
    for r in range(2, (ws.max_row or 1) + 1):
        if _norm_date(ws.cell(r, COL_DATE).value) != focus_date:
            continue
        team = team_nickname(_cell_str(ws.cell(r, COL_TEAM).value))
        if not team:
            continue
        rows.append(
            {
                "game_label": f"第{len(rows) + 1}場",
                "team": team,
                "flow": ws.cell(r, COL_FLOW).value,
                "side": _cell_str(ws.cell(r, COL_SIDE).value),
                "quant": _cell_str(ws.cell(r, COL_QUANT).value),
                "sp_head": _cell_str(ws.cell(r, COL_SP_HEAD).value),
                "sp_tail": _cell_str(ws.cell(r, COL_SP_TAIL).value),
                "sp_result": _cell_str(ws.cell(r, COL_SP_RESULT).value),
                "favml_head": _cell_str(ws.cell(r, COL_FAVML_HEAD).value),
                "favml_tail": _cell_str(ws.cell(r, COL_FAVML_TAIL).value),
                "recml_head": _cell_str(ws.cell(r, COL_RECML_HEAD).value),
                "recml_tail": _cell_str(ws.cell(r, COL_RECML_TAIL).value),
                "ml_result": _cell_str(ws.cell(r, COL_ML_RESULT).value),
                "fav2_head": _cell_str(ws.cell(r, COL_FAV2_HEAD).value),
                "fav2_tail": _cell_str(ws.cell(r, COL_FAV2_TAIL).value),
                "rec2_head": _cell_str(ws.cell(r, COL_REC2_HEAD).value),
                "rec2_tail": _cell_str(ws.cell(r, COL_REC2_TAIL).value),
                "two_result": _cell_str(ws.cell(r, COL_TWO_RESULT).value),
            }
        )
    return rows


def format_observation_text(
    focus_date: str,
    rows: list[dict],
    *,
    kind: str = KIND_OBS_PREGAME,
    exported_at: datetime | None = None,
) -> str:
    exported_at = exported_at or datetime.now()
    lines = [
        "【MLB 盤口觀察匯出｜資料來源】",
        f"類型：{kind}",
        f"日期：{focus_date}",
        "來源檔：盤口觀察.xlsx → 工作表1",
        "匯出範圍：當日所有場次（頭盤＝玖九第1筆快照；尾盤＝開賽前最後快照）",
        f"匯出時間：{exported_at:%Y-%m-%d %H:%M:%S}",
        "",
        "說明：",
        "- 本檔供 NotebookLM 與「賽前推薦.txt」交叉比對用",
        "- 頭→尾：出盤到賽前的盤口移動方向",
        "- H/M/R：讓分結果／獨贏結果／2分結果（賽前可能為 _）",
        "- 空值：_",
        "",
        "========================================",
    ]
    if not rows:
        lines.append("（此日期在「盤口觀察」查無場次）")
    else:
        for row in rows:
            lines.extend(
                [
                    row["game_label"],
                    f"讓分球隊：{_fmt_empty(row.get('team'))}",
                    f"主客：{_fmt_empty(row.get('side'))}",
                    f"金流：{_fmt_flow_export(row.get('flow'))}",
                    f"量化：{_fmt_empty(row.get('quant'))}",
                    f"讓分(頭)：{_fmt_cell(row.get('sp_head'))}",
                    f"讓分(尾)：{_fmt_cell(row.get('sp_tail'))}",
                    f"讓分走勢：{summarize_trend(row.get('sp_head'), row.get('sp_tail'), market='spread')}",
                    f"讓分結果：{_fmt_cell(row.get('sp_result'))}",
                    f"讓分獨贏(頭)：{_fmt_cell(row.get('favml_head'))}",
                    f"讓分獨贏(尾)：{_fmt_cell(row.get('favml_tail'))}",
                    f"讓分獨贏走勢：{summarize_trend(row.get('favml_head'), row.get('favml_tail'))}",
                    f"受讓獨贏(頭)：{_fmt_cell(row.get('recml_head'))}",
                    f"受讓獨贏(尾)：{_fmt_cell(row.get('recml_tail'))}",
                    f"受讓獨贏走勢：{summarize_trend(row.get('recml_head'), row.get('recml_tail'))}",
                    f"獨贏結果：{_fmt_cell(row.get('ml_result'))}",
                    f"讓2分贏(頭)：{_fmt_cell(row.get('fav2_head'))}",
                    f"讓2分贏(尾)：{_fmt_cell(row.get('fav2_tail'))}",
                    f"讓2分走勢：{summarize_trend(row.get('fav2_head'), row.get('fav2_tail'))}",
                    f"受讓2分贏(頭)：{_fmt_cell(row.get('rec2_head'))}",
                    f"受讓2分贏(尾)：{_fmt_cell(row.get('rec2_tail'))}",
                    f"受讓2分走勢：{summarize_trend(row.get('rec2_head'), row.get('rec2_tail'))}",
                    f"2分結果：{_fmt_cell(row.get('two_result'))}",
                    "",
                ]
            )
    lines.append("========================================")
    lines.append(f"合計場次：{len(rows)}")
    lines.append("")
    return "\n".join(lines)


def export_observation_for_notebooklm(
    focus_date: str,
    *,
    kind: str = KIND_OBS_PREGAME,
    observation_xlsx: str | None = None,
    export_dir: str | None = None,
) -> tuple[str, int]:
    """匯出單一日期盤口觀察。回傳 (檔案路徑, 場次數)。"""
    observation_xlsx = observation_xlsx or OBSERVATION_XLSX_FILE
    export_dir = export_dir or EXPORT_DIR
    os.makedirs(export_dir, exist_ok=True)

    if not os.path.exists(observation_xlsx):
        raise SystemExit(f"找不到檔案：{observation_xlsx}")

    wb = load_workbook(observation_xlsx, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = read_observation_rows(ws, focus_date)
    finally:
        wb.close()

    text = format_observation_text(focus_date, rows, kind=kind)
    out_path = os.path.join(export_dir, observation_filename(focus_date, kind))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    return out_path, len(rows)


def export_observations_for_pipeline(
    results_day: str,
    focus_day: str,
    *,
    observation_xlsx: str | None = None,
) -> dict:
    """賽果日匯出盤口結果（賽後）；關注日匯出盤口觀察（賽前）。"""
    r_path, r_n = export_observation_for_notebooklm(
        results_day,
        kind=KIND_OBS_RESULTS,
        observation_xlsx=observation_xlsx,
    )
    f_path, f_n = export_observation_for_notebooklm(
        focus_day,
        kind=KIND_OBS_PREGAME,
        observation_xlsx=observation_xlsx,
    )
    export_dir = os.path.dirname(os.path.abspath(f_path))
    return {
        "export_dir": export_dir,
        "obs_results_path": r_path,
        "obs_results_games": r_n,
        "obs_pregame_path": f_path,
        "obs_pregame_games": f_n,
    }
