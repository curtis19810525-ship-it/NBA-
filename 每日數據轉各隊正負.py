"""
每日數據轉各隊正負
從「每日數據」分頁讀取資料，轉換並寫入「各隊正負」分頁
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import re
import os

# 引入設定檔
try:
    from config import NBA_STATS_FILE, TEAM_STATS_START_DATE, TEAM_STATS_START_COLUMN
except ImportError:
    # 如果 config.py 不存在，使用預設路徑（OneDrive 目錄）
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"
    TEAM_STATS_START_DATE = "20260326"
    TEAM_STATS_START_COLUMN = 16  # P 欄

# MLB 30 支球隊（按「各隊正負」分頁的順序）
MLB_TEAMS_ORDER = [
    '金鶯', '紅襪', '洋基', '光芒', '藍鳥',
    '白襪', '守護者', '老虎', '皇家', '雙城',
    '太空人', '天使', '運動家', '水手', '遊騎兵',
    '勇士', '馬林魚', '大都會', '費城人', '國民',
    '小熊', '紅人', '釀酒人', '海盜', '紅雀',
    '響尾蛇', '落磯', '道奇', '教士', '巨人'
]

# 球隊名稱對應表（將「每日數據」中的變體對應到「各隊正負」分頁中的標準名稱）
TEAM_NAME_MAPPING = {
    '巴爾的摩金鶯': '金鶯',
    '波士頓紅襪': '紅襪',
    '紐約洋基': '洋基',
    '坦帕灣光芒': '光芒',
    '多倫多藍鳥': '藍鳥',
    '芝加哥白襪': '白襪',
    '克里夫蘭守護者': '守護者',
    '底特律老虎': '老虎',
    '堪薩斯城皇家': '皇家',
    '明尼蘇達雙城': '雙城',
    '休士頓太空人': '太空人',
    '洛杉磯天使': '天使',
    '奧克蘭運動家': '運動家',
    '運動家隊': '運動家',
    '西雅圖水手': '水手',
    '德州遊騎兵': '遊騎兵',
    '亞特蘭大勇士': '勇士',
    '邁阿密馬林魚': '馬林魚',
    '紐約大都會': '大都會',
    '費城費城人': '費城人',
    '華盛頓國民': '國民',
    '芝加哥小熊': '小熊',
    '辛辛那提紅人': '紅人',
    '密爾瓦基釀酒人': '釀酒人',
    '匹茲堡海盜': '海盜',
    '聖路易紅雀': '紅雀',
    '亞利桑那響尾蛇': '響尾蛇',
    '科羅拉多落磯': '落磯',
    '洛杉磯道奇': '道奇',
    '聖地牙哥教士': '教士',
    '舊金山巨人': '巨人',
}


def normalize_team_name(team_name):
    """
    將球隊名稱標準化（對應到「各隊正負」分頁中的標準名稱）
    
    Args:
        team_name: 原始球隊名稱
    
    Returns:
        str: 標準化的球隊名稱
    """
    if not team_name:
        return team_name
    
    # 先去除空白
    team_name = str(team_name).strip()
    
    # 使用對應表轉換
    normalized = TEAM_NAME_MAPPING.get(team_name, team_name)
    
    return normalized


def get_team_row(team_name):
    """
    取得球隊在「各隊正負」分頁中的起始行
    
    Args:
        team_name: 球隊名稱（可能是變體）
    
    Returns:
        int: 起始行（第1行是球隊名稱，第5行是W/L，第7行是O/X）
    """
    # 先標準化球隊名稱
    normalized_name = normalize_team_name(team_name)
    
    try:
        team_index = MLB_TEAMS_ORDER.index(normalized_name)
        # 每個球隊區塊佔10行，起始行 = (team_index * 10) + 1
        start_row = (team_index * 10) + 1
        return start_row
    except ValueError:
        print(f"警告：找不到球隊 '{team_name}'（標準化後：'{normalized_name}'），跳過")
        return None


def calculate_wl(away_score, home_score, is_away_team):
    """
    計算 W/L（直接比分判斷）
    
    Args:
        away_score: 客隊分數
        home_score: 主隊分數
        is_away_team: 是否為客隊（True=客隊，False=主隊）
    
    Returns:
        str: 'W' 或 'L' 或 ''（如果分數為空）
    """
    if away_score is None or home_score is None:
        return ''
    
    try:
        away_score_int = int(away_score)
        home_score_int = int(home_score)
        
        if is_away_team:
            # 客隊：分數高則 W，分數低則 L
            return 'W' if away_score_int > home_score_int else 'L'
        else:
            # 主隊：分數高則 W，分數低則 L
            return 'W' if home_score_int > away_score_int else 'L'
    except (ValueError, TypeError):
        return ''


def calculate_ox_team(away_handicap, home_handicap, away_score, home_score, is_away_team):
    """
    計算 O/X（根據讓分條件，不考慮比例）
    用於「各隊正負」分頁
    
    Args:
        away_handicap: 客隊讓分（D欄，如「6分輸」或空）
        home_handicap: 主隊讓分（G欄，如「13分贏」或空）
        away_score: 客隊分數
        home_score: 主隊分數
        is_away_team: 是否為客隊（True=客隊，False=主隊）
    
    Returns:
        str: 'O' 或 'X' 或 ''（空白）
    """
    # 檢查讓分和分數（0 分是有效分數）
    def _is_missing_score(v):
        if v is None:
            return True
        if isinstance(v, str) and not v.strip():
            return True
        return False

    if (not away_handicap and not home_handicap) or _is_missing_score(away_score) or _is_missing_score(home_score):
        return ''
    
    # 轉換分數為整數
    try:
        away_score_int = int(away_score)
        home_score_int = int(home_score)
    except (ValueError, TypeError):
        return ''
    
    # 判斷讓分條件
    if away_handicap:
        # 客隊讓分（D欄）
        handicap_match = re.search(r'(\d+)分(輸|贏)', str(away_handicap))
        if handicap_match:
            handicap_points = int(handicap_match.group(1))
            handicap_result = handicap_match.group(2)  # "輸" 或 "贏"
            
            # 計算實際分差（客隊分數 - 主隊分數）
            actual_diff = away_score_int - home_score_int
            
            # 判斷是否符合讓分條件
            if handicap_result == '贏':
                # 客讓X分贏：需贏至少 X 分（實際分差 ≥ X）
                condition_met = actual_diff >= handicap_points
            else:  # '輸'
                # 客讓X分輸：需贏至少 (X+1) 分（實際分差 ≥ X+1）
                condition_met = actual_diff >= (handicap_points + 1)
            
            # 判斷 O/X
            if is_away_team:
                # 客隊（讓分方）：條件滿足 → O，條件不滿足 → X
                return 'O' if condition_met else 'X'
            else:
                # 主隊（非讓分方）：與客隊相反
                return 'X' if condition_met else 'O'
    
    elif home_handicap:
        # 主隊讓分（G欄）
        handicap_match = re.search(r'(\d+)分(輸|贏)', str(home_handicap))
        if handicap_match:
            handicap_points = int(handicap_match.group(1))
            handicap_result = handicap_match.group(2)  # "輸" 或 "贏"
            
            # 計算實際分差（主隊分數 - 客隊分數）
            actual_diff = home_score_int - away_score_int
            
            # 判斷是否符合讓分條件
            if handicap_result == '贏':
                # 主讓X分贏：需贏至少 X 分（實際分差 ≥ X）
                condition_met = actual_diff >= handicap_points
            else:  # '輸'
                # 主讓X分輸：需贏至少 (X+1) 分（實際分差 ≥ X+1）
                condition_met = actual_diff >= (handicap_points + 1)
            
            # 判斷 O/X
            if is_away_team:
                # 客隊（非讓分方）：與主隊相反
                return 'X' if condition_met else 'O'
            else:
                # 主隊（讓分方）：條件滿足 → O，條件不滿足 → X
                return 'O' if condition_met else 'X'
    
    return ''


def convert_date_to_column(date_str):
    """
    將日期轉換為欄位字母
    3/26 → P 欄（第16欄）
    3/27 → Q 欄（第17欄）
    以此類推
    
    Args:
        date_str: 日期字串（YYYYMMDD 格式，如 "20251022"）
    
    Returns:
        str: 欄位字母（如 "P"）
    """
    try:
        # 解析日期
        if isinstance(date_str, datetime):
            date_obj = date_str
        else:
            date_obj = datetime.strptime(str(date_str), "%Y%m%d")
        
        # 計算從 TEAM_STATS_START_DATE 開始的天數
        start_date = datetime.strptime(TEAM_STATS_START_DATE, "%Y%m%d")
        days_diff = (date_obj - start_date).days
        
        # 起始欄位預設 P（第 16 欄）
        column_num = TEAM_STATS_START_COLUMN + days_diff
        
        return get_column_letter(column_num)
    except Exception as e:
        print(f"錯誤：無法轉換日期 '{date_str}' 為欄位：{e}")
        return None


def read_games_from_daily_data(ws_daily, target_date):
    """
    從「每日數據」分頁讀取指定日期的所有比賽（一列一場 A-J）
    A=日期, B=時間, C=客分, D=客隊, E=客讓分, F=客比例, G=主分, H=主隊, I=主讓分, J=主比例
    
    Returns:
        list: 比賽資料列表，每個元素包含 {away_team, home_team, away_score,
              home_score, away_handicap, home_handicap}
    """
    games = []
    for row in range(2, ws_daily.max_row + 1):
        date_value = ws_daily[f'A{row}'].value
        if not date_value:
            continue
        if isinstance(date_value, datetime):
            date_str = date_value.strftime("%Y%m%d")
        else:
            date_str = str(date_value).strip()
        if date_str != target_date:
            continue

        away_score = ws_daily[f'C{row}'].value
        away_team = ws_daily[f'D{row}'].value
        away_handicap = ws_daily[f'E{row}'].value
        home_score = ws_daily[f'G{row}'].value
        home_team = ws_daily[f'H{row}'].value
        home_handicap = ws_daily[f'I{row}'].value

        if away_team:
            away_team = str(away_team).strip()
        if home_team:
            home_team = str(home_team).strip()

        games.append({
            'away_team': away_team,
            'home_team': home_team,
            'away_score': away_score,
            'home_score': home_score,
            'away_handicap': away_handicap,
            'home_handicap': home_handicap
        })
    return games


def convert_daily_to_team_stats(excel_file, target_date, show_progress=True):
    """
    將「每日數據」的資料轉換到「各隊正負」分頁
    
    Args:
        excel_file: Excel 檔案路徑
        target_date: 目標日期（YYYYMMDD 格式字串）
        show_progress: 是否顯示進度
    """
    try:
        # 檢查檔案是否存在
        if not os.path.exists(excel_file):
            print(f"錯誤：找不到 Excel 檔案：{excel_file}")
            print(f"請確認：")
            print(f"  1. 檔案是否存在於該路徑")
            print(f"  2. 檔案名稱是否為 'MLB26-27數據.xlsx'（或與 config.py 一致）")
            print(f"  3. 如果檔案在其他位置，請修改 config.py 中的 NBA_STATS_FILE 設定")
            return False
        
        # 開啟 Excel 檔案
        wb = openpyxl.load_workbook(excel_file)
        
        # 檢查「每日數據」分頁是否存在
        if '每日數據' not in wb.sheetnames:
            print(f"錯誤：找不到「每日數據」分頁")
            return False
        
        # 檢查「各隊正負」分頁是否存在
        if '各隊正負' not in wb.sheetnames:
            print(f"錯誤：找不到「各隊正負」分頁")
            return False
        
        ws_daily = wb['每日數據']
        ws_team = wb['各隊正負']

        # ===== 近10場（最穩定版：Python 直接計算，寫入純文字，不依賴任何 Excel 函數）=====
        # 資料範圍固定：P:IC（你已確認）
        DATA_START_COL = 'P'
        DATA_END_COL = 'IC'
        data_start_idx = column_index_from_string(DATA_START_COL)
        data_end_idx = column_index_from_string(DATA_END_COL)

        def _last10_string(row_num: int) -> str:
            vals = []
            for col_idx in range(data_start_idx, data_end_idx + 1):
                v = ws_team.cell(row=row_num, column=col_idx).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                # 你確認非空只會是 W/L 或 O/X 的單字母
                vals.append(s)
            last10 = vals[-10:]
            return ".".join(last10)
        
        # 讀取該日期的所有比賽
        games = read_games_from_daily_data(ws_daily, target_date)
        
        if not games:
            if show_progress:
                print(f"   {target_date}：沒有找到比賽資料")
            return False
        
        # 計算日期對應的欄位
        date_column = convert_date_to_column(target_date)
        if not date_column:
            print(f"錯誤：無法轉換日期 '{target_date}' 為欄位")
            return False
        
        if show_progress:
            print(f"   {target_date}：找到 {len(games)} 場比賽，寫入 {date_column} 欄")
        
        # 處理每場比賽
        for game in games:
            away_team = game['away_team']
            home_team = game['home_team']
            away_score = game['away_score']
            home_score = game['home_score']
            away_handicap = game['away_handicap']
            home_handicap = game['home_handicap']
            
            # 處理客隊
            away_team_row = get_team_row(away_team)
            if away_team_row:
                # W/L：第5行（起始行 + 4）
                wl_row = away_team_row + 4
                wl = calculate_wl(away_score, home_score, is_away_team=True)
                if wl:
                    ws_team[f'{date_column}{wl_row}'] = wl
                
                # O/X：第7行（起始行 + 6）
                ox_row = away_team_row + 6
                ox = calculate_ox_team(away_handicap, home_handicap, 
                                       away_score, home_score, is_away_team=True)
                if ox:
                    ws_team[f'{date_column}{ox_row}'] = ox
            
            # 處理主隊
            home_team_row = get_team_row(home_team)
            if home_team_row:
                # W/L：第5行（起始行 + 4）
                wl_row = home_team_row + 4
                wl = calculate_wl(away_score, home_score, is_away_team=False)
                if wl:
                    ws_team[f'{date_column}{wl_row}'] = wl
                
                # O/X：第7行（起始行 + 6）
                ox_row = home_team_row + 6
                ox = calculate_ox_team(away_handicap, home_handicap, 
                                       away_score, home_score, is_away_team=False)
                if ox:
                    ws_team[f'{date_column}{ox_row}'] = ox

        # 更新每隊「近10場」顯示（合併格左上角：F5/F7/F15/F17...）
        for team_index in range(len(MLB_TEAMS_ORDER)):
            start_row = (team_index * 10) + 1
            wl_row = start_row + 4
            ox_row = start_row + 6
            ws_team.cell(row=wl_row, column=6).value = _last10_string(wl_row)
            ws_team.cell(row=ox_row, column=6).value = _last10_string(ox_row)
        
        # 儲存檔案
        wb.save(excel_file)
        
        if show_progress:
            print(f"   {target_date}：轉換完成")
        
        return True
        
    except Exception as e:
        print(f"錯誤：轉換失敗 - {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("使用方法：python 每日數據轉各隊正負.py <日期>")
        print("日期格式：YYYYMMDD（例如：20251022）")
        sys.exit(1)
    
    target_date = sys.argv[1]
    
    print("=" * 60)
    print("每日數據轉各隊正負")
    print("=" * 60)
    print(f"目標日期：{target_date}")
    print(f"Excel 檔案：{NBA_STATS_FILE}")
    print()
    
    success = convert_daily_to_team_stats(NBA_STATS_FILE, target_date)
    
    if success:
        print()
        print("=" * 60)
        print("[完成] 轉換成功")
        print("=" * 60)
    else:
        print()
        print("=" * 60)
        print("[失敗] 轉換失敗")
        print("=" * 60)
        sys.exit(1)
