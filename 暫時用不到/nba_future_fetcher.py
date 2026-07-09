"""
NBA 明日賽事爬蟲
爬取 Playsport 明日賽事資料，並自動管理 NBA25-26.xlsm 與 NBA25-26歷史資料.xlsm
"""

import os
import sys
import time
import re
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import requests

# Selenium 導入（用於載入 JavaScript 動態內容）
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("錯誤: 未安裝 Selenium")
    print("請執行: pip install selenium")

# Win32 COM 導入（用於控制 Excel，讓標籤區域記住正確位置）
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("提示: 未安裝 pywin32，將跳過 Excel 視圖刷新功能")
    print("如需此功能，請執行: pip install pywin32")

# 引入設定檔
try:
    from config import NBA_STATS_FILE, SCRIPT_DIR
except ImportError:
    # 如果 config.py 不存在，使用預設路徑
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    NBA_STATS_FILE = os.path.join(SCRIPT_DIR, "NBA25-26.xlsm")

# 歷史資料檔案路徑（與 NBA25-26.xlsm 同目錄）
NBA_HISTORY_FILE = os.path.join(os.path.dirname(NBA_STATS_FILE), "NBA25-26歷史資料.xlsm")

# 球隊名稱對應表（網站可能顯示的變體對應到標準名稱）
TEAM_NAME_MAPPING = {
    '塞爾提': '塞爾提克',  # 網站上可能顯示為「塞爾提」（被截斷）
    '塞爾蒂克': '塞爾提克',  # 可能的變體
    '塞爾提克': '塞爾提克',  # 確保正確名稱也能對應
}

# ESPN 球隊對應表（中文隊名 -> ESPN 縮寫和完整名稱）
ESPN_TEAM_MAPPING = {
    '騎士': {'abbr': 'cle', 'name': 'cleveland-cavaliers'},
    '尼克': {'abbr': 'ny', 'name': 'new-york-knicks'},
    '塞爾提克': {'abbr': 'bos', 'name': 'boston-celtics'},
    '塞爾提': {'abbr': 'bos', 'name': 'boston-celtics'},  # 變體
    '籃網': {'abbr': 'bkn', 'name': 'brooklyn-nets'},
    '76人': {'abbr': 'phi', 'name': 'philadelphia-76ers'},
    '暴龍': {'abbr': 'tor', 'name': 'toronto-raptors'},
    '公牛': {'abbr': 'chi', 'name': 'chicago-bulls'},
    '活塞': {'abbr': 'det', 'name': 'detroit-pistons'},
    '溜馬': {'abbr': 'ind', 'name': 'indiana-pacers'},
    '公鹿': {'abbr': 'mil', 'name': 'milwaukee-bucks'},
    '老鷹': {'abbr': 'atl', 'name': 'atlanta-hawks'},
    '黃蜂': {'abbr': 'cha', 'name': 'charlotte-hornets'},
    '熱火': {'abbr': 'mia', 'name': 'miami-heat'},
    '魔術': {'abbr': 'orl', 'name': 'orlando-magic'},
    '巫師': {'abbr': 'was', 'name': 'washington-wizards'},
    '金塊': {'abbr': 'den', 'name': 'denver-nuggets'},
    '灰狼': {'abbr': 'min', 'name': 'minnesota-timberwolves'},
    '雷霆': {'abbr': 'okc', 'name': 'oklahoma-city-thunder'},
    '拓荒者': {'abbr': 'por', 'name': 'portland-trail-blazers'},
    '爵士': {'abbr': 'utah', 'name': 'utah-jazz'},  # 注意：縮寫是 'utah' 不是 'uta'
    '勇士': {'abbr': 'gs', 'name': 'golden-state-warriors'},
    '快艇': {'abbr': 'lac', 'name': 'los-angeles-clippers'},
    '湖人': {'abbr': 'lal', 'name': 'los-angeles-lakers'},
    '太陽': {'abbr': 'phx', 'name': 'phoenix-suns'},
    '國王': {'abbr': 'sac', 'name': 'sacramento-kings'},
    '獨行俠': {'abbr': 'dal', 'name': 'dallas-mavericks'},
    '火箭': {'abbr': 'hou', 'name': 'houston-rockets'},
    '灰熊': {'abbr': 'mem', 'name': 'memphis-grizzlies'},
    '鵜鶘': {'abbr': 'no', 'name': 'new-orleans-pelicans'},
    '馬刺': {'abbr': 'sa', 'name': 'san-antonio-spurs'},
}


def get_target_date():
    """獲取目標日期（今日 + 1 天），格式為 YYYYMMDD"""
    tomorrow = datetime.now() + timedelta(days=1)
    return tomorrow.strftime("%Y%m%d")


def get_yesterday_date():
    """獲取昨日日期，格式為 YYYYMMDD"""
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime("%Y%m%d")


def setup_driver(headless=False):
    """設定 Chrome 瀏覽器驅動程式"""
    if not SELENIUM_AVAILABLE:
        return None
        
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')  # 無頭模式
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception as e:
        print(f"無法啟動 Chrome 瀏覽器: {e}")
        print("\n請確保：")
        print("1. 已安裝 Chrome 瀏覽器")
        print("2. 已安裝 selenium: pip install selenium")
        print("3. ChromeDriver 會自動下載（Selenium 4.x）")
        return None


def get_page_with_selenium(url, wait_time=15):
    """使用 Selenium 載入網頁（支援 JavaScript）"""
    driver = setup_driver(headless=False)  # 顯示瀏覽器視窗以便觀察
    if not driver:
        return None
    
    try:
        print(f"正在使用 Selenium 載入網頁...")
        print(f"網址: {url}")
        driver.get(url)
        
        print(f"等待 {wait_time} 秒讓 JavaScript 執行...")
        time.sleep(wait_time)
        
        # 等待表格出現
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            print("✓ 表格已載入")
        except:
            print("⚠ 未找到表格，繼續解析...")
        
        # 獲取頁面 HTML
        html = driver.page_source
        
        # 解析 HTML
        soup = BeautifulSoup(html, 'html.parser')
        return soup
        
    except Exception as e:
        print(f"✗ Selenium 載入失敗: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        print("關閉瀏覽器...")
        driver.quit()


def fetch_playsport_data(date_str):
    """
    從 Playsport 爬取指定日期的賽事資料
    
    Args:
        date_str: 日期字串，格式為 YYYYMMDD
        
    Returns:
        list: 賽事資料列表，每個元素為一個字典
    """
    url = f"https://www.playsport.cc/predict/scale?allianceid=3&gametime={date_str}&sid=0"
    
    print(f"正在爬取資料...")
    print(f"網址: {url}")
    print()
    
    # 使用 Selenium 載入網頁（因為需要 JavaScript）
    soup = get_page_with_selenium(url, wait_time=15)
    
    if not soup:
        print("❌ 無法載入網頁")
        return []
    
    try:
        
        # 解析賽事資料（參考現有爬蟲的邏輯）
        games = []
        
        # 尋找包含對戰資訊的連結（這是關鍵標識）
        battle_links = soup.find_all('a', href=re.compile(r'/gamesData/battle'))
        print(f"找到 {len(battle_links)} 個對戰資訊連結")
        
        if not battle_links:
            print("⚠ 警告: 未找到對戰資訊連結")
            return []
        
        # 解析每場比賽
        for link_idx, link in enumerate(battle_links):
            try:
                # 找到包含此連結的表格行
                row = link.find_parent('tr')
                if not row:
                    print(f"    ⚠ 第 {link_idx + 1} 個連結：找不到包含連結的表格行")
                    continue
                
                game_data = {
                    'time': '',
                    'guest_team': '',
                    'guest_spread': '',  # 保持空白（目前無資料）
                    'guest_percentage': '',  # 保持空白（目前無資料）
                    'home_team': '',
                    'home_spread': '',  # 保持空白（目前無資料）
                    'home_percentage': '',  # 保持空白（目前無資料）
                }
                
                # 獲取該行的所有儲存格
                cells = row.find_all(['td', 'th'])
                cell_texts = [cell.get_text(separator=' ', strip=True) for cell in cells]
                row_text = ' | '.join(cell_texts)
                
                # 調試：輸出第一場比賽的 HTML 結構
                if link_idx == 0:
                    print(f"\n[調試] 第一場比賽的 HTML 結構：")
                    print(f"  行文字內容: {row_text[:200]}...")
                    print(f"  儲存格數量: {len(cells)}")
                    for i, cell in enumerate(cells[:5]):  # 只顯示前5個儲存格
                        print(f"  儲存格 {i+1}: {cell.get_text(strip=True)[:50]}")
                
                # 解析時間（格式如 "308 AM 06:00" 或 "AM 08:00"）
                time_match = re.search(r'(AM|PM)\s+(\d{2}:\d{2})', row_text, re.IGNORECASE)
                if time_match:
                    game_data['time'] = time_match.group(2)  # 只取時間部分（例如：08:00）
                else:
                    # 嘗試匹配 "308 AM 06:00" 格式
                    time_match = re.search(r'(\d{3})\s+(AM|PM)\s+(\d{2}:\d{2})', row_text, re.IGNORECASE)
                    if time_match:
                        game_data['time'] = time_match.group(3)
                    else:
                        # 如果沒有 AM/PM，嘗試直接匹配時間格式
                        time_match = re.search(r'(\d{2}:\d{2})', row_text)
                        if time_match:
                            game_data['time'] = time_match.group(1)
                
                # 根據實際網頁結構：
                # 1. 當前行（row）：包含時間和「對戰資訊」連結
                # 2. 下一行（next_row）：可能包含客隊和主隊，需要檢查所有儲存格
                # 根據調試輸出，主隊在下一行儲存格0，客隊可能在當前行或下一行的其他儲存格
                
                next_row = row.find_next_sibling('tr')
                if next_row:
                    next_cells = next_row.find_all(['td', 'th'])
                    
                    if link_idx < 3:  # 顯示前3場的調試資訊
                        print(f"  [調試] 第 {link_idx + 1} 場 - 下一行所有儲存格:")
                        for i, cell in enumerate(next_cells[:5]):  # 顯示前5個
                            print(f"    儲存格 {i}: '{cell.get_text(strip=True)[:50]}'")
                    
                    # 從下一行提取主隊（儲存格0）和客隊（其他儲存格或當前行）
                    # 根據網頁結構：下一行儲存格0是主隊
                    home_team_cell0 = ''
                    guest_team_from_next_row = []
                    
                    # 先提取儲存格0（主隊）
                    if len(next_cells) > 0:
                        cell0_text = next_cells[0].get_text(strip=True)
                        if cell0_text and '對戰' not in cell0_text and '資訊' not in cell0_text:
                            # 提取包含數字的球隊名稱
                            number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell0_text)
                            if number_match:
                                home_team_cell0 = number_match.group(0)
                                if '預測' not in home_team_cell0 and '比例' not in home_team_cell0 and '讓分' not in home_team_cell0:
                                    game_data['home_team'] = TEAM_NAME_MAPPING.get(home_team_cell0, home_team_cell0)
                                    if link_idx < 3:
                                        print(f"  [提取] 下一行儲存格0（主隊-數字）: '{home_team_cell0}'")
                            else:
                                # 提取純中文球隊名稱
                                chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell0_text)
                                common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                                valid_home = [name for name in chinese_names if name not in common_words and 2 <= len(name) <= 4 and '對戰' not in name and '資訊' not in name and '預測' not in name]
                                if valid_home:
                                    home_team_cell0 = valid_home[0]
                                    game_data['home_team'] = TEAM_NAME_MAPPING.get(home_team_cell0, home_team_cell0)
                                    if link_idx < 3:
                                        print(f"  [提取] 下一行儲存格0（主隊-中文）: '{home_team_cell0}'")
                    
                    # 從下一行其他儲存格（儲存格1以後）提取客隊
                    for cell_idx in range(1, len(next_cells)):
                        cell = next_cells[cell_idx]
                        cell_text = cell.get_text(strip=True)
                        if not cell_text:
                            continue
                        
                        # 跳過包含「對戰資訊」的儲存格
                        if '對戰' in cell_text or '資訊' in cell_text or '對戰資' in cell_text:
                            continue
                        
                        # 提取球隊名稱（不包含已經找到的主隊）
                        number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell_text)
                        if number_match:
                            team_name = number_match.group(0)
                            if '預測' not in team_name and '比例' not in team_name and '讓分' not in team_name and team_name != home_team_cell0:
                                guest_team_from_next_row.append(team_name)
                                if link_idx < 3:
                                    print(f"  [提取] 下一行儲存格 {cell_idx}（客隊-數字）: '{team_name}'")
                                break  # 找到一個就夠了
                        
                        chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell_text)
                        common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                        valid_teams = [name for name in chinese_names if name not in common_words and 2 <= len(name) <= 4 and '對戰' not in name and '資訊' not in name and '預測' not in name and name != home_team_cell0]
                        if valid_teams:
                            guest_team_from_next_row.append(valid_teams[0])
                            if link_idx < 3:
                                print(f"  [提取] 下一行儲存格 {cell_idx}（客隊-中文）: '{valid_teams[0]}'")
                            break  # 找到一個就夠了
                    
                    # 如果下一行沒找到客隊，從當前行提取客隊
                    if not game_data['guest_team']:
                        current_row_teams = []
                        current_cells = row.find_all(['td', 'th'])
                        for cell in current_cells:
                            cell_text = cell.get_text(strip=True)
                            if not cell_text:
                                continue
                            
                            # 跳過包含「對戰資訊」的儲存格
                            if '對戰資訊' in cell_text or '對戰資' in cell_text or '對戰' in cell_text or '資訊' in cell_text:
                                continue
                            
                            # 提取包含數字的球隊名稱（排除已找到的主隊）
                            number_match = re.search(r'\d+[\u4e00-\u9fff]{1,3}', cell_text)
                            if number_match:
                                team_name = number_match.group(0)
                                if ('預測' not in team_name and '比例' not in team_name and 
                                    '讓分' not in team_name and team_name != home_team_cell0):
                                    current_row_teams.append(team_name)
                                    if link_idx < 3:
                                        print(f"  [提取] 當前行儲存格（客隊-數字）: '{team_name}'")
                                    break  # 找到一個就夠了
                            
                            # 提取純中文球隊名稱（排除已找到的主隊）
                            chinese_names = re.findall(r'[\u4e00-\u9fff]{2,4}', cell_text)
                            common_words = {'對戰', '資訊', '對戰資訊', '對戰資', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過', '預測', '人預測', '主', '客'}
                            valid_teams = [name for name in chinese_names 
                                         if name not in common_words and 2 <= len(name) <= 4 
                                         and '對戰' not in name and '資訊' not in name 
                                         and '預測' not in name and name != home_team_cell0]
                            if valid_teams:
                                current_row_teams.append(valid_teams[0])
                                if link_idx < 3:
                                    print(f"  [提取] 當前行儲存格（客隊-中文）: '{valid_teams[0]}'")
                                break  # 找到一個就夠了
                        
                        if current_row_teams:
                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(current_row_teams[0], current_row_teams[0])
                    
                    if link_idx < 3:
                        print(f"  [方法1] 第 {link_idx + 1} 場 - 結果: 客隊='{game_data.get('guest_team', '')}', 主隊='{game_data.get('home_team', '')}'")
                
                # 方法2：如果下一行沒找到，嘗試從連結中提取（備用方法）
                if not game_data['guest_team'] or not game_data['home_team']:
                    team_links = row.find_all('a', href=re.compile(r'/gamesData/teams'))
                    if len(team_links) >= 2:
                        guest_team_raw = team_links[0].get_text(strip=True)
                        home_team_raw = team_links[1].get_text(strip=True)
                        
                        # 使用球隊名稱對應表
                        if not game_data['guest_team']:
                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(guest_team_raw, guest_team_raw)
                        if not game_data['home_team']:
                            game_data['home_team'] = TEAM_NAME_MAPPING.get(home_team_raw, home_team_raw)
                        
                        if link_idx == 0:
                            print(f"  [方法2] 從連結找到: 客隊='{guest_team_raw}' -> '{game_data['guest_team']}', 主隊='{home_team_raw}' -> '{game_data['home_team']}'")
                
                # 方法3：如果還是沒找到，嘗試從當前行提取（最後備用，但要小心不要提取到"對戰資訊"）
                if not game_data['guest_team'] or not game_data['home_team']:
                    # 從當前行提取所有可能的中文球隊名稱
                    all_chinese = re.findall(r'[\u4e00-\u9fff]{2,4}', row_text)
                    common_words = {'對戰', '資訊', '對戰資訊', '讓分', '比例', '主場', '客隊', '主隊', '比賽', '時間', '國際', '運彩', '大小', '不讓', '會員', '主推', '近', '日', '過'}
                    all_valid_teams = [name for name in all_chinese if name not in common_words and len(name) >= 2]
                    
                    # 額外過濾：如果包含"對戰"或"資訊"的文字，也要過濾掉
                    all_valid_teams = [name for name in all_valid_teams if '對戰' not in name and '資訊' not in name]
                    
                    # 去重但保持順序
                    seen = set()
                    unique_teams = []
                    for team in all_valid_teams:
                        if team not in seen:
                            seen.add(team)
                            unique_teams.append(team)
                    
                    if len(unique_teams) >= 2:
                        if not game_data['guest_team']:
                            game_data['guest_team'] = TEAM_NAME_MAPPING.get(unique_teams[0], unique_teams[0])
                        if not game_data['home_team']:
                            game_data['home_team'] = TEAM_NAME_MAPPING.get(unique_teams[1], unique_teams[1])
                        if link_idx < 3:
                            print(f"  [方法3] 從當前行提取: 客隊='{unique_teams[0]}', 主隊='{unique_teams[1]}'")
                
                # 最終檢查：確保沒有"對戰資訊"、"預測"等被當作球隊名稱
                invalid_keywords = ['對戰', '資訊', '對戰資訊', '預測', '人預測', '比例', '讓分']
                
                if game_data['guest_team']:
                    for keyword in invalid_keywords:
                        if keyword in game_data['guest_team']:
                            if link_idx < 3:
                                print(f"  [警告] 客隊包含無效關鍵字 '{keyword}'，清空: '{game_data['guest_team']}'")
                            game_data['guest_team'] = ''
                            break
                
                if game_data['home_team']:
                    for keyword in invalid_keywords:
                        if keyword in game_data['home_team']:
                            if link_idx < 3:
                                print(f"  [警告] 主隊包含無效關鍵字 '{keyword}'，清空: '{game_data['home_team']}'")
                            game_data['home_team'] = ''
                            break
                
                if link_idx < 3:
                    print(f"  [最終結果] 客隊='{game_data['guest_team']}', 主隊='{game_data['home_team']}'")
                
                # ========== 提取讓分和比例資料 ==========
                # 從「國際盤」的「讓分?」欄位提取讓分格式（格式：主5分輸 或 客X分輸/贏）
                # 從「國際盤」的「讓分?」欄位提取比例（39% 和 61%）
                
                # 合併當前行和下一行的文字，用於搜尋讓分和比例
                combined_text_for_spread = row_text
                if next_row:
                    next_row_text = next_row.get_text(separator=' ', strip=True)
                    combined_text_for_spread = row_text + ' | ' + next_row_text
                
                # 提取客隊讓分（格式：客X分輸 或 客X分贏）
                # 從「國際盤」的「讓分?」欄位提取，例如：客1分輸50% 或 客2分贏50%
                # 匹配：客 X分輸 或 客X分輸 或 客 X分贏 或 客X分贏
                guest_spread_match = re.search(r'客\s*(\d+)\s*分\s*(輸|贏)', combined_text_for_spread)
                if guest_spread_match:
                    spread_points = guest_spread_match.group(1)
                    spread_result = guest_spread_match.group(2)  # "輸" 或 "贏"
                    spread_value = f"{spread_points}分{spread_result}"
                    game_data['guest_spread'] = spread_value
                    if link_idx < 3:
                        print(f"  [提取] 客隊讓分（從讓分?欄位）: '{spread_value}'")
                
                # 提取主隊讓分（格式：主X分輸 或 主X分贏）
                # 從「國際盤」的「讓分?」欄位提取，例如：主5分輸50% 或 主9分輸50%
                # 匹配：主 X分輸 或 主X分輸 或 主 X分贏 或 主X分贏
                home_spread_match = re.search(r'主\s*(\d+)\s*分\s*(輸|贏)', combined_text_for_spread)
                if home_spread_match:
                    spread_points = home_spread_match.group(1)
                    spread_result = home_spread_match.group(2)  # "輸" 或 "贏"
                    spread_value = f"{spread_points}分{spread_result}"
                    game_data['home_spread'] = spread_value
                    if link_idx < 3:
                        print(f"  [提取] 主隊讓分（從讓分?欄位）: '{spread_value}'")
                
                # 提取客隊比例（從「國際盤」的「讓分?」欄位，格式：39%）
                # 尋找「讓分?」欄位中的客隊比例
                # 格式可能是：Guest (巫師): 39% 或 客 39% 或 客39%
                # 優先匹配「讓分?」欄位中的比例（不是「讓分」欄位中的50%）
                # 方法1：尋找「客」後面、「人預測」前面的百分比（這是預測比例）
                guest_percentage_match = re.search(r'客[^%]*?(\d+)%\s+\d+\s*人預測', combined_text_for_spread)
                if guest_percentage_match:
                    percentage_value = guest_percentage_match.group(1) + '%'
                    game_data['guest_percentage'] = percentage_value
                    if link_idx < 3:
                        print(f"  [提取] 客隊比例（從人預測前）: '{percentage_value}'")
                else:
                    # 方法2：尋找「客」後面的所有百分比，排除50%（讓分欄位中的），取其他百分比
                    guest_percentages = re.findall(r'客[^%]*?(\d+)%', combined_text_for_spread)
                    if guest_percentages:
                        # 過濾掉50%，取第一個非50%的百分比
                        for pct in guest_percentages:
                            if pct != '50':
                                percentage_value = pct + '%'
                                game_data['guest_percentage'] = percentage_value
                                if link_idx < 3:
                                    print(f"  [提取] 客隊比例: '{percentage_value}'")
                                break
                        # 如果都是50%，取最後一個
                        if not game_data['guest_percentage'] and guest_percentages:
                            percentage_value = guest_percentages[-1] + '%'
                            game_data['guest_percentage'] = percentage_value
                            if link_idx < 3:
                                print(f"  [提取] 客隊比例（備用）: '{percentage_value}'")
                
                # 提取主隊比例（從「國際盤」的「讓分?」欄位，格式：61%）
                # 尋找「讓分?」欄位中的主隊比例
                # 格式可能是：Home (黃蜂): 61% 或 主 61% 或 主61%
                # 優先匹配「讓分?」欄位中的比例（不是「讓分」欄位中的50%）
                # 方法1：尋找「主」後面、「人預測」前面的百分比（這是預測比例）
                home_percentage_match = re.search(r'主[^%]*?(\d+)%\s+\d+\s*人預測', combined_text_for_spread)
                if home_percentage_match:
                    percentage_value = home_percentage_match.group(1) + '%'
                    game_data['home_percentage'] = percentage_value
                    if link_idx < 3:
                        print(f"  [提取] 主隊比例（從人預測前）: '{percentage_value}'")
                else:
                    # 方法2：尋找「主」後面的所有百分比，排除50%（讓分欄位中的），取其他百分比
                    home_percentages = re.findall(r'主[^%]*?(\d+)%', combined_text_for_spread)
                    if home_percentages:
                        # 過濾掉50%，取第一個非50%的百分比
                        for pct in home_percentages:
                            if pct != '50':
                                percentage_value = pct + '%'
                                game_data['home_percentage'] = percentage_value
                                if link_idx < 3:
                                    print(f"  [提取] 主隊比例: '{percentage_value}'")
                                break
                        # 如果都是50%，取最後一個
                        if not game_data['home_percentage'] and home_percentages:
                            percentage_value = home_percentages[-1] + '%'
                            game_data['home_percentage'] = percentage_value
                            if link_idx < 3:
                                print(f"  [提取] 主隊比例（備用）: '{percentage_value}'")
                
                # 調試輸出
                if link_idx < 3:
                    print(f"  [讓分比例] 客隊讓分='{game_data['guest_spread']}', 客隊比例='{game_data['guest_percentage']}'")
                    print(f"  [讓分比例] 主隊讓分='{game_data['home_spread']}', 主隊比例='{game_data['home_percentage']}'")
                
                # 如果找到客隊和主隊，則加入列表
                if game_data['guest_team'] and game_data['home_team']:
                    games.append(game_data)
                    print(f"    ✓ 解析第 {len(games)} 場: {game_data['time']} {game_data['guest_team']} vs {game_data['home_team']}")
                else:
                    print(f"    ⚠ 跳過（缺少球隊資訊）:")
                    print(f"       時間: '{game_data['time']}'")
                    print(f"       客隊: '{game_data['guest_team']}'")
                    print(f"       主隊: '{game_data['home_team']}'")
            except Exception as e:
                print(f"    ⚠ 解析比賽時發生錯誤: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        if not games:
            print("⚠ 警告：未找到任何賽事資料")
            print("請檢查網站結構是否變更，或該日期是否有賽事")
        else:
            print(f"\n✓ 成功解析 {len(games)} 場賽事")
            print("\n" + "=" * 60)
            print("抓取到的資料詳細內容：")
            print("=" * 60)
            for idx, game in enumerate(games, 1):
                print(f"\n第 {idx} 場比賽：")
                print(f"  時間 (time): '{game.get('time', '')}'")
                print(f"  客隊 (guest_team): '{game.get('guest_team', '')}'")
                print(f"  客隊讓分 (guest_spread): '{game.get('guest_spread', '')}'")
                print(f"  客隊比例 (guest_percentage): '{game.get('guest_percentage', '')}'")
                print(f"  主隊 (home_team): '{game.get('home_team', '')}'")
                print(f"  主隊讓分 (home_spread): '{game.get('home_spread', '')}'")
                print(f"  主隊比例 (home_percentage): '{game.get('home_percentage', '')}'")
                print(f"  -> Excel 位置: B{idx * 4 + 2} = {game.get('time', '')}, C{idx * 4 + 2} = {game.get('guest_team', '')}, F{idx * 4 + 2} = {game.get('home_team', '')}")
            print("=" * 60)
            
        return games
        
    except Exception as e:
        print(f"❌ 解析錯誤: {e}")
        import traceback
        traceback.print_exc()
        return []


def fetch_espn_team_stats(team_name_chinese):
    """
    從 ESPN 抓取球隊統計數據（使用 requests，快速方法）
    
    Args:
        team_name_chinese: 中文隊名（例如：'騎士'）
        
    Returns:
        dict: 包含統計數據的字典，格式為：
        {
            'GP': '31',
            'PTS': '120',
            'OR': '12.4',
            'DR': '32.1',
            'REB': '44.5',
            'AST': '27.6',
            'STL': '8.9',
            'BLK': '5.1',
            'TO': '13.3',
            'PF': '21.3',
            'AST/TO': '2.1'
        }
        如果抓取失敗，返回 None
    """
    # 先將中文隊名標準化（處理「塞爾提」變體）
    normalized_name = TEAM_NAME_MAPPING.get(team_name_chinese, team_name_chinese)
    
    # 查找 ESPN 對應資訊
    if normalized_name not in ESPN_TEAM_MAPPING:
        print(f"  ⚠ 找不到球隊 '{team_name_chinese}' 的 ESPN 對應資訊")
        return None
    
    team_info = ESPN_TEAM_MAPPING[normalized_name]
    abbr = team_info['abbr']
    name = team_info['name']
    url = f"https://www.espn.com/nba/team/stats/_/name/{abbr}/{name}"
    
    print(f"  正在抓取 {team_name_chinese} ({abbr}) 的統計數據...")
    
    # 優先使用 requests（快速方法）
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            stats = _parse_espn_stats(soup)
            if stats:
                print(f"    ✓ 成功抓取 {team_name_chinese} 的統計數據（使用 requests）")
                return stats
            else:
                print(f"    ⚠ 無法解析 {team_name_chinese} 的統計數據")
        else:
            print(f"    ⚠ HTTP 錯誤: {response.status_code}")
    except Exception as e:
        print(f"    ⚠ requests 方法失敗: {e}")
        # 如果 requests 失敗，嘗試使用 Selenium（備用方案）
        if SELENIUM_AVAILABLE:
            try:
                print(f"    改用 Selenium 載入網頁（備用方案）...")
                soup = get_page_with_selenium(url, wait_time=10)
                if soup:
                    stats = _parse_espn_stats(soup)
                    if stats:
                        print(f"    ✓ 成功抓取 {team_name_chinese} 的統計數據（使用 Selenium）")
                        return stats
            except Exception as e2:
                print(f"    ⚠ Selenium 方法也失敗: {e2}")
    
    print(f"    ❌ 無法抓取 {team_name_chinese} 的統計數據")
    return None


def _parse_espn_stats(soup):
    """
    解析 ESPN 統計數據表格（Player Stats 的 Total 行）
    使用快速方法：優先使用 class="Table Table--align-right" 定位，直接取最後一行作為 Total 行
    
    Args:
        soup: BeautifulSoup 物件
        
    Returns:
        dict: 統計數據字典，如果解析失敗返回 None
    """
    stats = {
        'GP': '',
        'PTS': '',
        'OR': '',
        'DR': '',
        'REB': '',
        'AST': '',
        'STL': '',
        'BLK': '',
        'TO': '',
        'PF': '',
        'AST/TO': ''
    }
    
    try:
        # 方法1：優先使用 class="Table Table--align-right" 精準定位（快速方法）
        table = soup.find("table", class_="Table Table--align-right")
        
        if table:
            print(f"    [調試] 使用 class='Table Table--align-right' 找到表格")
            # 找到表格，使用快速方法
            rows = table.find_all("tr")
            if not rows or len(rows) < 2:
                # 如果沒有行或只有表頭，跳到備用方法
                table = None
            else:
                # 步驟 1: 尋找表頭行（第一行）
                header_row = None
                first_row_cells = rows[0].find_all(['th', 'td'])
                if first_row_cells:
                    header_row = [cell.get_text(strip=True) for cell in first_row_cells]
                    print(f"    [調試] 表頭欄位: {header_row[:10]}...")  # 顯示前10個欄位
                    # 確認表頭包含必要的欄位（至少需要 GP 和 PTS）
                    if 'GP' in header_row and 'PTS' in header_row:
                        # 檢查是否有 OR 和 DR（有些球隊可能沒有）
                        has_or = 'OR' in header_row
                        has_dr = 'DR' in header_row
                        if not has_or or not has_dr:
                            print(f"    [調試] ⚠ 表頭缺少 OR 或 DR: OR={has_or}, DR={has_dr}")
                        # 即使沒有 OR 或 DR，也繼續處理（其他欄位仍然可以提取）
                        # 步驟 2: 取最後一行作為 Total 行（快速方法）
                        total_row = rows[-1]
                        total_row_cells = total_row.find_all(['td', 'th'])
                        
                        if total_row_cells:
                            # 步驟 3: 提取 Total 行的數據
                            total_data_values = [cell.get_text(strip=True) for cell in total_row_cells]
                            
                            try:
                                # 找到每個欄位在表頭中的位置（索引）
                                gp_idx = header_row.index('GP')
                                pts_idx = header_row.index('PTS')
                                
                                # OR 和 DR 可能不存在，需要檢查
                                or_idx = header_row.index('OR') if 'OR' in header_row else None
                                dr_idx = header_row.index('DR') if 'DR' in header_row else None
                                
                                reb_idx = header_row.index('REB')
                                ast_idx = header_row.index('AST')
                                stl_idx = header_row.index('STL')
                                blk_idx = header_row.index('BLK')
                                to_idx = header_row.index('TO')
                                pf_idx = header_row.index('PF')
                                
                                # 提取 AST/TO（如果存在）
                                ast_to_idx = None
                                if 'AST/TO' in header_row:
                                    ast_to_idx = header_row.index('AST/TO')
                                
                                # 從 Total 行提取數據
                                stats['GP'] = total_data_values[gp_idx] if gp_idx < len(total_data_values) else ''
                                stats['PTS'] = total_data_values[pts_idx] if pts_idx < len(total_data_values) else ''
                                stats['OR'] = total_data_values[or_idx] if or_idx is not None and or_idx < len(total_data_values) else ''
                                stats['DR'] = total_data_values[dr_idx] if dr_idx is not None and dr_idx < len(total_data_values) else ''
                                stats['REB'] = total_data_values[reb_idx] if reb_idx < len(total_data_values) else ''
                                stats['AST'] = total_data_values[ast_idx] if ast_idx < len(total_data_values) else ''
                                stats['STL'] = total_data_values[stl_idx] if stl_idx < len(total_data_values) else ''
                                stats['BLK'] = total_data_values[blk_idx] if blk_idx < len(total_data_values) else ''
                                stats['TO'] = total_data_values[to_idx] if to_idx < len(total_data_values) else ''
                                stats['PF'] = total_data_values[pf_idx] if pf_idx < len(total_data_values) else ''
                                
                                # 提取 AST/TO（網頁上已有，不需要計算）
                                if ast_to_idx is not None and ast_to_idx < len(total_data_values):
                                    stats['AST/TO'] = total_data_values[ast_to_idx]
                                else:
                                    # 如果網頁上沒有 AST/TO 欄位，則計算
                                    try:
                                        if stats['AST'] and stats['TO']:
                                            ast_val = float(stats['AST'])
                                            to_val = float(stats['TO'])
                                            if to_val > 0:
                                                stats['AST/TO'] = f"{ast_val / to_val:.1f}"
                                    except:
                                        pass
                                
                                # 確認已提取到必要數據（至少 GP 和 PTS）
                                if stats['GP'] and stats['PTS']:
                                    print(f"    [調試] ✓ 成功提取數據（方法1）")
                                    return stats
                                else:
                                    print(f"    [調試] ⚠ 數據不完整: GP={stats['GP']}, PTS={stats['PTS']}")
                                    
                            except (ValueError, IndexError) as e:
                                # 如果提取失敗，繼續使用備用方法
                                print(f"    [調試] ⚠ 方法1提取失敗: {e}")
                                pass
        
        # 方法2：備用方法（如果 class 定位失敗，使用通用搜尋）
        if not stats.get('GP') or not stats.get('PTS'):
            print(f"    [調試] 方法1失敗，使用備用方法...")
            # 尋找所有表格
            tables = soup.find_all('table')
            print(f"    [調試] 找到 {len(tables)} 個表格")
            
            # 尋找 Player Stats 表格（包含 "GP" 和 "PTS" 的表格）
            for table_idx, table in enumerate(tables):
                table_text = table.get_text()
                # 確認是 Player Stats 表格
                if 'GP' in table_text and 'PTS' in table_text:
                    print(f"    [調試] 表格 #{table_idx + 1} 包含 GP 和 PTS")
                    rows = table.find_all('tr')
                    if not rows:
                        print(f"    [調試] ⚠ 表格 #{table_idx + 1} 沒有行")
                        continue
                    
                    print(f"    [調試] 表格 #{table_idx + 1} 共有 {len(rows)} 行")
                    
                    # 尋找表頭行
                    header_row = None
                    for row_idx, row in enumerate(rows):
                        cells = row.find_all(['th', 'td'])
                        if not cells:
                            continue
                        cell_texts = [cell.get_text(strip=True) for cell in cells]
                        # 至少需要 GP 和 PTS（OR 和 DR 可能不存在）
                        if 'GP' in cell_texts and 'PTS' in cell_texts:
                            header_row = cell_texts
                            print(f"    [調試] ✓ 在第 {row_idx + 1} 行找到表頭（備用方法）: {cell_texts[:5]}...")
                            break
                    
                    if not header_row:
                        print(f"    [調試] ⚠ 表格 #{table_idx + 1} 未找到完整表頭（GP, PTS, OR, DR）")
                        continue
                    
                    # 取最後一行作為 Total 行（備用方法也使用簡單邏輯）
                    last_row = rows[-1]
                    total_row_cells = last_row.find_all(['td', 'th'])
                    if total_row_cells and len(total_row_cells) >= len(header_row) - 1:
                        total_data_values = [cell.get_text(strip=True) for cell in total_row_cells]
                        
                        try:
                            # 找到每個欄位在表頭中的位置
                            gp_idx = header_row.index('GP')
                            pts_idx = header_row.index('PTS')
                            
                            # OR 和 DR 可能不存在，需要檢查
                            or_idx = header_row.index('OR') if 'OR' in header_row else None
                            dr_idx = header_row.index('DR') if 'DR' in header_row else None
                            
                            reb_idx = header_row.index('REB')
                            ast_idx = header_row.index('AST')
                            stl_idx = header_row.index('STL')
                            blk_idx = header_row.index('BLK')
                            to_idx = header_row.index('TO')
                            pf_idx = header_row.index('PF')
                            
                            stats['GP'] = total_data_values[gp_idx] if gp_idx < len(total_data_values) else ''
                            stats['PTS'] = total_data_values[pts_idx] if pts_idx < len(total_data_values) else ''
                            stats['OR'] = total_data_values[or_idx] if or_idx is not None and or_idx < len(total_data_values) else ''
                            stats['DR'] = total_data_values[dr_idx] if dr_idx is not None and dr_idx < len(total_data_values) else ''
                            stats['REB'] = total_data_values[reb_idx] if reb_idx < len(total_data_values) else ''
                            stats['AST'] = total_data_values[ast_idx] if ast_idx < len(total_data_values) else ''
                            stats['STL'] = total_data_values[stl_idx] if stl_idx < len(total_data_values) else ''
                            stats['BLK'] = total_data_values[blk_idx] if blk_idx < len(total_data_values) else ''
                            stats['TO'] = total_data_values[to_idx] if to_idx < len(total_data_values) else ''
                            stats['PF'] = total_data_values[pf_idx] if pf_idx < len(total_data_values) else ''
                            
                            # 提取 AST/TO
                            ast_to_idx = None
                            if 'AST/TO' in header_row:
                                ast_to_idx = header_row.index('AST/TO')
                            if ast_to_idx is not None and ast_to_idx < len(total_data_values):
                                stats['AST/TO'] = total_data_values[ast_to_idx]
                            else:
                                try:
                                    if stats['AST'] and stats['TO']:
                                        ast_val = float(stats['AST'])
                                        to_val = float(stats['TO'])
                                        if to_val > 0:
                                            stats['AST/TO'] = f"{ast_val / to_val:.1f}"
                                except:
                                    pass
                            
                            if stats['GP'] and stats['PTS']:
                                print(f"    [調試] ✓ 成功提取數據（方法2）")
                                return stats
                            else:
                                print(f"    [調試] ⚠ 數據不完整: GP={stats['GP']}, PTS={stats['PTS']}")
                        except (ValueError, IndexError) as e:
                            print(f"    [調試] ⚠ 提取數據時發生錯誤: {e}")
                            continue
        
    except Exception as e:
        print(f"    解析 ESPN 統計數據時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
    
    return None if not stats.get('GP') else stats


def fetch_espn_stats_for_games(games):
    """
    為所有比賽的球隊抓取 ESPN 統計數據
    
    Args:
        games: 比賽列表，每個元素包含 'guest_team' 和 'home_team'
        
    Returns:
        dict: 球隊名稱 -> 統計數據的對應字典
    """
    if not games:
        return {}
    
    print("\n" + "=" * 60)
    print("步驟 3.5: 抓取 ESPN 球隊統計數據")
    print("=" * 60)
    
    # 收集所有需要抓取的球隊（去重）
    teams_to_fetch = set()
    for game in games:
        guest_team = game.get('guest_team', '')
        home_team = game.get('home_team', '')
        if guest_team:
            teams_to_fetch.add(guest_team)
        if home_team:
            teams_to_fetch.add(home_team)
    
    print(f"需要抓取 {len(teams_to_fetch)} 支球隊的統計數據")
    print(f"球隊列表: {', '.join(sorted(teams_to_fetch))}")
    print()
    
    # 抓取每支球隊的統計數據
    team_stats = {}
    for team_name in sorted(teams_to_fetch):
        stats = fetch_espn_team_stats(team_name)
        if stats:
            team_stats[team_name] = stats
        time.sleep(1)  # 避免請求過快
    
    print(f"\n✓ 成功抓取 {len(team_stats)} 支球隊的統計數據")
    if team_stats:
        print(f"[調試] 成功抓取的球隊列表:")
        for team_name, stats in team_stats.items():
            print(f"  - {team_name}: GP={stats.get('GP', 'N/A')}, PTS={stats.get('PTS', 'N/A')}, OR={stats.get('OR', 'N/A')}, DR={stats.get('DR', 'N/A')}")
    print("=" * 60)
    
    return team_stats


def ensure_history_file():
    """確保歷史資料檔案存在，若不存在則建立"""
    if not os.path.exists(NBA_HISTORY_FILE):
        print(f"歷史資料檔案不存在，正在建立: {NBA_HISTORY_FILE}")
        try:
            ensure_xlsm_file(NBA_HISTORY_FILE)
            # 確保有 '歷史資料' 分頁（使用 keep_vba=True 以保留 VBA 程式碼）
            wb = load_workbook(NBA_HISTORY_FILE, keep_vba=True)
            if '歷史資料' not in wb.sheetnames:
                if wb.worksheets:
                    wb.worksheets[0].title = '歷史資料'
                else:
                    wb.create_sheet('歷史資料')
            wb.save(NBA_HISTORY_FILE)
            wb.close()
            print(f"✓ 已建立歷史資料檔案")
        except Exception as e:
            print(f"❌ 建立歷史資料檔案時發生錯誤: {e}")
            raise
    else:
        print(f"✓ 歷史資料檔案已存在: {NBA_HISTORY_FILE}")


def _create_xlsm_file(file_path, default_sheet_name=None):
    """
    使用 win32com 建立真正的 .xlsm 檔案
    
    Args:
        file_path: 目標檔案路徑（必須是 .xlsm 副檔名）
        default_sheet_name: 預設分頁名稱（如果為 None，則使用 'Sheet1'）
    """
    if not file_path.lower().endswith('.xlsm'):
        raise ValueError(f"檔案路徑必須是 .xlsm 格式: {file_path}")
    
    if not WIN32COM_AVAILABLE:
        raise RuntimeError(
            "無法建立 .xlsm 檔案：需要 pywin32 套件。\n"
            "請執行: pip install pywin32\n"
            "或者，請先在 Excel 中手動建立 .xlsm 檔案。"
        )
    
    try:
        print(f"  正在使用 Excel 建立 .xlsm 檔案: {file_path}")
        
        # 建立 Excel 應用程式物件
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # 建立新的工作簿
        workbook = excel.Workbooks.Add()
        
        # 設定預設分頁名稱
        if default_sheet_name:
            if workbook.Worksheets.Count > 0:
                workbook.Worksheets(1).Name = default_sheet_name
        
        # 儲存為 .xlsm 格式（xlOpenXMLWorkbookMacroEnabled = 52）
        workbook.SaveAs(os.path.abspath(file_path), FileFormat=52)
        
        # 關閉檔案
        workbook.Close(SaveChanges=False)
        
        # 關閉 Excel
        excel.Quit()
        
        # 清理 COM 物件
        del workbook
        del excel
        
        # 等待 Excel 完全釋放檔案鎖定
        import time
        time.sleep(0.5)
        
        print(f"  ✓ 已成功建立 .xlsm 檔案")
        
    except Exception as e:
        raise RuntimeError(f"建立 .xlsm 檔案時發生錯誤: {e}")


def ensure_xlsm_file(file_path):
    """
    確保 .xlsm 檔案存在且格式正確
    
    如果檔案不存在，會：
    1. 先檢查是否有對應的 .xlsx 檔案
    2. 如果有，使用 win32com 轉換為 .xlsm
    3. 如果沒有，使用 win32com 建立新的 .xlsm 檔案
    
    Args:
        file_path: 目標檔案路徑（必須是 .xlsm 副檔名）
    """
    if not file_path.lower().endswith('.xlsm'):
        raise ValueError(f"檔案路徑必須是 .xlsm 格式: {file_path}")
    
    # 檢查檔案是否已存在
    if os.path.exists(file_path):
        # 先嘗試用 openpyxl 開啟（檢查檔案是否損壞）
        try:
            wb = load_workbook(file_path, keep_vba=False)
            wb.close()
        except Exception as e:
            print(f"⚠ 檔案存在但無法用 openpyxl 開啟，將重新建立: {e}")
            # 備份舊檔案
            backup_path = file_path + '.backup_' + datetime.now().strftime('%Y%m%d_%H%M%S')
            try:
                import shutil
                shutil.copy2(file_path, backup_path)
                print(f"  已備份舊檔案到: {backup_path}")
            except:
                pass
            # 刪除問題檔案
            try:
                os.remove(file_path)
            except:
                pass
            # 繼續建立新檔案
        else:
            # openpyxl 可以開啟，但需要檢查是否真的是 .xlsm 格式（用 win32com 驗證）
            if WIN32COM_AVAILABLE:
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    
                    workbook = excel.Workbooks.Open(os.path.abspath(file_path))
                    workbook.Close(SaveChanges=False)
                    excel.Quit()
                    
                    del workbook
                    del excel
                    
                    # Excel 可以開啟，格式正確
                    import time
                    time.sleep(0.2)  # 等待釋放
                    return
                    
                except Exception as e:
                    # Excel 無法開啟，檔案格式有問題（實際上是 .xlsx 格式）
                    print(f"⚠ 檔案存在但格式不正確（實際上是 .xlsx 格式），正在修復...")
                    # 備份舊檔案
                    backup_path = file_path + '.backup_' + datetime.now().strftime('%Y%m%d_%H%M%S')
                    try:
                        import shutil
                        shutil.copy2(file_path, backup_path)
                        print(f"  已備份舊檔案到: {backup_path}")
                    except:
                        pass
                    
                    # 轉換檔案格式
                    try:
                        temp_xlsx = file_path + '.temp_convert.xlsx'
                        shutil.copy2(file_path, temp_xlsx)
                        
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                        
                        workbook = excel.Workbooks.Open(os.path.abspath(temp_xlsx))
                        workbook.SaveAs(os.path.abspath(file_path), FileFormat=52)
                        workbook.Close(SaveChanges=False)
                        excel.Quit()
                        
                        del workbook
                        del excel
                        
                        # 刪除暫存檔案
                        if os.path.exists(temp_xlsx):
                            os.remove(temp_xlsx)
                        
                        import time
                        time.sleep(0.5)
                        
                        print(f"  ✓ 檔案格式已修復為真正的 .xlsm 格式")
                        return
                    except Exception as convert_e:
                        print(f"  ⚠ 格式轉換失敗: {convert_e}")
                        # 刪除有問題的檔案，讓程式重新建立
                        try:
                            os.remove(file_path)
                        except:
                            pass
                        if os.path.exists(temp_xlsx):
                            try:
                                os.remove(temp_xlsx)
                            except:
                                pass
            else:
                # 沒有 win32com，無法驗證，假設格式正確
                return
    
    # 檢查是否有對應的 .xlsx 檔案
    xlsx_path = file_path.replace('.xlsm', '.xlsx')
    if os.path.exists(xlsx_path):
        print(f"發現舊的 .xlsx 檔案，正在轉換為 .xlsm: {xlsx_path}")
        if not WIN32COM_AVAILABLE:
            raise RuntimeError(
                "無法轉換 .xlsx 為 .xlsm：需要 pywin32 套件。\n"
                "請執行: pip install pywin32\n"
                "或者，請先在 Excel 中手動將檔案另存為 .xlsm 格式。"
            )
        
        try:
            # 使用 win32com 開啟 .xlsx 並另存為 .xlsm
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            workbook = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            workbook.SaveAs(os.path.abspath(file_path), FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
            workbook.Close(SaveChanges=False)
            excel.Quit()
            
            del workbook
            del excel
            
            # 等待 Excel 完全釋放檔案鎖定
            import time
            time.sleep(0.5)
            
            print(f"  ✓ 已成功轉換為 .xlsm 檔案")
            return
            
        except Exception as e:
            print(f"  ⚠ 轉換失敗: {e}")
            print(f"  將建立新的 .xlsm 檔案")
    
    # 建立新的 .xlsm 檔案
    _create_xlsm_file(file_path)


def refresh_excel_view(file_path):
    """
    自動開啟並關閉 Excel 檔案，讓 Excel 記住正確的標籤區域位置
    
    Args:
        file_path: Excel 檔案路徑
    """
    if not WIN32COM_AVAILABLE:
        print("⚠ 跳過 Excel 視圖刷新（未安裝 pywin32）")
        return
    
    if not os.path.exists(file_path):
        print(f"⚠ 檔案不存在，跳過視圖刷新: {file_path}")
        return
    
    try:
        print("正在刷新 Excel 視圖（讓標籤區域記住正確位置）...")
        
        # 建立 Excel 應用程式物件
        excel = win32com.client.Dispatch("Excel.Application")
        
        # 隱藏 Excel 視窗（避免閃爍）
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # 開啟檔案
        workbook = excel.Workbooks.Open(os.path.abspath(file_path))
        
        # 確保活動工作表是最左側的分頁（第一個工作表）
        if workbook.Worksheets.Count > 0:
            workbook.Worksheets(1).Activate()
        
        # 等待 Excel 完成操作（讓標籤區域更新）
        excel.Calculate()
        time.sleep(0.5)  # 額外等待，確保視圖更新
        
        # 儲存檔案（讓 Excel 記住視圖設定）
        workbook.Save()
        
        # 關閉檔案
        workbook.Close()
        
        # 關閉 Excel
        excel.Quit()
        
        # 清理 COM 物件
        del workbook
        del excel
        
        print("✓ 已完成 Excel 視圖刷新")
        
    except Exception as e:
        print(f"⚠ 刷新 Excel 視圖時發生錯誤（不影響主要功能）: {e}")


def backup_yesterday_sheet():
    """備份昨日分頁到歷史資料檔案，並從主檔案中刪除"""
    yesterday_date = get_yesterday_date()
    
    if not os.path.exists(NBA_STATS_FILE):
        print(f"主檔案不存在，跳過備份: {NBA_STATS_FILE}")
        return
    
    try:
        # 開啟主檔案（使用 keep_vba=True 以保留 VBA 程式碼）
        wb_main = load_workbook(NBA_STATS_FILE, keep_vba=True)
        
        # 檢查是否存在昨日分頁
        if yesterday_date not in wb_main.sheetnames:
            print(f"昨日分頁 '{yesterday_date}' 不存在，跳過備份")
            wb_main.close()
            return
        
        print(f"發現昨日分頁 '{yesterday_date}'，開始備份...")
        
        # 複製分頁資料
        source_sheet = wb_main[yesterday_date]
        
        # 確保歷史資料檔案存在
        ensure_history_file()
        
        # 嘗試開啟歷史資料檔案，如果損壞則重新建立
        wb_history = None
        try:
            wb_history = load_workbook(NBA_HISTORY_FILE, keep_vba=True)
        except Exception as e:
            print(f"⚠ 歷史資料檔案損壞或無法開啟: {e}")
            print(f"   正在備份損壞的檔案並重新建立...")
            
            # 備份損壞的檔案（如果存在）
            if os.path.exists(NBA_HISTORY_FILE):
                backup_path = NBA_HISTORY_FILE.replace('.xlsm', '_損壞備份_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsm')
                try:
                    import shutil
                    shutil.copy2(NBA_HISTORY_FILE, backup_path)
                    print(f"   ✓ 已備份損壞的檔案到: {backup_path}")
                except Exception as backup_e:
                    print(f"   ⚠ 無法備份損壞的檔案: {backup_e}")
            
            # 重新建立歷史資料檔案
            try:
                ensure_xlsm_file(NBA_HISTORY_FILE)
                wb_history = load_workbook(NBA_HISTORY_FILE, keep_vba=True)
                # 確保有 '歷史資料' 分頁
                if '歷史資料' not in wb_history.sheetnames:
                    if wb_history.worksheets:
                        wb_history.worksheets[0].title = '歷史資料'
                    else:
                        wb_history.create_sheet('歷史資料')
                wb_history.save(NBA_HISTORY_FILE)
                wb_history.close()
                print(f"   ✓ 已重新建立歷史資料檔案")
            except Exception as create_e:
                print(f"   ❌ 無法重新建立歷史資料檔案: {create_e}")
                print(f"   備份操作將跳過，請手動修復歷史資料檔案後重新執行")
                wb_main.close()
                return
        
        # 檢查歷史檔案中是否已存在該分頁
        if yesterday_date in wb_history.sheetnames:
            print(f"⚠ 歷史檔案中已存在分頁 '{yesterday_date}'，將覆蓋")
            wb_history.remove(wb_history[yesterday_date])
        
        # 建立新分頁
        new_sheet = wb_history.create_sheet(yesterday_date)
        
        # 複製所有資料
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection
                    new_cell.alignment = cell.alignment
        
        # 儲存歷史資料檔案
        wb_history.save(NBA_HISTORY_FILE)
        wb_history.close()
        print(f"✓ 已備份到歷史資料檔案")
        
        # 從主檔案中刪除昨日分頁
        wb_main.remove(source_sheet)
        
        # 設定活動工作表為最左側的分頁（第一個工作表）
        if wb_main.worksheets:
            wb_main.active = wb_main.worksheets[0]
            first_sheet_name = wb_main.worksheets[0].title
            print(f"✓ 已設定活動工作表為最左側的分頁: '{first_sheet_name}'")
        
        wb_main.save(NBA_STATS_FILE)
        wb_main.close()
        print(f"✓ 已從主檔案中刪除昨日分頁")
        
        # 刷新 Excel 視圖（讓標籤區域記住正確位置）
        refresh_excel_view(NBA_STATS_FILE)
        
    except Exception as e:
        print(f"❌ 備份過程發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def cleanup_old_date_sheets():
    """
    清理舊的日期分頁，只保留最近 3 天的日期分頁
    其他日期分頁移動到歷史資料檔案
    """
    try:
        print("\n" + "=" * 60)
        print("步驟 6: 清理舊的日期分頁...")
        print("=" * 60)
        
        if not os.path.exists(NBA_STATS_FILE):
            print(f"主檔案不存在，跳過清理: {NBA_STATS_FILE}")
            return
        
        # 計算最近 3 天的日期範圍（昨天、今天、明天）
        today = datetime.now()
        yesterday = (today - timedelta(days=1)).strftime("%Y%m%d")
        today_str = today.strftime("%Y%m%d")
        tomorrow = (today + timedelta(days=1)).strftime("%Y%m%d")
        
        keep_dates = {yesterday, today_str, tomorrow}
        
        print(f"當前日期: {today.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"保留的日期分頁：{yesterday}（昨天）、{today_str}（今天）、{tomorrow}（明天）")
        print(f"保留日期集合: {sorted(keep_dates)}")
        
        # 開啟主檔案（使用 keep_vba=True 以保留 VBA 程式碼）
        wb_main = load_workbook(NBA_STATS_FILE, keep_vba=True)
        
        # 找出所有日期分頁（8位數格式）
        date_sheets = []
        for sheet_name in wb_main.sheetnames:
            # 檢查是否為日期格式（8位數）
            if len(sheet_name) == 8 and sheet_name.isdigit():
                date_sheets.append(sheet_name)
        
        print(f"\n找到 {len(date_sheets)} 個日期分頁: {', '.join(sorted(date_sheets))}")
        
        # 找出需要移動的日期分頁
        sheets_to_move = [s for s in date_sheets if s not in keep_dates]
        
        # 調試：顯示每個日期分頁的判斷結果
        print(f"\n[調試] 日期分頁判斷結果：")
        for sheet_name in sorted(date_sheets):
            status = "保留" if sheet_name in keep_dates else "移動"
            print(f"  {sheet_name}: {status}")
        
        if not sheets_to_move:
            print("\n✓ 沒有需要移動的日期分頁")
            wb_main.close()
            return
        
        print(f"\n需要移動到歷史資料的日期分頁: {', '.join(sorted(sheets_to_move))}")
        print(f"共 {len(sheets_to_move)} 個分頁需要移動")
        
        # 確保歷史資料檔案存在
        ensure_history_file()
        
        # 嘗試開啟歷史資料檔案，如果損壞則重新建立
        wb_history = None
        try:
            wb_history = load_workbook(NBA_HISTORY_FILE, keep_vba=True)
            print(f"✓ 成功開啟歷史資料檔案")
        except Exception as e:
            print(f"⚠ 歷史資料檔案損壞或無法開啟: {e}")
            print(f"   正在備份損壞的檔案並重新建立...")
            
            # 備份損壞的檔案（如果存在）
            if os.path.exists(NBA_HISTORY_FILE):
                backup_path = NBA_HISTORY_FILE.replace('.xlsm', '_損壞備份_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsm')
                try:
                    import shutil
                    shutil.copy2(NBA_HISTORY_FILE, backup_path)
                    print(f"   ✓ 已備份損壞的檔案到: {backup_path}")
                except Exception as backup_e:
                    print(f"   ⚠ 無法備份損壞的檔案: {backup_e}")
            
            # 重新建立歷史資料檔案
            try:
                wb_history = Workbook()
                # Excel 檔案必須至少有一個可見的分頁，所以保留預設的 'Sheet'
                # 或者可以重新命名為更合適的名稱
                if 'Sheet' in wb_history.sheetnames:
                    wb_history['Sheet'].title = '歷史資料'
                wb_history.save(NBA_HISTORY_FILE)
                print(f"   ✓ 已重新建立歷史資料檔案")
            except Exception as create_e:
                print(f"   ❌ 無法重新建立歷史資料檔案: {create_e}")
                print(f"   清理操作將跳過，請手動修復歷史資料檔案後重新執行")
                wb_main.close()
                return
        
        moved_count = 0
        for sheet_name in sheets_to_move:
            try:
                print(f"\n處理分頁: {sheet_name}")
                
                # 從主檔案中獲取分頁
                source_sheet = wb_main[sheet_name]
                
                # 檢查歷史檔案中是否已存在該分頁
                if sheet_name in wb_history.sheetnames:
                    print(f"  ⚠ 歷史檔案中已存在分頁 '{sheet_name}'，將覆蓋")
                    wb_history.remove(wb_history[sheet_name])
                
                # 建立新分頁
                new_sheet = wb_history.create_sheet(sheet_name)
                
                # 複製所有資料和格式
                max_row = source_sheet.max_row
                max_col = source_sheet.max_column
                
                for row_idx in range(1, max_row + 1):
                    for col_idx in range(1, max_col + 1):
                        source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                        target_cell = new_sheet.cell(row=row_idx, column=col_idx)
                        
                        # 複製值
                        target_cell.value = source_cell.value
                        
                        # 複製格式（簡化處理，只複製基本格式）
                        try:
                            if source_cell.has_style:
                                if source_cell.alignment:
                                    source_align = source_cell.alignment
                                    target_cell.alignment = Alignment(
                                        horizontal=source_align.horizontal,
                                        vertical=source_align.vertical,
                                        text_rotation=source_align.text_rotation,
                                        wrap_text=source_align.wrap_text,
                                        shrink_to_fit=source_align.shrink_to_fit,
                                        indent=source_align.indent
                                    )
                                if source_cell.font:
                                    source_font = source_cell.font
                                    target_cell.font = Font(
                                        name=source_font.name,
                                        size=source_font.size,
                                        bold=source_font.bold,
                                        italic=source_font.italic,
                                        underline=source_font.underline,
                                        strike=source_font.strike,
                                        color=source_font.color
                                    )
                                if source_cell.fill and source_cell.fill.patternType:
                                    source_fill = source_cell.fill
                                    target_cell.fill = PatternFill(
                                        fill_type=source_fill.patternType,
                                        start_color=source_fill.start_color,
                                        end_color=source_fill.end_color
                                    )
                                if source_cell.number_format:
                                    target_cell.number_format = source_cell.number_format
                        except Exception:
                            # 格式複製失敗不影響數據複製
                            pass
                
                # 複製欄寬
                for col_letter in source_sheet.column_dimensions:
                    if col_letter in source_sheet.column_dimensions:
                        source_dim = source_sheet.column_dimensions[col_letter]
                        target_dim = new_sheet.column_dimensions[col_letter]
                        if source_dim.width:
                            target_dim.width = source_dim.width
                
                # 從主檔案中刪除分頁
                wb_main.remove(source_sheet)
                moved_count += 1
                print(f"  ✓ 已移動分頁 '{sheet_name}' 到歷史資料檔案")
                
            except Exception as e:
                print(f"  ❌ 移動分頁 '{sheet_name}' 時發生錯誤: {e}")
                import traceback
                traceback.print_exc()
        
        # 儲存歷史資料檔案
        try:
            wb_history.save(NBA_HISTORY_FILE)
            wb_history.close()
            print(f"\n✓ 已儲存歷史資料檔案: {moved_count} 個分頁")
        except PermissionError:
            print(f"\n❌ 無法儲存歷史資料檔案，檔案可能被其他程式開啟: {NBA_HISTORY_FILE}")
            print("   請關閉 Excel 後重新執行")
            wb_history.close()
            # 即使歷史檔案無法保存，也要嘗試保存主檔案
        except Exception as e:
            print(f"\n❌ 儲存歷史資料檔案時發生錯誤: {e}")
            wb_history.close()
        
        # 設定活動工作表為最左側的分頁
        if wb_main.worksheets:
            wb_main.active = wb_main.worksheets[0]
            first_sheet_name = wb_main.worksheets[0].title
            print(f"✓ 已設定活動工作表為最左側的分頁: '{first_sheet_name}'")
        
        # 儲存主檔案
        try:
            wb_main.save(NBA_STATS_FILE)
            wb_main.close()
            print(f"✓ 已儲存主檔案: 已移除 {moved_count} 個舊日期分頁")
            
            # 刷新 Excel 視圖
            refresh_excel_view(NBA_STATS_FILE)
        except PermissionError:
            print(f"\n❌ 無法儲存主檔案，檔案可能被其他程式開啟: {NBA_STATS_FILE}")
            print("   請關閉 Excel 後重新執行")
            print(f"   ⚠ 注意：已從記憶體中移除 {moved_count} 個分頁，但未保存到檔案")
            wb_main.close()
        except Exception as e:
            print(f"\n❌ 儲存主檔案時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            wb_main.close()
        
    except Exception as e:
        print(f"❌ 清理日期分頁時發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def update_analysis_sheet_references(wb):
    """
    更新三個分析分頁的 A1 單元格，寫入對應的日期分頁名稱
    用於動態引用日期分頁的資料（例如：=INDIRECT(A1&"!B2")）
    
    Args:
        wb: openpyxl Workbook 物件
    """
    try:
        print("\n" + "=" * 60)
        print("步驟 5: 更新分析分頁的日期引用...")
        print("=" * 60)
        
        # 計算三個日期（對應的日期分頁名稱）
        today = datetime.now()
        yesterday_date = (today - timedelta(days=1)).strftime("%Y%m%d")  # 昨天的日期
        today_date = today.strftime("%Y%m%d")  # 今天的日期
        tomorrow_date = (today + timedelta(days=1)).strftime("%Y%m%d")  # 明天的日期
        
        analysis_sheets = {
            '昨日分析': yesterday_date,
            '今日分析': today_date,
            '明日分析': tomorrow_date
        }
        
        print(f"日期對應關係：")
        print(f"  昨日分析 -> A1 = '{yesterday_date}'（昨天的日期）")
        print(f"  今日分析 -> A1 = '{today_date}'（今天的日期）")
        print(f"  明日分析 -> A1 = '{tomorrow_date}'（明天的日期）")
        
        for analysis_sheet_name, date_str in analysis_sheets.items():
            print(f"\n處理分頁: {analysis_sheet_name}")
            
            # 檢查分析分頁是否存在，如果不存在則創建
            if analysis_sheet_name not in wb.sheetnames:
                print(f"  ⚠ 分頁 '{analysis_sheet_name}' 不存在，正在創建...")
                analysis_ws = wb.create_sheet(analysis_sheet_name)
            else:
                analysis_ws = wb[analysis_sheet_name]
            
            # 檢查對應的日期分頁是否存在
            if date_str in wb.sheetnames:
                # 日期分頁存在，寫入日期字串到 A1
                analysis_ws['A1'] = date_str
                print(f"  ✓ A1 = '{date_str}'（日期分頁存在）")
            else:
                # 日期分頁不存在，寫入提示文字
                analysis_ws['A1'] = '尚未創建'
                print(f"  ⚠ A1 = '尚未創建'（日期分頁 '{date_str}' 不存在）")
        
        print(f"\n✓ 已完成三個分析分頁的 A1 更新")
        print(f"  提示：您可以在分析分頁中使用 =INDIRECT($A$1&\"!B2\") 來動態引用日期分頁的資料")
        
    except Exception as e:
        print(f"  ❌ 更新分析分頁引用時發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def write_to_excel(games, date_str, team_stats=None):
    """
    將賽事資料寫入 Excel
    
    Args:
        games: 賽事資料列表（可能為空）
        date_str: 日期字串，格式為 YYYYMMDD
        team_stats: 球隊統計數據字典（可選），格式為 {球隊名稱: {GP: ..., PTS: ..., ...}}
    """
    
    try:
        # 確保主檔案存在且格式正確
        if not os.path.exists(NBA_STATS_FILE):
            print(f"主檔案不存在，正在建立: {NBA_STATS_FILE}")
            try:
                ensure_xlsm_file(NBA_STATS_FILE)
            except RuntimeError as e:
                print(f"❌ {e}")
                raise
        
        # 開啟主檔案（使用 keep_vba=True 以保留 VBA 程式碼）
        wb = load_workbook(NBA_STATS_FILE, keep_vba=True)
        
        # 檢查分頁是否已存在
        if date_str in wb.sheetnames:
            print(f"⚠ 分頁 '{date_str}' 已存在，將覆蓋")
            wb.remove(wb[date_str])
        
        # 建立新分頁
        ws = wb.create_sheet(date_str)
        
        # 將新分頁移動到最前面（最左側）
        # 獲取所有分頁列表
        sheets = wb._sheets
        # 找到新建立的分頁索引
        new_sheet_index = sheets.index(ws)
        # 從列表中移除
        sheets.pop(new_sheet_index)
        # 插入到最前面（索引 0）
        sheets.insert(0, ws)
        print(f"✓ 已將分頁 '{date_str}' 移動到最前面")
        
        # 如果沒有比賽資料，寫入「本日沒有比賽」
        if not games:
            print(f"\n⚠ 沒有賽事資料，將寫入「本日沒有比賽」")
            # 在 B2 位置寫入「本日沒有比賽」（對應第一場比賽時間的位置）
            ws['B2'] = '本日沒有比賽'
            # 設定對齊方式（置中對齊）
            ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
            print(f"  B2 = '本日沒有比賽'")
        else:
            # 寫入資料
            print(f"\n開始寫入 {len(games)} 場賽事到 Excel...")
            
            # 調試：顯示 team_stats 的狀態
            if team_stats:
                print(f"\n[調試] team_stats 字典中有 {len(team_stats)} 支球隊的數據:")
                for team_name, stats in team_stats.items():
                    print(f"  - {team_name}: GP={stats.get('GP', '')}, PTS={stats.get('PTS', '')}")
            else:
                print(f"\n[調試] ⚠ team_stats 為空，無法寫入 ESPN 數據")
            
            for idx, game in enumerate(games):
                # 計算起始列：第 N 場資料行 = (N-1)*4 + 2
                # 第1場：資料在 B2~H2，標題在 B1~H1
                # 第2場：資料在 B6~H6，標題在 B5~H5
                # 第3場：資料在 B10~H10，標題在 B9~H9
                data_row = idx * 4 + 2  # 資料行（B2, B6, B10, ...）
                title_row = data_row - 1  # 標題行（B1, B5, B9, ...）
                
                print(f"\n寫入第 {idx + 1} 場比賽（資料行: {data_row}, 標題行: {title_row}）:")
                
                # 標題行（B1, B16, B31, ...）
                ws[f'B{title_row}'] = '比賽時間'
                ws[f'C{title_row}'] = '客隊'
                ws[f'D{title_row}'] = '讓分'
                ws[f'E{title_row}'] = '比例'
                ws[f'F{title_row}'] = '主場'
                ws[f'G{title_row}'] = '讓分'
                ws[f'H{title_row}'] = '比例'
                print(f"  標題行 {title_row}: B{title_row}=比賽時間, C{title_row}=客隊, D{title_row}=讓分, E{title_row}=比例, F{title_row}=主場, G{title_row}=讓分, H{title_row}=比例")
                
                # 資料行（B2, B17, B32, ...）
                # B 欄：比賽時間
                time_value = game.get('time', '')
                ws[f'B{data_row}'] = time_value
                print(f"  B{data_row} = '{time_value}' (比賽時間)")
                
                # C 欄：客隊名稱
                guest_team_value = game.get('guest_team', '')
                ws[f'C{data_row}'] = guest_team_value
                print(f"  C{data_row} = '{guest_team_value}' (客隊)")
                
                # D 欄：客隊讓分（只顯示讓分方的隊伍，受讓方留空）
                guest_spread_value = game.get('guest_spread', '')
                # 如果客隊有讓分，顯示讓分；否則留空
                if guest_spread_value:
                    ws[f'D{data_row}'] = guest_spread_value
                    print(f"  D{data_row} = '{guest_spread_value}' (客隊讓分)")
                else:
                    ws[f'D{data_row}'] = ''  # 受讓方留空
                    print(f"  D{data_row} = '' (客隊無讓分，留空)")
                
                # E 欄：客隊預測比例
                guest_percentage_value = game.get('guest_percentage', '')
                ws[f'E{data_row}'] = guest_percentage_value
                print(f"  E{data_row} = '{guest_percentage_value}' (客隊比例)")
                
                # F 欄：主隊名稱
                home_team_value = game.get('home_team', '')
                ws[f'F{data_row}'] = home_team_value
                print(f"  F{data_row} = '{home_team_value}' (主隊)")
                
                # G 欄：主隊讓分（只顯示讓分方的隊伍，受讓方留空）
                home_spread_value = game.get('home_spread', '')
                # 如果主隊有讓分，顯示讓分；否則留空
                if home_spread_value:
                    ws[f'G{data_row}'] = home_spread_value
                    print(f"  G{data_row} = '{home_spread_value}' (主隊讓分)")
                else:
                    ws[f'G{data_row}'] = ''  # 受讓方留空
                    print(f"  G{data_row} = '' (主隊無讓分，留空)")
                
                # H 欄：主隊預測比例
                home_percentage_value = game.get('home_percentage', '')
                ws[f'H{data_row}'] = home_percentage_value
                print(f"  H{data_row} = '{home_percentage_value}' (主隊比例)")
                
                # 設定對齊方式
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    # 標題行（置中對齊）
                    title_cell = ws[f'{col}{title_row}']
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 資料行（左對齊）
                    data_cell = ws[f'{col}{data_row}']
                    data_cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 寫入 ESPN 統計數據（I~U 欄位）
                if team_stats:
                    print(f"  [調試] team_stats 不為空，開始寫入 ESPN 數據")
                    # 寫入 ESPN 標題行（只在第一場比賽時寫入）
                    if idx == 0:
                        espn_headers = {
                            'I': '客/主',
                            'J': '名稱',
                            'K': 'GP(以出賽)',
                            'L': 'PTS(得分)',
                            'M': 'OR(命中率)',
                            'N': 'DR(籃板率)',
                            'O': 'REB(籃板球)',
                            'P': 'AST(助攻)',
                            'Q': 'STL(抄截)',
                            'R': 'BLK(火鍋)',
                            'S': 'TO(失誤)',
                            'T': 'PF(犯規)',
                            'U': 'AST/TO(貢獻度)'
                        }
                        for col, header in espn_headers.items():
                            ws[f'{col}{title_row}'] = header
                            ws[f'{col}{title_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        print(f"  已寫入 ESPN 標題行 {title_row}: I{title_row}~U{title_row}")
                    
                    # 寫入客隊統計數據（I2, J2, K2, ...）
                    guest_team = game.get('guest_team', '')
                    print(f"  [調試] 檢查客隊 '{guest_team}' 是否在 team_stats 中: {guest_team in team_stats if guest_team else False}")
                    if guest_team and guest_team in team_stats:
                        guest_stats = team_stats[guest_team]
                        ws[f'I{data_row}'] = '客隊'
                        ws[f'J{data_row}'] = guest_team
                        ws[f'K{data_row}'] = guest_stats.get('GP', '')
                        ws[f'L{data_row}'] = guest_stats.get('PTS', '')
                        ws[f'M{data_row}'] = guest_stats.get('OR', '')
                        ws[f'N{data_row}'] = guest_stats.get('DR', '')
                        ws[f'O{data_row}'] = guest_stats.get('REB', '')
                        ws[f'P{data_row}'] = guest_stats.get('AST', '')
                        ws[f'Q{data_row}'] = guest_stats.get('STL', '')
                        ws[f'R{data_row}'] = guest_stats.get('BLK', '')
                        ws[f'S{data_row}'] = guest_stats.get('TO', '')
                        ws[f'T{data_row}'] = guest_stats.get('PF', '')
                        ws[f'U{data_row}'] = guest_stats.get('AST/TO', '')
                        print(f"  已寫入客隊 {guest_team} 的 ESPN 統計數據: I{data_row}~U{data_row}")
                    
                    # 寫入主隊統計數據（I3, J3, K3, ...）
                    home_team = game.get('home_team', '')
                    print(f"  [調試] 檢查主隊 '{home_team}' 是否在 team_stats 中: {home_team in team_stats if home_team else False}")
                    if home_team and home_team in team_stats:
                        home_stats = team_stats[home_team]
                        home_data_row = data_row + 1  # 主隊在下一行
                        ws[f'I{home_data_row}'] = '主隊'
                        ws[f'J{home_data_row}'] = home_team
                        ws[f'K{home_data_row}'] = home_stats.get('GP', '')
                        ws[f'L{home_data_row}'] = home_stats.get('PTS', '')
                        ws[f'M{home_data_row}'] = home_stats.get('OR', '')
                        ws[f'N{home_data_row}'] = home_stats.get('DR', '')
                        ws[f'O{home_data_row}'] = home_stats.get('REB', '')
                        ws[f'P{home_data_row}'] = home_stats.get('AST', '')
                        ws[f'Q{home_data_row}'] = home_stats.get('STL', '')
                        ws[f'R{home_data_row}'] = home_stats.get('BLK', '')
                        ws[f'S{home_data_row}'] = home_stats.get('TO', '')
                        ws[f'T{home_data_row}'] = home_stats.get('PF', '')
                        ws[f'U{home_data_row}'] = home_stats.get('AST/TO', '')
                        print(f"  已寫入主隊 {home_team} 的 ESPN 統計數據: I{home_data_row}~U{home_data_row}")
                    
                    # 設定 ESPN 數據的對齊方式
                    for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                        if guest_team and guest_team in team_stats:
                            ws[f'{col}{data_row}'].alignment = Alignment(horizontal='left', vertical='center')
                        if home_team and home_team in team_stats:
                            ws[f'{col}{data_row + 1}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 更新三個分析分頁的 A1（昨日分析、今日分析、明日分析）
        update_analysis_sheet_references(wb)
        
        # 設定活動工作表為最左側的分頁（第一個工作表）
        if wb.worksheets:
            wb.active = wb.worksheets[0]
            first_sheet_name = wb.worksheets[0].title
            print(f"✓ 已設定活動工作表為最左側的分頁: '{first_sheet_name}'")
        
        # 儲存檔案
        wb.save(NBA_STATS_FILE)
        wb.close()
        if games:
            print(f"✓ 已成功寫入 {len(games)} 場賽事到分頁 '{date_str}'")
        else:
            print(f"✓ 已成功建立分頁 '{date_str}'（本日沒有比賽）")
        
        # 刷新 Excel 視圖（讓標籤區域記住正確位置）
        refresh_excel_view(NBA_STATS_FILE)
        
    except PermissionError as e:
        print(f"\n❌ 無法寫入檔案（檔案可能正在被 Excel 開啟）")
        print(f"   檔案路徑: {NBA_STATS_FILE}")
        print(f"\n請執行以下步驟：")
        print(f"   1. 關閉所有 Excel 視窗（包括 {os.path.basename(NBA_STATS_FILE)}）")
        print(f"   2. 在工作管理員中確認沒有 EXCEL.EXE 程序在執行")
        print(f"   3. 重新執行爬蟲程式")
        print(f"\n錯誤詳情: {e}")
    except Exception as e:
        print(f"❌ 寫入 Excel 時發生錯誤: {e}")
        import traceback
        traceback.print_exc()


def main():
    """主程式"""
    print("=" * 60)
    print("NBA 明日賽事爬蟲")
    print("=" * 60)
    print()
    
    # 獲取目標日期
    target_date = get_target_date()
    print(f"目標日期: {target_date}")
    print()
    
    # 確保歷史資料檔案存在
    print("步驟 1: 檢查歷史資料檔案...")
    ensure_history_file()
    print()
    
    # 備份昨日分頁
    print("步驟 2: 檢查並備份昨日分頁...")
    backup_yesterday_sheet()
    print()
    
    # 爬取資料
    print("步驟 3: 爬取明日賽事資料...")
    games = fetch_playsport_data(target_date)
    print()
    
    # 抓取 ESPN 統計數據（步驟 3.5）
    team_stats = {}
    if games:
        team_stats = fetch_espn_stats_for_games(games)
        print()
    
    # 寫入 Excel（即使沒有比賽也會建立分頁）
    print("步驟 4: 寫入 Excel...")
    write_to_excel(games, target_date, team_stats)
    print()
    
    # 清理舊的日期分頁（只保留最近 3 天）
    cleanup_old_date_sheets()
    print()
    
    print("=" * 60)
    print("執行完成！")
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n程式已中斷")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ 發生未預期的錯誤: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

