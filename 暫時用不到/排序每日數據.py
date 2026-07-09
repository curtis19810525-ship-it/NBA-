"""
手動執行「每日數據」分頁的日期排序（舊到新，一列一場 A-J）
同日期內再依 B 欄比賽時間由早到晚（與 24h 寫入一致，避免 12:10 在末段）
"""

import os
import re
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font

# 引入設定檔
try:
    from config import NBA_XLSX_FILE
except ImportError:
    NBA_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"


def _daily_b_time_to_minutes(b_val):
    """B 欄比賽時間轉成從 00:00 起分鐘數，供同日期內排序。"""
    if b_val is None or b_val == "":
        return 99999
    if isinstance(b_val, datetime):
        return b_val.hour * 60 + b_val.minute
    s = str(b_val).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return 99999
    return int(m.group(1)) * 60 + int(m.group(2))


def sort_daily_data_by_date(ws):
    """
    對「每日數據」分頁按日期排序（舊到新），一列一場，A-J，跨日空一列
    注意：此函數會重新寫入整個分頁，請確保檔案已備份
    """
    try:
        max_rows_to_process = 5000
        data_rows = []  # (date_str, row_data) 每筆為一列 A-J

        for row in range(2, min(ws.max_row + 1, max_rows_to_process + 1)):
            date_value = ws[f'A{row}'].value
            if not date_value:
                continue
            date_str = date_value.strftime("%Y%m%d") if isinstance(date_value, datetime) else str(date_value).strip()

            row_data = []
            for col in range(1, 11):  # A-J
                cell = ws.cell(row=row, column=col)
                alignment_info = None
                if cell.alignment:
                    try:
                        alignment_info = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    except Exception:
                        alignment_info = Alignment(horizontal="center", vertical="center")
                font_info = None
                if cell.font:
                    try:
                        font_info = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            color=cell.font.color
                        )
                    except Exception:
                        font_info = None
                row_data.append({
                    'value': cell.value,
                    'alignment': alignment_info,
                    'number_format': cell.number_format,
                    'font': font_info
                })
            data_rows.append((date_str, row_data))

        if not data_rows:
            print("⚠ 未找到任何資料列，跳過排序")
            return

        def _row_b_minutes(row_d):
            if not row_d or len(row_d) < 2:
                return 99999
            return _daily_b_time_to_minutes(row_d[1].get("value"))

        data_rows.sort(
            key=lambda x: (x[0] or "", _row_b_minutes(x[1]))
        )

        for row in range(2, min(ws.max_row + 1, 10000)):
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.value = None
                cell.border = Border()

        current_row = 2
        prev_date = None
        for date_str, row_data in data_rows:
            if prev_date and prev_date != date_str:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).border = Border()
                current_row += 1
            for col_idx, cell_data in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                if cell_data.get('value') is not None:
                    cell.value = cell_data['value']
                if cell_data.get('alignment'):
                    try:
                        cell.alignment = cell_data['alignment']
                    except Exception:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell_data.get('number_format'):
                    cell.number_format = cell_data['number_format']
                if cell_data.get('font'):
                    try:
                        cell.font = cell_data['font']
                    except Exception:
                        pass
                cell.border = Border()
            current_row += 1
            prev_date = date_str

        print("✓ 已按日期（舊到新）與同日期內比賽時間（早到晚）排序")
        
    except Exception as e:
        print(f"⚠ 排序時發生錯誤：{e}")
        import traceback
        traceback.print_exc()
        raise  # 重新拋出異常，讓主函數處理


def main():
    """主函數"""
    print("=" * 60)
    print("排序「每日數據」分頁（日期舊到新；同日依比賽時間早到晚）")
    print("=" * 60)
    
    excel_file = NBA_XLSX_FILE
    
    if not os.path.exists(excel_file):
        print(f"✗ 檔案不存在：{excel_file}")
        sys.exit(1)
    
    # 檢查檔案是否被鎖定
    try:
        test_file = open(excel_file, 'r+b')
        test_file.close()
    except PermissionError:
        print(f"✗ 檔案正在被其他程式使用：{excel_file}")
        print(f"  請關閉 Excel 或其他正在使用此檔案的程式後再試")
        sys.exit(1)
    
    try:
        print(f"\n正在讀取檔案：{excel_file}")
        wb = load_workbook(excel_file, read_only=False)
        
        # 檢查「每日數據」分頁
        if "每日數據" not in wb.sheetnames:
            print("✗ 找不到「每日數據」分頁")
            wb.close()
            sys.exit(1)
        
        ws = wb["每日數據"]
        print(f"✓ 找到「每日數據」分頁")
        print(f"  目前總行數：{ws.max_row}")
        
        # 執行排序
        print("\n開始執行排序...")
        sort_daily_data_by_date(ws)
        
        # 儲存檔案
        print("\n正在儲存檔案...")
        wb.save(excel_file)
        print(f"✓ 排序完成並已儲存")
        
        wb.close()
        
        print("\n" + "=" * 60)
        print("✓ 執行完成！")
        print("=" * 60)
        
    except KeyError as e:
        error_msg = str(e)
        if "[Content_Types].xml" in error_msg or "archive" in error_msg.lower():
            print(f"✗ Excel 檔案損壞或格式異常：{excel_file}")
            print(f"  錯誤訊息：{error_msg}")
            print(f"\n  建議：請從備份還原檔案")
        else:
            print(f"✗ 讀取 Excel 檔案時發生錯誤：{error_msg}")
        sys.exit(1)
    except PermissionError:
        print(f"✗ 無法儲存檔案：檔案可能正在被其他程式使用")
        print(f"  請關閉 Excel 後再試")
        sys.exit(1)
    except Exception as e:
        error_msg = str(e)
        print(f"✗ 執行時發生錯誤：{error_msg}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
