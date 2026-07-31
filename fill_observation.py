# -*- coding: utf-8 -*-
"""
填入盤口觀察

把「玖九盤口變化 → 比賽結果／MLB快照」的場次與盤口頭尾，
自動填入 OneDrive 的「盤口觀察.xlsx → 工作表1」。
若 MLB26-27「紀錄」同日有對應讓分球隊，再補金流／量化（沒有則留空）。
賽前：紀錄有場次即可建列並填頭盤；比賽結果有資料再填 H／M／R。

規則（與使用者討論定案）：
  - 場次來源：玖九「比賽結果」（不再因「紀錄」缺日而整天 0 場）
  - 吃日期區間：python fill_observation.py YYYYMMDD YYYYMMDD（單日可兩個都填同一天；結束早於開始須重輸、不對調）
  - 模型欄、頭、結果欄：逐格只填空白、不覆蓋（含手動修正）
  - 尾欄（G/J/L/O/Q）：每次重跑以玖九最新「開賽前最後快照」覆蓋更新
  - 舊日期：找到 (日期+讓分球隊) 的列補空格；新日期：一場一列接在表尾
  - 頭 = 玖九該對戰第 1 筆快照；尾 = 開賽前最後 1 筆快照
  - 結果欄 H/M/R 用「對照表」翻譯；對照表沒有的字串不亂填、會列出來

寫入前自動備份盤口觀察.xlsx 至「回朔點備分」，僅保留最近 5 份。
獨立執行時會取得管線鎖；由 run_all 呼叫時請加 --no-lock。
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries

from analysis_helpers import (
    _norm_date,
    _norm_time,
    all_pregame_snapshots,
    find_jiujiu_game_by_fav,
    find_result_row,
    kickoff_matches_focus,
    read_daily_games,
    read_result_rows,
    read_snapshot_rows,
    team_nickname,
)
from config import JIUJIU_XLSX_FILE, MLB_XLSX_FILE, OBSERVATION_XLSX_FILE
from odds_format import decimal_to_jiujiu_line, format_odds

BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "回朔點備分")
OBSERVATION_BACKUP_KEEP = int(os.environ.get("OBSERVATION_BACKUP_KEEP", "5"))

SHEET_RECORD = "紀錄"
# 紀錄分頁的直欄組起始欄（每 10 欄一組）：A、K、U、AE …
RECORD_BLOCK_STEP = 10


def record_block_cols(ws) -> tuple[int, ...]:
    """依工作表寬度列出所有直欄組起始欄（1, 11, 21, 31 …）。"""
    max_c = max(getattr(ws, "max_column", 1) or 1, 1)
    return tuple(range(1, max_c + 1, RECORD_BLOCK_STEP))

# 盤口觀察「工作表1」欄位（1-indexed）
COL_DATE = 1        # A 日期
COL_TEAM = 2        # B 讓分球隊
COL_FLOW = 3        # C 金流
COL_SIDE = 4        # D 主/客
COL_QUANT = 5       # E 量化（紀錄的合理性）
COL_SP_HEAD = 6     # F 讓分(頭)
COL_SP_TAIL = 7     # G 讓分(尾)
COL_SP_RESULT = 8   # H 結果（讓分 O/X）
COL_FAVML_HEAD = 9  # I 讓分獨贏(頭)
COL_FAVML_TAIL = 10 # J 讓分獨贏(尾)
COL_RECML_HEAD = 11 # K 受讓獨贏(頭)
COL_RECML_TAIL = 12 # L 受讓獨贏(尾)
COL_ML_RESULT = 13  # M 獨贏結果（讓/受）
COL_FAV2_HEAD = 14  # N 讓2分贏(頭)
COL_FAV2_TAIL = 15  # O 讓2分贏(尾)
COL_REC2_HEAD = 16  # P 受讓2分贏(頭)
COL_REC2_TAIL = 17  # Q 受讓2分贏(尾)
COL_TWO_RESULT = 18 # R 2分結果（O/X）

# 尾欄：重跑時以最新玖九快照覆蓋（G/J/L/O/Q）
TAIL_COLS = (
    COL_SP_TAIL,
    COL_FAVML_TAIL,
    COL_RECML_TAIL,
    COL_FAV2_TAIL,
    COL_REC2_TAIL,
)

# 用來核對標題的關鍵字（避免欄位錯位就寫入）
HEADER_CHECKS = {
    COL_DATE: "日期",
    COL_TEAM: "讓分球隊",
    COL_SP_HEAD: "讓分",
    COL_ML_RESULT: "獨贏結果",
}

# ===== 結果欄對照表（H/M/R）=====
# H 讓分結果：比賽結果「讓分結算」字串 → O/X
SPREAD_SETTLE_MAP = {
    "過": "O",
    "沒過": "X",
}
# R 2分結果：比賽結果「兩分贏結果」字串 → O/X（先判受讓再判讓，避免子字串誤判）
TWO_RESULT_RULES = (
    ("受讓2分", "X"),
    ("讓2分", "O"),
)


def ensure_excel_files_closed(paths: list[str]) -> None:
    """偵測檔案被占用（多半 Excel 未關閉）；有占用則中止並提示先存檔關閉。"""
    locked: list[str] = []
    for path in paths:
        if not path or not os.path.exists(path):
            continue
        try:
            with open(path, "r+b"):
                pass
        except PermissionError:
            locked.append(path)
        except OSError as e:
            winerr = getattr(e, "winerror", None)
            if winerr == 32 or e.errno in (11, 13, 16):
                locked.append(path)
    if locked:
        lines = "\n".join(f"  - {p}" for p in locked)
        raise SystemExit(
            "偵測到以下檔案仍被占用（請先在 Excel「存檔」並關閉後再執行）：\n"
            f"{lines}"
        )


def parse_yyyymmdd(s: str) -> datetime:
    return datetime.strptime(str(s or "").strip(), "%Y%m%d")


def prompt_date(prompt_text: str) -> str:
    while True:
        try:
            value = input(prompt_text).strip()
        except (EOFError, KeyboardInterrupt):
            print("\n已取消。")
            raise SystemExit(1)
        try:
            parse_yyyymmdd(value)
            return value
        except ValueError:
            print("日期格式錯誤，請使用 YYYYMMDD（例如 20251022）。")


def prompt_date_range_interactive() -> tuple[datetime, datetime]:
    """兩行起迄；結束早於開始則整段重輸（不對調）。"""
    while True:
        start_str = prompt_date("請輸入開始日期 (YYYYMMDD): ")
        end_str = prompt_date("請輸入結束日期 (YYYYMMDD): ")
        start = parse_yyyymmdd(start_str)
        end = parse_yyyymmdd(end_str)
        if end < start:
            print("錯誤：結束日期早於開始日期，請重新輸入起迄兩行。")
            continue
        return start, end


def parse_cli_date_range(start_str: str, end_str: str) -> tuple[datetime, datetime]:
    start = parse_yyyymmdd(start_str)
    end = parse_yyyymmdd(end_str)
    if end < start:
        raise SystemExit(
            "錯誤：結束日期早於開始日期。請重新執行並輸入正確起迄（不會自動對調）。"
        )
    return start, end


def daterange(start: datetime, end: datetime):
    cur = start
    while cur <= end:
        yield cur.strftime("%Y%m%d")
        cur += timedelta(days=1)


def parse_team_venue(text: str) -> tuple[str, str]:
    """'勇士主場' → ('勇士', '主')；'水手客場' → ('水手', '客')。"""
    t = str(text or "").strip()
    if t.endswith("主場"):
        return t[:-2].strip(), "主"
    if t.endswith("客場"):
        return t[:-2].strip(), "客"
    return t, ""


def _cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def read_record_games(ws, focus_date: str) -> list[dict]:
    """從『紀錄』分頁讀出某日所有場次的模型欄。"""
    games, _diag = read_record_games_with_diag(ws, focus_date)
    return games


def read_record_games_with_diag(ws, focus_date: str) -> tuple[list[dict], dict]:
    """回傳 (games, 診斷)。診斷用於查無資料時說明是「找不到日期」還是「日期有但解析不到場次」。"""
    diag: dict = {
        "focus_date": focus_date,
        "date_found": False,
        "date_row": None,
        "date_col": None,
        "a1_raw": ws.cell(1, 1).value,
        "a1_norm": _norm_date(ws.cell(1, 1).value),
        "a1_type": type(ws.cell(1, 1).value).__name__,
        "labels_seen": [],
        "parsed_games": 0,
        "skipped_no_team": 0,
    }
    game_label_re = re.compile(r"第\s*\d+\s*場")

    for dc in record_block_cols(ws):
        date_row = None
        for r in range(1, (ws.max_row or 1) + 1):
            if _norm_date(ws.cell(r, dc).value) == focus_date:
                date_row = r
                break
        if date_row is None:
            continue

        diag["date_found"] = True
        diag["date_row"] = date_row
        diag["date_col"] = dc

        games: list[dict] = []
        r = date_row + 1
        blanks = 0
        while r <= (ws.max_row or 1):
            a = ws.cell(r, dc).value
            cell_date = _norm_date(a)
            if cell_date and cell_date != focus_date:
                break
            label = _cell_str(a)
            if label and len(diag["labels_seen"]) < 8:
                diag["labels_seen"].append(label)
            if game_label_re.match(label):
                team_venue = _cell_str(ws.cell(r, dc + 1).value)
                time_str = _norm_time(ws.cell(r + 1, dc).value)
                flow = ws.cell(r + 1, dc + 4).value
                quant = _cell_str(ws.cell(r + 1, dc + 5).value)
                team, side = parse_team_venue(team_venue)
                if team:
                    games.append(
                        {
                            "game_label": re.sub(r"\s+", "", label),
                            "time": time_str,
                            "fav_team": team,
                            "fav_nick": team_nickname(team),
                            "side": side,
                            "flow": flow,
                            "quant": quant,
                        }
                    )
                else:
                    diag["skipped_no_team"] += 1
                blanks = 0
                r += 2
                continue
            if label == "" and _cell_str(ws.cell(r + 1, dc).value) == "":
                blanks += 1
                if blanks >= 3:
                    break
            r += 1
        diag["parsed_games"] = len(games)
        if games:
            return games, diag
    return [], diag


def diagnose_record_miss(ws, focus_date: str, diag: dict) -> None:
    """查無資料時印出可操作的診斷。"""
    print(f"[診斷] 目標日期：{focus_date}")
    print(
        f"[診斷] 紀錄!A1 原始值={diag.get('a1_raw')!r}　"
        f"型別={diag.get('a1_type')}　正規化={diag.get('a1_norm')!r}"
    )
    if not diag.get("date_found"):
        print(
            "[診斷] 在「紀錄」各直欄組中找不到此日期表頭。"
            "請確認已存檔、已關閉 Excel，且 A1（或該日區塊頂端）為 YYYYMMDD／日期。"
        )
        # 掃描前 30 列、前 5 組直欄，列出程式認得出的日期
        found = []
        for dc in record_block_cols(ws)[:5]:
            for r in range(1, min(31, (ws.max_row or 1) + 1)):
                n = _norm_date(ws.cell(r, dc).value)
                if n:
                    found.append(f"{get_column_letter(dc)}{r}={n}")
        if found:
            print("[診斷] 目前認得到的日期樣例：" + "、".join(found[:12]))
        else:
            print("[診斷] 前幾個直欄組幾乎認不到任何日期（可能是 data_only 讀到空值／未存檔）。")
        return
    print(
        f"[診斷] 已找到日期表頭：欄{diag.get('date_col')} 列{diag.get('date_row')}，"
        f"但解析出場次數={diag.get('parsed_games')}　"
        f"有「第N場」但隊名解析失敗={diag.get('skipped_no_team')}"
    )
    if diag.get("labels_seen"):
        print("[診斷] 日期下方看到的標籤：" + "、".join(diag["labels_seen"]))
    print("[診斷] 預期結構：日期列下一列起為「第1場」+ 右側「xxx主場/客場」，再下一列為時間。")


def find_daily_game(daily: list[dict], time_str: str, fav_nick: str) -> dict | None:
    cands = [d for d in daily if _norm_time(d.get("time")) == time_str]
    for d in cands:
        if team_nickname(d.get("away", "")) == fav_nick or team_nickname(d.get("home", "")) == fav_nick:
            return d
    return cands[0] if len(cands) == 1 else None


def resolve_game_teams(
    daily: list[dict],
    snapshots: list[dict],
    results: list[dict],
    focus_date: str,
    time_str: str,
    fav_nick: str,
) -> tuple[str, str, str]:
    """回傳 (away, home, 來源)；來源為 daily / jiujiu / 空字串。"""
    d = find_daily_game(daily, time_str, fav_nick)
    if d:
        return d.get("away", ""), d.get("home", ""), "daily"
    pair = find_jiujiu_game_by_fav(results, snapshots, focus_date, time_str, fav_nick)
    if pair:
        return pair[0], pair[1], "jiujiu"
    return "", "", ""


def _clean_home_name(name: str) -> str:
    return _cell_str(name).replace("(主)", "").strip()


def results_for_date(results: list[dict], focus_date: str) -> list[dict]:
    """篩出開賽時間落在 focus_date 的比賽結果列。"""
    out: list[dict] = []
    for row in results:
        kick = str(row.get("開賽時間", "") or "").strip()
        t = _norm_time(kick)
        if not t:
            continue
        if kickoff_matches_focus(kick, focus_date, t):
            out.append(row)
    return out


def parse_fav_from_result(result: dict) -> tuple[str, str, bool] | None:
    """
    從比賽結果解析讓分方。
    回傳 (顯示隊名, 暱稱, 是否主場讓分)；解析失敗則 None。
    """
    away = _cell_str(result.get("客隊"))
    home = _clean_home_name(result.get("主隊"))
    spread_side = _cell_str(result.get("讓分方"))
    away_nick = team_nickname(away) or away
    home_nick = team_nickname(home) or home

    fav_token = ""
    m = re.match(r"^(.+?)讓", spread_side)
    if m:
        fav_token = m.group(1).strip()
    fav_nick = team_nickname(fav_token) or fav_token

    def _hit(token: str, nick: str, full: str) -> bool:
        if not token and not nick:
            return False
        if nick and fav_nick and nick == fav_nick:
            return True
        if token and nick and (token == nick or token in full or nick in token):
            return True
        if nick and nick in spread_side:
            return True
        return False

    if _hit(fav_token, home_nick, home):
        return home_nick or home, home_nick or home, True
    if _hit(fav_token, away_nick, away):
        return away_nick or away, away_nick or away, False

    # 讓分方空白時無法可靠判斷，略過
    if not spread_side:
        return None
    return None


def fmt_flow(value) -> str:
    if value is None or str(value).strip() == "":
        return ""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    pct = round(v * 100) if 0 < abs(v) <= 1 else round(v)
    return f"{pct}%"


def fmt_handicap(value) -> str:
    """玖九讓球盤 → 玖九碼字串（1-70）。數值或玖九碼皆可。"""
    s = str(value or "").strip()
    if not s:
        return ""
    return decimal_to_jiujiu_line(s)


def _fav_handicap_raw(snap: dict, fav_is_home: bool):
    return snap.get("讓球_主盤") if fav_is_home else snap.get("讓球_客盤")


def pick_head_tail(snaps: list[dict], fav_is_home: bool) -> tuple[dict, dict]:
    """頭/尾取「有掛出讓分盤」的快照之第一筆與最後一筆（略過盤口尚未開出的空白筆）。"""
    hc_side = fav_is_home
    valid = [s for s in snaps if fmt_handicap(_fav_handicap_raw(s, hc_side))]
    if not valid:
        hc_side = not fav_is_home
        valid = [s for s in snaps if fmt_handicap(_fav_handicap_raw(s, hc_side))]
    if not valid:
        return {}, {}

    def pack(snap: dict) -> dict:
        fields = side_fields(snap, fav_is_home)
        if hc_side != fav_is_home:
            fields["sp"] = fmt_handicap(_fav_handicap_raw(snap, hc_side))
        return fields

    return pack(valid[0]), pack(valid[-1])


def side_fields(snap: dict, fav_is_home: bool) -> dict:
    """從單一快照取讓分方/受讓方的盤口值。"""
    if not snap:
        return {}
    if fav_is_home:
        sp = snap.get("讓球_主盤")
        fav_ml, rec_ml = snap.get("獨贏_主"), snap.get("獨贏_客")
        fav2, rec2 = snap.get("一輸_主賠率"), snap.get("一輸_客賠率")
    else:
        sp = snap.get("讓球_客盤")
        fav_ml, rec_ml = snap.get("獨贏_客"), snap.get("獨贏_主")
        fav2, rec2 = snap.get("一輸_客賠率"), snap.get("一輸_主賠率")
    return {
        "sp": fmt_handicap(sp),
        "fav_ml": format_odds(str(fav_ml or "")),
        "rec_ml": format_odds(str(rec_ml or "")),
        "fav2": format_odds(str(fav2 or "")),
        "rec2": format_odds(str(rec2 or "")),
    }


def translate_results(result: dict, fav_is_home: bool, unknowns: list[str]) -> dict:
    """用對照表把比賽結果翻成 H/M/R；未知字串記錄到 unknowns 並留空。"""
    out = {"spread": "", "ml": "", "two": ""}
    if not result:
        return out

    settle = _cell_str(result.get("讓分結算"))
    if settle:
        if settle in SPREAD_SETTLE_MAP:
            out["spread"] = SPREAD_SETTLE_MAP[settle]
        else:
            unknowns.append(f"讓分結算「{settle}」")

    winner = _cell_str(result.get("勝方"))
    if winner in ("客", "主"):
        fav_char = "主" if fav_is_home else "客"
        out["ml"] = "讓" if winner == fav_char else "受"
    elif winner:
        unknowns.append(f"勝方「{winner}」")

    two = _cell_str(result.get("兩分贏結果"))
    if two:
        matched = False
        for kw, val in TWO_RESULT_RULES:
            if kw in two:
                out["two"] = val
                matched = True
                break
        if not matched:
            unknowns.append(f"兩分贏結果「{two}」")
    return out


def find_last_data_row(ws) -> int:
    """以 B 欄（讓分球隊）找最後一筆資料列。"""
    last = 1
    for r in range(2, ws.max_row + 1):
        if _cell_str(ws.cell(r, COL_TEAM).value):
            last = r
    return last


def set_if_blank(ws, row: int, col: int, value, date_fmt: str | None = None) -> bool:
    """空白才寫入；回傳是否實際寫入。"""
    if value is None or str(value).strip() == "":
        return False
    cell = ws.cell(row, col)
    if cell.value is not None and str(cell.value).strip() != "":
        return False
    cell.value = value
    if date_fmt and col == COL_DATE:
        cell.number_format = date_fmt
    return True


def set_tail(ws, row: int, col: int, value) -> bool:
    """尾欄：有新值就覆蓋（反映開賽前最後快照）；回傳是否實際變更。"""
    if value is None or str(value).strip() == "":
        return False
    cell = ws.cell(row, col)
    new_s = str(value).strip()
    old_s = _cell_str(cell.value)
    if old_s == new_s:
        return False
    cell.value = value
    return True


def write_cell(
    ws, row: int, col: int, value, *, date_fmt: str | None = None
) -> bool:
    """依欄位類型寫入：尾欄覆蓋，其餘空白才填。"""
    if col in TAIL_COLS:
        return set_tail(ws, row, col, value)
    return set_if_blank(ws, row, col, value, date_fmt)


def verify_headers(ws) -> None:
    bad = []
    for col, kw in HEADER_CHECKS.items():
        head = _cell_str(ws.cell(1, col).value)
        if kw not in head:
            bad.append(f"{get_column_letter(col)}1 應含「{kw}」，實際為「{head}」")
    if bad:
        raise SystemExit("盤口觀察標題列與預期不符，為安全起見中止：\n  " + "\n  ".join(bad))


def _prune_observation_backups(base_name: str, ext: str, *, keep: int) -> list[str]:
    """刪除超過 keep 份的舊備份，回傳已刪除檔名。"""
    if keep < 1 or not os.path.isdir(BACKUP_DIR):
        return []
    prefix = f"{base_name}_備份_"
    candidates: list[tuple[float, str]] = []
    for name in os.listdir(BACKUP_DIR):
        if not (name.startswith(prefix) and name.endswith(ext)):
            continue
        full = os.path.join(BACKUP_DIR, name)
        if os.path.isfile(full):
            candidates.append((os.path.getmtime(full), full))
    candidates.sort(key=lambda item: item[0], reverse=True)
    removed = []
    for _, full in candidates[keep:]:
        try:
            os.remove(full)
            removed.append(os.path.basename(full))
        except OSError:
            pass
    return removed


def backup_file(path: str, *, keep: int = OBSERVATION_BACKUP_KEEP) -> str:
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.basename(path)
    base_name, ext = os.path.splitext(filename)
    bak_name = f"{base_name}_備份_{ts}{ext}"
    bak = os.path.join(BACKUP_DIR, bak_name)
    shutil.copy2(path, bak)
    _prune_observation_backups(base_name, ext, keep=keep)
    return bak


def save_workbook(wb, path: str) -> None:
    tmp = path + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    wb.save(tmp)
    try:
        check = load_workbook(tmp, read_only=True)
        check.close()
    except Exception as e:
        os.remove(tmp)
        raise SystemExit(f"產生的檔案無法通過驗證，未覆蓋原檔：{e}") from e
    os.remove(path)
    shutil.move(tmp, path)


def print_lookup_table() -> None:
    print("【結果欄對照表（核對用）】")
    print("  H 讓分結果：" + "、".join(f"{k}→{v}" for k, v in SPREAD_SETTLE_MAP.items()))
    print("  M 獨贏結果：勝方=讓分方→讓、勝方=受讓方→受")
    print("  R 2分結果：" + "、".join(f"含「{k}」→{v}" for k, v in TWO_RESULT_RULES))
    print("  尾欄 G/J/L/O/Q：重跑時覆蓋為最新開賽前快照")
    print()


STD_FONT_NAME = "新細明體"
STD_FONT_SIZE = 11
CENTER_COLS = (COL_SP_RESULT, COL_ML_RESULT, COL_TWO_RESULT)


def get_main_table(ws):
    """取第一張 Excel 表格（盤口觀察只有一張）。"""
    if not ws.tables:
        return None
    name = list(ws.tables.keys())[0]
    return ws.tables[name]


def style_appended_rows(ws, row_start: int, row_end: int, max_col: int, date_fmt: str) -> None:
    """對表格外新列套用與既有列一致的字型、日期格式、結果欄置中。"""
    for r in range(row_start, row_end + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).font = Font(name=STD_FONT_NAME, size=STD_FONT_SIZE)
        ws.cell(r, COL_DATE).number_format = date_fmt
        for c in CENTER_COLS:
            ws.cell(r, c).alignment = Alignment(horizontal="center")


def extend_table_and_format(ws, table, orig_bounds, date_fmt: str) -> int | None:
    """把超出原表格範圍的新列納入表格並套格式；回傳擴張後的最後列（無擴張則 None）。"""
    if table is None:
        return None
    min_c, min_r, max_c, max_r = orig_bounds
    final_last = find_last_data_row(ws)
    if final_last <= max_r:
        return None
    style_appended_rows(ws, max_r + 1, final_last, max_c, date_fmt)
    table.ref = (
        f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{final_last}"
    )
    return final_last


def build(start: datetime, end: datetime) -> None:
    for p in (JIUJIU_XLSX_FILE, OBSERVATION_XLSX_FILE):
        if not os.path.exists(p):
            raise SystemExit(f"找不到檔案：{p}")

    print("=" * 60)
    print("填入盤口觀察")
    print("=" * 60)
    print(f"日期範圍：{start:%Y/%m/%d} ~ {end:%Y/%m/%d}")
    print(f"目標檔：{OBSERVATION_XLSX_FILE}")
    print("場次來源：玖九「比賽結果」；金流／量化：紀錄有則補、無則留空")
    print()
    print("【重要】Excel 只顯示記憶體內容；本程式讀的是「已存檔」的磁碟檔。")
    print("請先在 Excel 對 MLB／玖九／盤口觀察按「存檔」並關閉，再繼續。")
    print()
    ensure_excel_files_closed(
        [MLB_XLSX_FILE, JIUJIU_XLSX_FILE, OBSERVATION_XLSX_FILE]
    )
    print_lookup_table()

    ws_record = None
    mlb_wb = None
    if os.path.exists(MLB_XLSX_FILE):
        print("讀取 MLB26-27（紀錄、每日數據；用於補金流／量化）…")
        try:
            mlb_wb = load_workbook(MLB_XLSX_FILE, data_only=True)
            if SHEET_RECORD in mlb_wb.sheetnames:
                ws_record = mlb_wb[SHEET_RECORD]
            else:
                print("[注意] MLB26-27 找不到「紀錄」分頁，金流／量化將留空。")
        except Exception as e:
            print(f"[注意] 無法讀取 MLB26-27（金流／量化將留空）：{e}")
            mlb_wb = None
            ws_record = None
    else:
        print("[注意] 找不到 MLB26-27，金流／量化將留空。")

    print("讀取玖九（MLB快照、比賽結果）…")
    jj_wb = load_workbook(JIUJIU_XLSX_FILE, read_only=True, data_only=True)
    snapshots = read_snapshot_rows(jj_wb)
    results = read_result_rows(jj_wb)
    jj_wb.close()
    if not results:
        raise SystemExit("玖九「比賽結果」沒有資料，無法建立場次。")

    obs_wb = load_workbook(OBSERVATION_XLSX_FILE)
    ws = obs_wb[obs_wb.sheetnames[0]]
    verify_headers(ws)
    date_fmt = ws.cell(2, COL_DATE).number_format or "yyyy-mm-dd"

    main_table = get_main_table(ws)
    orig_table_bounds = range_boundaries(main_table.ref) if main_table else None

    existing: dict[tuple[str, str], int] = {}
    last_row = find_last_data_row(ws)
    for r in range(2, last_row + 1):
        d = _norm_date(ws.cell(r, COL_DATE).value)
        team = team_nickname(_cell_str(ws.cell(r, COL_TEAM).value))
        if d and team:
            existing[(d, team)] = r
    append_row = last_row + 1

    unknowns: list[str] = []
    total_games = 0
    new_rows = 0
    filled_cells = 0
    tail_updates = 0
    record_filled = 0
    no_source_dates: list[str] = []
    skipped_no_fav = 0
    pregame_rows = 0

    for focus_date in daterange(start, end):
        day_results = results_for_date(results, focus_date)
        record_games: list[dict] = []
        if ws_record is not None:
            record_games, _diag = read_record_games_with_diag(ws_record, focus_date)
        record_map = {
            g["fav_nick"]: g for g in record_games if g.get("fav_nick")
        }

        # 合併場次：比賽結果（有賽果）+ 紀錄（賽前建頭盤）
        jobs: list[dict] = []
        seen_nicks: set[str] = set()

        for result in day_results:
            parsed = parse_fav_from_result(result)
            if not parsed:
                skipped_no_fav += 1
                spread = _cell_str(result.get("讓分方"))
                if spread:
                    unknowns.append(f"讓分方無法解析「{spread}」")
                continue
            fav_team, fav_nick, fav_is_home = parsed
            if fav_nick in seen_nicks:
                continue
            seen_nicks.add(fav_nick)
            jobs.append(
                {
                    "source": "result",
                    "result": result,
                    "fav_team": fav_team,
                    "fav_nick": fav_nick,
                    "fav_is_home": fav_is_home,
                    "away": _cell_str(result.get("客隊")),
                    "home": _clean_home_name(result.get("主隊")),
                    "time": _norm_time(result.get("開賽時間")),
                }
            )

        for g in record_games:
            fav_nick = g.get("fav_nick") or ""
            if not fav_nick or fav_nick in seen_nicks:
                continue
            seen_nicks.add(fav_nick)
            side = _cell_str(g.get("side"))
            jobs.append(
                {
                    "source": "record",
                    "result": None,
                    "fav_team": _cell_str(g.get("fav_team")) or fav_nick,
                    "fav_nick": fav_nick,
                    "fav_is_home": side == "主",
                    "away": "",
                    "home": "",
                    "time": _norm_time(g.get("time")),
                    "record": g,
                }
            )

        if not jobs:
            no_source_dates.append(focus_date)
            continue

        daily: list[dict] = []
        if mlb_wb is not None and "每日數據" in getattr(mlb_wb, "sheetnames", []):
            try:
                daily = read_daily_games(mlb_wb, focus_date)
            except Exception:
                daily = []

        date_obj = datetime.strptime(focus_date, "%Y%m%d")

        for job in jobs:
            fav_team = job["fav_team"]
            fav_nick = job["fav_nick"]
            fav_is_home = job["fav_is_home"]
            away = job.get("away") or ""
            home = job.get("home") or ""
            time_str = job.get("time") or ""
            result = job.get("result")
            total_games += 1
            if job.get("source") == "record" and result is None:
                pregame_rows += 1

            if not away or not home:
                a2, h2, _src = resolve_game_teams(
                    daily, snapshots, results, focus_date, time_str, fav_nick
                )
                away = away or a2
                home = home or h2

            # 若仍無客主，無法取快照頭尾；仍可寫模型欄
            head = tail = {}
            if away and home and time_str:
                snaps = all_pregame_snapshots(
                    snapshots, focus_date, time_str, away, home
                )
                if snaps:
                    # 若僅有紀錄來源，依隊名再確認主客
                    if team_nickname(home) == fav_nick:
                        fav_is_home = True
                    elif team_nickname(away) == fav_nick:
                        fav_is_home = False
                    head, tail = pick_head_tail(snaps, fav_is_home)

            if result is None and away and home and time_str:
                result = find_result_row(results, focus_date, time_str, away, home)

            res = translate_results(result or {}, fav_is_home, unknowns)
            rec = job.get("record") or record_map.get(fav_nick)
            if rec:
                record_filled += 1
                flow = fmt_flow(rec.get("flow"))
                quant = _cell_str(rec.get("quant"))
                side = _cell_str(rec.get("side")) or ("主" if fav_is_home else "客")
                display_team = _cell_str(rec.get("fav_team")) or fav_team
            else:
                flow = ""
                quant = ""
                side = "主" if fav_is_home else "客"
                display_team = fav_team

            key = (focus_date, fav_nick)
            if key in existing:
                row = existing[key]
                is_new = False
            else:
                row = append_row
                append_row += 1
                existing[key] = row
                is_new = True
                new_rows += 1

            if is_new:
                if set_if_blank(ws, row, COL_DATE, date_obj, date_fmt):
                    filled_cells += 1

            writes = [
                (COL_TEAM, display_team),
                (COL_FLOW, flow),
                (COL_SIDE, side),
                (COL_QUANT, quant),
                (COL_SP_HEAD, head.get("sp", "")),
                (COL_SP_TAIL, tail.get("sp", "")),
                (COL_SP_RESULT, res["spread"]),
                (COL_FAVML_HEAD, head.get("fav_ml", "")),
                (COL_FAVML_TAIL, tail.get("fav_ml", "")),
                (COL_RECML_HEAD, head.get("rec_ml", "")),
                (COL_RECML_TAIL, tail.get("rec_ml", "")),
                (COL_ML_RESULT, res["ml"]),
                (COL_FAV2_HEAD, head.get("fav2", "")),
                (COL_FAV2_TAIL, tail.get("fav2", "")),
                (COL_REC2_HEAD, head.get("rec2", "")),
                (COL_REC2_TAIL, tail.get("rec2", "")),
                (COL_TWO_RESULT, res["two"]),
            ]
            for col, val in writes:
                if write_cell(ws, row, col, val, date_fmt=date_fmt):
                    filled_cells += 1
                    if col in TAIL_COLS and not is_new:
                        tail_updates += 1

    if mlb_wb is not None:
        mlb_wb.close()

    if no_source_dates:
        print()
        print(
            "無場次可寫的日期（玖九比賽結果與紀錄皆無）："
            + ", ".join(no_source_dates)
        )
    if skipped_no_fav:
        print(f"[注意] 讓分方無法解析而略過：{skipped_no_fav} 場")
    if pregame_rows:
        print(f"賽前建列（僅頭盤／模型，尚無比賽結果）：{pregame_rows} 場")

    if filled_cells == 0 and new_rows == 0:
        print("沒有需要填入的空格（可能皆已填過，或區間內無場次）。未變更檔案。")
        obs_wb.close()
        _report(
            no_source_dates,
            unknowns,
            total_games,
            new_rows,
            filled_cells,
            tail_updates,
            record_filled=record_filled,
        )
        return

    if main_table is not None:
        extended_to = extend_table_and_format(ws, main_table, orig_table_bounds, date_fmt)
        if extended_to:
            print(f"表格「{main_table.displayName}」範圍已擴張至第 {extended_to} 列並套用格式")

    bak = backup_file(OBSERVATION_XLSX_FILE)
    print(f"已建立還原點（備份）：{bak}（回朔點備分，保留最近 {OBSERVATION_BACKUP_KEEP} 份）")
    save_workbook(obs_wb, OBSERVATION_XLSX_FILE)
    obs_wb.close()
    print(f"[完成] 已更新：{OBSERVATION_XLSX_FILE}")
    _report(
        no_source_dates,
        unknowns,
        total_games,
        new_rows,
        filled_cells,
        tail_updates,
        record_filled=record_filled,
    )


def _report(
    no_source_dates,
    unknowns,
    total_games,
    new_rows,
    filled_cells,
    tail_updates=0,
    record_filled=0,
) -> None:
    print()
    print("-" * 60)
    print(
        f"處理場次：{total_games}　新增列：{new_rows}　填入格數：{filled_cells}"
        + (f"　尾欄更新：{tail_updates}" if tail_updates else "")
        + (f"　紀錄補金流／量化：{record_filled}" if record_filled else "")
    )
    if no_source_dates:
        print(f"玖九比賽結果查無資料的日期：{', '.join(no_source_dates)}")
    if unknowns:
        from collections import Counter

        print("[注意] 對照表未涵蓋／讓分方無法解析的字串：")
        for s, n in Counter(unknowns).most_common():
            print(f"   {s} ×{n}")
    print("-" * 60)


def main() -> None:
    for stream in (sys.stdout, sys.stderr, sys.stdin):
        try:
            stream.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError):
            pass
    parser = argparse.ArgumentParser(description="填入盤口觀察")
    parser.add_argument(
        "start_date",
        nargs="?",
        help="開始日期 YYYYMMDD",
    )
    parser.add_argument(
        "end_date",
        nargs="?",
        help="結束日期 YYYYMMDD（單日可與開始日期相同）",
    )
    parser.add_argument(
        "--no-lock",
        action="store_true",
        help="由父行程（run_all）已持鎖時略過防並行",
    )
    args = parser.parse_args()

    lock_held = False
    if not args.no_lock:
        from pipeline_lock import acquire_pipeline_lock, release_pipeline_lock

        acquire_pipeline_lock(
            os.path.dirname(os.path.abspath(__file__)),
            owner="填入盤口觀察",
        )
        lock_held = True

    try:
        if args.start_date and args.end_date:
            start, end = parse_cli_date_range(args.start_date, args.end_date)
        elif args.start_date:
            start = end = parse_yyyymmdd(args.start_date)
        else:
            start, end = prompt_date_range_interactive()

        build(start, end)
    finally:
        if lock_held:
            from pipeline_lock import release_pipeline_lock

            release_pipeline_lock()


if __name__ == "__main__":
    main()
