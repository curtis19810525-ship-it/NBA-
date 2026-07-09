# -*- coding: utf-8 -*-
"""
爬取 ESPN MLB 各隊 Total（獨立腳本）
- 資料來源：ESPN MLB team stats（All Splits + Expanded）
- 寫入分頁：各隊total
- 每次執行：清空第 2 列起，重寫 30 隊資料（即時總覽）
"""

import os
import re
import sys
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from config import NBA_STATS_FILE
except ImportError:
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# 固定 MLB 30 隊順序（與各隊正負一致）
MLB_TEAMS = [
    ("金鶯", "bal", "baltimore-orioles"),
    ("紅襪", "bos", "boston-red-sox"),
    ("洋基", "nyy", "new-york-yankees"),
    ("光芒", "tb", "tampa-bay-rays"),
    ("藍鳥", "tor", "toronto-blue-jays"),
    ("白襪", "chw", "chicago-white-sox"),
    ("守護者", "cle", "cleveland-guardians"),
    ("老虎", "det", "detroit-tigers"),
    ("皇家", "kc", "kansas-city-royals"),
    ("雙城", "min", "minnesota-twins"),
    ("太空人", "hou", "houston-astros"),
    ("天使", "laa", "los-angeles-angels"),
    ("運動家", "ath", "athletics"),
    ("水手", "sea", "seattle-mariners"),
    ("遊騎兵", "tex", "texas-rangers"),
    ("勇士", "atl", "atlanta-braves"),
    ("馬林魚", "mia", "miami-marlins"),
    ("大都會", "nym", "new-york-mets"),
    ("費城人", "phi", "philadelphia-phillies"),
    ("國民", "wsh", "washington-nationals"),
    ("小熊", "chc", "chicago-cubs"),
    ("紅人", "cin", "cincinnati-reds"),
    ("釀酒人", "mil", "milwaukee-brewers"),
    ("海盜", "pit", "pittsburgh-pirates"),
    ("紅雀", "stl", "st-louis-cardinals"),
    ("響尾蛇", "ari", "arizona-diamondbacks"),
    ("落磯", "col", "colorado-rockies"),
    ("道奇", "lad", "los-angeles-dodgers"),
    ("教士", "sd", "san-diego-padres"),
    ("巨人", "sf", "san-francisco-giants"),
]

TOTAL_HEADERS = [
    "球隊",
    "GP/出賽",
    "AB/打數",
    "R/得分",
    "H/安打",
    "2B/二壘打",
    "3B/三壘打",
    "HR/全壘打",
    "RBI/打點",
    "TB/壘打數",
    "BB/保送",
    "SO/三振",
    "SB/盜壘",
    "AVG/打擊率",
    "OBP/上壘率",
    "SLG/長打率",
    "OPS/整體攻擊指數",
    "WAR/勝場貢獻值",
    "RC/得分創造",
    "RC/27/每27出局得分創造",
    "BB/PA/保送率",
    "BB/K/保送三振比",
    "ISOP/純長打率",
    "SECA/次級攻擊指數",
    "P/PA/每打席用球數",
    "XBH/長打數",
    "PA/打席",
    "AB/HR/全壘打打數比",
]


def _n(s):
    return re.sub(r"\s+", "", str(s).strip().upper())


def _t(cell):
    return re.sub(r"\s+", " ", cell.get_text(" ", strip=True)).strip()


def _table_headers(table):
    tr = table.find("tr")
    if not tr:
        return []
    return [_n(_t(c)) for c in tr.find_all(["th", "td"])]


def _rows(table):
    rows = []
    for tr in table.find_all("tr")[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        vals = [_t(c) for c in cells]
        if any(vals):
            rows.append(vals)
    return rows


def _pair_name_and_stats(name_rows, stat_rows, stat_headers):
    out = []
    n = min(len(name_rows), len(stat_rows))
    for i in range(n):
        nm = name_rows[i][0].strip() if name_rows[i] else ""
        vals = stat_rows[i]
        if not nm:
            continue
        if len(vals) < len(stat_headers):
            vals += [""] * (len(stat_headers) - len(vals))
        if len(vals) > len(stat_headers):
            vals = vals[:len(stat_headers)]
        out.append((nm, dict(zip(stat_headers, vals))))
    return out


def fetch_team_maps(abbr, slug):
    url = f"https://www.espn.com/mlb/team/stats/_/name/{abbr}/{slug}"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    tables = soup.find_all("table")
    if len(tables) < 4:
        return {}, {}

    basic_need = {"GP", "AB", "R", "H", "2B", "3B", "HR", "RBI", "TB", "BB", "SO", "SB", "AVG", "OBP", "SLG", "OPS", "WAR"}
    ex_need = {"GP", "AB", "RC", "RC/27", "BB/PA", "BB/K", "ISOP", "SECA", "P/PA", "XBH", "PA", "AB/HR"}

    basic_map = {}
    ex_map = {}
    headers_list = [_table_headers(tb) for tb in tables]

    for i in range(len(tables) - 1):
        h1 = headers_list[i]
        h2 = headers_list[i + 1]
        if h1 == ["NAME"] and basic_need.issubset(set(h2)) and not basic_map:
            name_rows = _rows(tables[i])
            stat_rows = _rows(tables[i + 1])
            for name, mp in _pair_name_and_stats(name_rows, stat_rows, h2):
                basic_map[name] = mp
        if h1 == ["NAME"] and ex_need.issubset(set(h2)) and not ex_map:
            name_rows = _rows(tables[i])
            stat_rows = _rows(tables[i + 1])
            for name, mp in _pair_name_and_stats(name_rows, stat_rows, h2):
                ex_map[name] = mp

    return basic_map, ex_map


def parse_team_total(team_cn, abbr, slug):
    try:
        basic_map, ex_map = fetch_team_maps(abbr, slug)
    except Exception as e:
        print(f"  抓取失敗 {team_cn}: {e}")
        return [team_cn] + [""] * (len(TOTAL_HEADERS) - 1)

    total_basic = basic_map.get("Total", {})
    total_ex = ex_map.get("Total", {})

    row = [
        team_cn,
        total_basic.get("GP", ""),
        total_basic.get("AB", ""),
        total_basic.get("R", ""),
        total_basic.get("H", ""),
        total_basic.get("2B", ""),
        total_basic.get("3B", ""),
        total_basic.get("HR", ""),
        total_basic.get("RBI", ""),
        total_basic.get("TB", ""),
        total_basic.get("BB", ""),
        total_basic.get("SO", ""),
        total_basic.get("SB", ""),
        total_basic.get("AVG", ""),
        total_basic.get("OBP", ""),
        total_basic.get("SLG", ""),
        total_basic.get("OPS", ""),
        total_basic.get("WAR", ""),
        total_ex.get("RC", ""),
        total_ex.get("RC/27", ""),
        total_ex.get("BB/PA", ""),
        total_ex.get("BB/K", ""),
        total_ex.get("ISOP", ""),
        total_ex.get("SECA", ""),
        total_ex.get("P/PA", ""),
        total_ex.get("XBH", ""),
        total_ex.get("PA", ""),
        total_ex.get("AB/HR", ""),
    ]
    return row


def write_totals_to_excel(rows):
    wb = load_workbook(NBA_STATS_FILE)
    if "各隊total" not in wb.sheetnames:
        wb.create_sheet("各隊total")
    ws = wb["各隊total"]

    for c, h in enumerate(TOTAL_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    wb.save(NBA_STATS_FILE)
    wb.close()


def main():
    print("=" * 60)
    print("爬取 ESPN MLB 各隊 Total")
    print("=" * 60)
    print(f"Excel: {NBA_STATS_FILE}")
    if not os.path.exists(NBA_STATS_FILE):
        print("錯誤：Excel 檔案不存在")
        return 1

    out_rows = []
    for idx, (team_cn, abbr, slug) in enumerate(MLB_TEAMS, start=1):
        print(f"[{idx:02d}/30] {team_cn}")
        out_rows.append(parse_team_total(team_cn, abbr, slug))

    write_totals_to_excel(out_rows)
    print("完成：已覆蓋寫入分頁「各隊total」。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
