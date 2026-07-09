# -*- coding: utf-8 -*-
"""
爬取 NBA 今日先發名單
資料來源: https://stats.nba.com/js/data/leaders/00_daily_lineups_YYYYMMDD.json
寫入 MLB26-27數據.xlsx「先發名單」分頁，方案 A：一列一場，欄位 日期|客隊|主隊|客隊先發1~5|客隊替補1~8|主隊先發1~5|主隊替補1~8
無賽事時保留上次資料並加註「今日無賽事」
"""

import os
import sys
import json
import unicodedata
import requests
from datetime import datetime
from openpyxl import load_workbook


def normalize_player_name_display(name):
    """
    去掉變音符，與 ESPN「球員狀態」B 欄拼法一致（例：Luka Dončić -> Luka Doncic）。
    供「數據分析」從先發名單引用後，可用 SEARCH 對到球員狀態。
    """
    if not name:
        return name
    s = str(name).strip()
    if not s:
        return s
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

try:
    from zoneinfo import ZoneInfo
    USE_ZONEINFO = True
except ImportError:
    USE_ZONEINFO = False

try:
    from config import NBA_STATS_FILE
except ImportError:
    NBA_STATS_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"

# NBA 縮寫 -> 中文隊名（與各隊正負一致）
NBA_ABBR_TO_ZH = {
    'BOS': '賽爾提克', 'BKN': '籃網', 'NY': '尼克', 'NYK': '尼克',
    'PHI': '76人', 'TOR': '暴龍', 'CHI': '公牛', 'CLE': '騎士',
    'DET': '活塞', 'IND': '溜馬', 'MIL': '公鹿', 'ATL': '老鷹',
    'CHA': '黃蜂', 'MIA': '熱火', 'ORL': '魔術', 'WAS': '巫師',
    'GSW': '勇士', 'LAC': '快艇', 'LAL': '湖人', 'PHX': '太陽',
    'SAC': '國王', 'DEN': '金塊', 'MIN': '灰狼', 'OKC': '雷霆',
    'POR': '拓荒者', 'UTA': '爵士', 'DAL': '獨行俠', 'HOU': '火箭',
    'MEM': '灰熊', 'NOP': '鵜鶘', 'SAS': '馬刺',
}

SHEET_NAME = '先發名單'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': 'https://www.nba.com/',
}


def get_today_yyyymmdd():
    """取得美東時區的今日日期（NBA 以美東為準）"""
    if USE_ZONEINFO:
        now_et = datetime.now(ZoneInfo("America/New_York"))
    else:
        try:
            import pytz
            now_et = datetime.now(pytz.timezone("America/New_York"))
        except ImportError:
            from datetime import timezone, timedelta
            et = timezone(timedelta(hours=-5))  # EST 近似值
            now_et = datetime.now(et)
    return now_et.strftime('%Y%m%d')


def fetch_daily_lineups(date_yyyymmdd):
    """抓取指定日期的先發名單 JSON"""
    url = f"https://stats.nba.com/js/data/leaders/00_daily_lineups_{date_yyyymmdd}.json"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  抓取失敗: {e}")
        return None


def get_starters_and_bench(players_list):
    """取先發 5 人（有 position）及替補 8 人，替補排除先發名單，回傳 (starters, bench)"""
    starters = []
    rest = []
    for p in players_list:
        name = normalize_player_name_display(p.get('playerName') or '')
        if p.get('position') and len(starters) < 5:
            starters.append(name)
        else:
            rest.append(name)
    while len(starters) < 5:
        starters.append('')
    # 替補排除先發（避免重疊）
    starter_set = {s for s in starters if s}
    bench_raw = [r for r in rest if r and r not in starter_set]
    bench = (bench_raw + [''] * 8)[:8]
    return starters, bench


def team_abbr_to_zh(abbr):
    if not abbr:
        return ''
    return NBA_ABBR_TO_ZH.get(abbr.upper(), abbr)


def parse_games(data):
    """解析 JSON 為 [日期, 客隊, 主隊, 客先發1~5, 客替補1~8, 主先發1~5, 主替補1~8]"""
    rows = []
    games = data.get('games') or []
    date_str = get_today_yyyymmdd()
    for g in games:
        away = g.get('awayTeam') or {}
        home = g.get('homeTeam') or {}
        away_zh = team_abbr_to_zh(away.get('teamAbbreviation') or '')
        home_zh = team_abbr_to_zh(home.get('teamAbbreviation') or '')
        away_starters, away_bench = get_starters_and_bench(away.get('players') or [])
        home_starters, home_bench = get_starters_and_bench(home.get('players') or [])
        row = [date_str, away_zh, home_zh] + away_starters + away_bench + home_starters + home_bench
        rows.append(row)
    return rows


def ensure_sheet(wb, name):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def write_lineups(wb, rows, no_games_today=False):
    """寫入「先發名單」：有資料則清空第2列起重寫；無賽事則保留並加註"""
    sh = ensure_sheet(wb, SHEET_NAME)
    # 表頭：有資料時一律重寫，確保欄位對齊 D-H 客隊先發1~5, I-P 客隊替補6~13, 主隊先發1~5, 主隊替補6~13
    headers = ['日期', '客隊', '主隊',
               '客隊先發1', '客隊先發2', '客隊先發3', '客隊先發4', '客隊先發5',
               '客隊替補6', '客隊替補7', '客隊替補8', '客隊替補9', '客隊替補10', '客隊替補11', '客隊替補12', '客隊替補13',
               '主隊先發1', '主隊先發2', '主隊先發3', '主隊先發4', '主隊先發5',
               '主隊替補6', '主隊替補7', '主隊替補8', '主隊替補9', '主隊替補10', '主隊替補11', '主隊替補12', '主隊替補13']
    for c, h in enumerate(headers, 1):
        sh.cell(1, c, h)
    if no_games_today:
        next_row = sh.max_row + 1
        sh.cell(next_row, 1, get_today_yyyymmdd())
        sh.cell(next_row, 2, '今日無賽事')
        for c in range(3, 30):
            sh.cell(next_row, c, '')
        print(f"  今日無賽事，已加註於第 {next_row} 列")
        return
    # 清空第 2 列起
    max_row = sh.max_row
    for r in range(max_row, 1, -1):
        sh.delete_rows(r)
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            sh.cell(row=i, column=j, value=val)
    print(f"  已寫入「先發名單」{len(rows)} 場（第2列起）")


def main():
    print("=" * 60)
    print("爬取 NBA 今日先發名單")
    print("=" * 60)
    print(f"Excel: {NBA_STATS_FILE}")
    if not os.path.exists(NBA_STATS_FILE):
        print("錯誤：Excel 檔案不存在")
        return 1
    today = get_today_yyyymmdd()
    print(f"日期: {today}")
    data = fetch_daily_lineups(today)
    if data is None:
        print("無法取得資料，保留原分頁內容並加註「今日無賽事」")
        wb = load_workbook(NBA_STATS_FILE)
        write_lineups(wb, [], no_games_today=True)
        wb.save(NBA_STATS_FILE)
        return 0
    rows = parse_games(data)
    if not rows:
        print("今日無賽事，保留上次資料並加註。")
        wb = load_workbook(NBA_STATS_FILE)
        write_lineups(wb, [], no_games_today=True)
        wb.save(NBA_STATS_FILE)
        return 0
    wb = load_workbook(NBA_STATS_FILE)
    write_lineups(wb, rows)
    wb.save(NBA_STATS_FILE)
    print("完成。")
    return 0


if __name__ == '__main__':
    sys.exit(main())
