# -*- coding: utf-8 -*-
"""分析母版：讀取 MLB 總表、每日數據、玖九快照。"""

from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

try:
    from config import JIUJIU_XLSX_FILE, MLB_XLSX_FILE
except ImportError:
    MLB_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\MLB26-27數據.xlsx"
    JIUJIU_XLSX_FILE = r"C:\Users\curti\OneDrive\MLB26\玖九盤口變化.xlsx"

_JIUJIU_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "抓取玖九盤口動態資訊")
)
if _JIUJIU_DIR not in sys.path:
    sys.path.insert(0, _JIUJIU_DIR)

from mlb_filter import parse_start_time, resolve_mlb_team, team_nickname  # noqa: E402
from odds_format import format_game_fields_for_timeline, format_odds, format_timeline_handicap_cell  # noqa: E402

SNAPSHOT_COLS = [
    "抓取時間",
    "觸發方式",
    "分頁",
    "開賽時間",
    "客隊",
    "主隊",
    "讓球_客盤",
    "讓球_客賠率",
    "讓球_主盤",
    "讓球_主賠率",
    "獨贏_客",
    "獨贏_主",
    "一輸_客線",
    "一輸_客賠率",
    "一輸_主賠率",
    "原始區塊",
]

RESULT_COLS = [
    "開賽時間",
    "客隊",
    "主隊",
    "客得分",
    "主得分",
    "比分",
    "勝方",
    "讓分方",
    "讓分結算",
    "客隊獨贏",
    "主隊獨贏",
    "獨贏結果",
    "客隊兩分差",
    "主隊兩分差",
    "兩分贏結果",
    "結果更新時間",
]

TZ = parse_start_time.__globals__["TZ"]
RECENT_CAPTURES = 5


def _row_dict(headers: list[str], row: tuple) -> dict[str, Any]:
    return {h: (row[i] if i < len(row) else "") for i, h in enumerate(headers)}


def _norm_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, int):
        s = str(value)
        return s if len(s) == 8 and s.isdigit() else ""
    s = str(value).strip().replace("-", "").replace("/", "")
    if len(s) >= 8 and s[:8].isdigit():
        return s[:8]
    return ""


def _norm_time(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    s = str(value).strip()
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def _parse_scrape_dt(s: str) -> datetime:
    s = str(s or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=TZ)
        except ValueError:
            continue
    return datetime.min.replace(tzinfo=TZ)


def _display_home(name: str) -> str:
    name = str(name or "").strip()
    if name and "(主)" not in name:
        return f"{name}(主)"
    return name


def game_key(time_str: str, away: str, home: str) -> tuple[str, str, str]:
    return (_norm_time(time_str), team_nickname(away), team_nickname(home))


def teams_match(away_a: str, home_a: str, away_b: str, home_b: str) -> bool:
    ra, rb = resolve_mlb_team(away_a), resolve_mlb_team(away_b)
    ha, hb = resolve_mlb_team(home_a), resolve_mlb_team(home_b)
    if not ra or not rb or not ha or not hb:
        return False
    return ra == rb and ha == hb


def jiujiu_kickoff(focus_date: str, time_str: str) -> str:
    mm = focus_date[4:6]
    dd = focus_date[6:8]
    return f"{mm}-{dd} {_norm_time(time_str)}"


def kickoff_matches_focus(開賽時間: str, focus_date: str, time_str: str) -> bool:
    """開賽時間須對到關注日 + 時分，避免跨日誤配。"""
    if _norm_time(開賽時間) != time_str:
        return False
    ref = date(int(focus_date[:4]), int(focus_date[4:6]), int(focus_date[6:8]))
    kickoff = jiujiu_kickoff(focus_date, time_str)
    dt = parse_start_time(str(開賽時間 or "").strip(), ref) or parse_start_time(kickoff, ref)
    if not dt:
        s = str(開賽時間 or "")
        mm, dd = focus_date[4:6], focus_date[6:8]
        return f"{mm}-{dd}" in s or f"{mm}/{dd}" in s
    return dt.date() == ref


def read_daily_games(wb: Workbook, focus_date: str) -> list[dict[str, Any]]:
    from 每日數據轉正負盤 import read_games_from_daily_data

    ws = wb["每日數據"]
    raw = read_games_from_daily_data(ws, focus_date)
    team_rows = {_norm_time(g["time"]): g for g in _read_daily_games_with_teams(wb, focus_date)}
    games: list[dict[str, Any]] = []
    for i, g in enumerate(raw, start=1):
        t = _norm_time(g.get("time"))
        teams = team_rows.get(t, {})
        games.append(
            {
                "game_no": i,
                "date": focus_date,
                "time": t,
                "away": teams.get("away", ""),
                "home": teams.get("home", ""),
                "away_handicap": g.get("away_handicap"),
                "home_handicap": g.get("home_handicap"),
                "away_ratio": g.get("away_ratio"),
                "home_ratio": g.get("home_ratio"),
                "away_score": g.get("away_score"),
                "home_score": g.get("home_score"),
            }
        )
    return games


def _read_daily_games_with_teams(wb: Workbook, focus_date: str) -> list[dict[str, Any]]:
    ws = wb["每日數據"]
    games: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if _norm_date(ws.cell(row, 1).value) != focus_date:
            continue
        away = str(ws.cell(row, 4).value or "").strip()
        home = str(ws.cell(row, 8).value or "").strip()
        if not away and not home:
            continue
        games.append(
            {
                "game_no": len(games) + 1,
                "date": focus_date,
                "time": _norm_time(ws.cell(row, 2).value),
                "away": away,
                "home": home,
                "away_handicap": ws.cell(row, 5).value,
                "home_handicap": ws.cell(row, 9).value,
                "away_ratio": ws.cell(row, 6).value,
                "home_ratio": ws.cell(row, 10).value,
                "away_score": ws.cell(row, 3).value,
                "home_score": ws.cell(row, 7).value,
            }
        )
    games.sort(key=lambda g: g["time"] or "00:00")
    for i, g in enumerate(games, start=1):
        g["game_no"] = i
    return games


def read_zongbiao_games(wb: Workbook, focus_date: str) -> dict[tuple[str, str, str], dict[str, Any]]:
    if "總表" not in wb.sheetnames:
        return {}
    ws = wb["總表"]
    out: dict[tuple[str, str, str], dict[str, Any]] = {}
    r = 1
    while r <= ws.max_row:
        if _norm_date(ws.cell(r, 1).value) != focus_date:
            r += 1
            continue
        start = r
        pending_manual: dict[str, Any] = {
            "handicap_side": ws.cell(start + 1, 2).value,
            "receiver_side": ws.cell(start + 1, 4).value,
            "pass_label": ws.cell(start + 1, 6).value,
        }
        r = start + 2
        while r <= ws.max_row:
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if _norm_date(a) and _norm_date(a) != focus_date:
                break
            if b == "讓分方":
                pending_manual = {
                    "handicap_side": ws.cell(r + 1, 2).value,
                    "receiver_side": ws.cell(r + 1, 4).value,
                    "pass_label": ws.cell(r + 1, 6).value,
                }
                r += 2
                continue
            if isinstance(a, str) and re.match(r"第\d+場", a.strip()):
                time_str = _norm_time(ws.cell(r + 1, 1).value)
                val_row = r + 2
                away = str(ws.cell(r + 4, 3).value or "").strip()
                home = str(ws.cell(r + 4, 6).value or "").strip()
                key = game_key(time_str, away, home)
                out[key] = {
                    "game_label": str(a).strip(),
                    "venue": str(ws.cell(r, 2).value or "").strip(),
                    "fair_away": ws.cell(val_row, 2).value,
                    "fair_home": ws.cell(val_row, 4).value,
                    "handicap_line": ws.cell(val_row, 5).value,
                    "flow_home": ws.cell(val_row, 7).value,
                    "reasonableness": ws.cell(val_row, 8).value,
                    "away": away,
                    "home": home,
                    **pending_manual,
                }
                pending_manual = {}
                r += 6
                continue
            r += 1
        break
    return out


def read_snapshot_rows(wb: Workbook) -> list[dict[str, Any]]:
    if "MLB快照" not in wb.sheetnames:
        return []
    ws = wb["MLB快照"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(_row_dict(SNAPSHOT_COLS, row))
    return rows


def read_result_rows(wb: Workbook) -> list[dict[str, Any]]:
    if "比賽結果" not in wb.sheetnames:
        return []
    ws = wb["比賽結果"]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    if not headers[0]:
        headers = RESULT_COLS
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(_row_dict(headers[: len(RESULT_COLS)], row))
    return rows


def all_pregame_snapshots(
    snapshots: list[dict[str, Any]],
    focus_date: str,
    time_str: str,
    away: str,
    home: str,
) -> list[dict[str, Any]]:
    """該場開賽前全部快照（去重、由舊到新）。"""
    kickoff = jiujiu_kickoff(focus_date, time_str)
    start = parse_start_time(kickoff, date(int(focus_date[:4]), int(focus_date[4:6]), int(focus_date[6:8])))
    matched: list[dict[str, Any]] = []
    for row in snapshots:
        if not teams_match(away, home, row.get("客隊", ""), row.get("主隊", "")):
            continue
        if not kickoff_matches_focus(str(row.get("開賽時間", "")), focus_date, time_str):
            continue
        if start and _parse_scrape_dt(str(row.get("抓取時間", ""))) > start:
            continue
        matched.append(row)
    by_ts: dict[str, dict[str, Any]] = {}
    for row in matched:
        ts = str(row.get("抓取時間", "") or "").strip()
        by_ts[ts] = row
    return sorted(by_ts.values(), key=lambda r: _parse_scrape_dt(str(r.get("抓取時間", ""))))


def format_snap_header(ts: str) -> str:
    s = str(ts or "").strip()
    if not s:
        return ""
    dt = _parse_scrape_dt(s)
    if dt.year > 2000:
        return dt.strftime("%m-%d %H:%M")
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})\s+(\d{1,2}):(\d{2})", s)
    if m:
        return f"{m.group(2)}-{m.group(3)} {int(m.group(4)):02d}:{m.group(5)}"
    return s[:16]


def format_jiujiu_cells(snapshot: dict[str, Any]) -> tuple[str, str, str, str, str, str]:
    if not snapshot:
        return ("", "", "", "", "", "")
    g = format_game_fields_for_timeline(snapshot)
    spread_a = format_timeline_handicap_cell(g.get("讓球_客盤", ""), g.get("讓球_客賠率", ""))
    spread_h = format_timeline_handicap_cell(g.get("讓球_主盤", ""), g.get("讓球_主賠率", ""))
    ml_a = format_odds(str(g.get("獨贏_客", "") or ""))
    ml_h = format_odds(str(g.get("獨贏_主", "") or ""))
    two_a = format_odds(str(g.get("一輸_客賠率", "") or ""))
    two_h = format_odds(str(g.get("一輸_主賠率", "") or ""))
    return spread_a, spread_h, ml_a, ml_h, two_a, two_h


def snapshot_timeline_cells(snapshot: dict[str, Any]) -> list[str]:
    return list(format_jiujiu_cells(snapshot))


def _empty_cells() -> list[str]:
    return ["", "", "", "", "", ""]


def build_jiujiu_dj_layout(all_snaps: list[dict[str, Any]]) -> dict[str, Any]:
    """
    D/J 欄版面：
    - D = 第 1 次；E~I = 最近 5 次（不足 5 次方案甲：E 起依序填）；J = 最後 1 次
    回傳 6 列（客/主 × 讓分/獨贏/2分）的各欄數值。
    """
    empty = _empty_cells()
    if not all_snaps:
        return {
            "header_d": "",
            "headers_e_i": [""] * RECENT_CAPTURES,
            "header_j": "",
            "col_d": empty[:],
            "cols_e_i": [empty[:] for _ in range(RECENT_CAPTURES)],
            "col_j": empty[:],
        }

    first = all_snaps[0]
    last = all_snaps[-1]
    recent = all_snaps[-RECENT_CAPTURES:]

    headers_e_i = [format_snap_header(str(s.get("抓取時間", ""))) for s in recent]
    while len(headers_e_i) < RECENT_CAPTURES:
        headers_e_i.append("")

    cols_e_i = [snapshot_timeline_cells(s) for s in recent]
    while len(cols_e_i) < RECENT_CAPTURES:
        cols_e_i.append(empty[:])

    return {
        "header_d": format_snap_header(str(first.get("抓取時間", ""))),
        "headers_e_i": headers_e_i,
        "header_j": format_snap_header(str(last.get("抓取時間", ""))),
        "col_d": snapshot_timeline_cells(first),
        "cols_e_i": cols_e_i,
        "col_j": snapshot_timeline_cells(last),
    }


def ratio_display(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        v = float(value)
        if 0 < v <= 1:
            return format_odds(str(v))
        return format_odds(str(v / 100 if v > 1 else v))
    s = str(value).replace("%", "").strip()
    try:
        v = float(s)
        if v > 1:
            v = v / 100
        return format_odds(str(v))
    except ValueError:
        return s


def find_result_row(
    results: list[dict[str, Any]],
    focus_date: str,
    time_str: str,
    away: str,
    home: str,
) -> dict[str, Any] | None:
    for row in results:
        if not teams_match(away, home, row.get("客隊", ""), row.get("主隊", "")):
            continue
        if kickoff_matches_focus(str(row.get("開賽時間", "")), focus_date, time_str):
            return row
    return None


def find_jiujiu_game_by_fav(
    results: list[dict[str, Any]],
    snapshots: list[dict[str, Any]],
    focus_date: str,
    time_str: str,
    fav_nick: str,
) -> tuple[str, str] | None:
    """每日數據配對失敗時，依開賽時間 + 讓分球隊暱稱從玖九找回客主隊全名。"""
    nick = team_nickname(fav_nick) or str(fav_nick or "").strip()
    if not nick:
        return None
    t = _norm_time(time_str)

    def _match_row(row: dict[str, Any]) -> tuple[str, str] | None:
        if not kickoff_matches_focus(str(row.get("開賽時間", "")), focus_date, t):
            return None
        away = str(row.get("客隊", "") or "").strip()
        home = str(row.get("主隊", "") or "").strip()
        if team_nickname(away) == nick or team_nickname(home) == nick:
            return away, home
        return None

    for row in results:
        hit = _match_row(row)
        if hit:
            return hit
    for row in snapshots:
        hit = _match_row(row)
        if hit:
            return hit
    return None


def format_score_display(game: dict[str, Any], result: dict[str, Any] | None) -> str:
    if result and str(result.get("比分", "") or "").strip():
        s = str(result["比分"]).strip()
        if "@" in s:
            return s
        if "-" in s:
            a, _, b = s.partition("-")
            return f"{a.strip()} @ {b.strip()}"
        return s
    away = game.get("away_score")
    home = game.get("home_score")
    if result:
        if away is None and result.get("客得分") is not None:
            away = result.get("客得分")
        if home is None and result.get("主得分") is not None:
            home = result.get("主得分")
    if away is not None and home is not None and str(away) != "" and str(home) != "":
        return f"{away} @ {home}"
    return ""


def post_settlement_cells(result: dict[str, Any]) -> dict[str, str]:
    """賽後 K/L 欄（各盤種區塊）。"""
    if not result:
        return {
            "spread_k_away": "",
            "spread_k_home": "",
            "spread_l": "",
            "ml_l": "",
            "two_k_away": "",
            "two_k_home": "",
            "two_l": "",
        }
    spread_side = str(result.get("讓分方", "") or "").strip()
    spread_k_away, spread_k_home = "", spread_side
    m = re.match(r"(.+?讓\d+[+-]\d+)", spread_side)
    if m:
        spread_k_home = m.group(1)
        spread_k_away = spread_side.replace(spread_k_home, "").strip() or spread_k_home
    ml_result = str(result.get("獨贏結果", "") or "").strip()
    ml_l = "勝" if ml_result and "獨贏" in ml_result else ml_result
    two_result = str(result.get("兩分贏結果", "") or "").strip()
    two_l = "沒過" if "受讓" in two_result else ("勝" if two_result else "")
    if "讓2分" in two_result and "沒" not in two_l:
        two_l = "勝"
    return {
        "spread_k_away": spread_k_away,
        "spread_k_home": spread_k_home,
        "spread_l": str(result.get("讓分結算", "") or ""),
        "ml_l": ml_l,
        "two_k_away": "",
        "two_k_home": two_result[:20] if two_result else "",
        "two_l": two_l or two_result,
    }


def calc_daily_ox(game: dict[str, Any], result: dict[str, Any] | None = None) -> str:
    from 每日數據轉正負盤 import calculate_ox

    away_score = game.get("away_score")
    home_score = game.get("home_score")
    if result:
        if away_score is None and result.get("客得分") is not None:
            away_score = result.get("客得分")
        if home_score is None and result.get("主得分") is not None:
            home_score = result.get("主得分")
    away_ratio = game.get("away_ratio")
    home_ratio = game.get("home_ratio")
    if isinstance(away_ratio, (int, float)) and away_ratio <= 1:
        away_ratio = away_ratio * 100
    if isinstance(home_ratio, (int, float)) and home_ratio <= 1:
        home_ratio = home_ratio * 100
    return calculate_ox(
        game.get("away_handicap"),
        game.get("home_handicap"),
        away_score,
        home_score,
        away_ratio,
        home_ratio,
    )


def merge_games(
    daily: list[dict[str, Any]],
    zongbiao: dict[tuple[str, str, str], dict[str, Any]],
    snapshots: list[dict[str, Any]],
    results: list[dict[str, Any]],
    focus_date: str,
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for g in daily:
        key = game_key(g["time"], g["away"], g["home"])
        zb = zongbiao.get(key, {})
        all_snaps = all_pregame_snapshots(
            snapshots, focus_date, g["time"], g["away"], g["home"]
        )
        jj = build_jiujiu_dj_layout(all_snaps)
        result = find_result_row(results, focus_date, g["time"], g["away"], g["home"])
        res = result or {}
        merged.append(
            {
                **g,
                "game_label": zb.get("game_label") or f"第{g['game_no']}場",
                "matchup": f"{g['away']} @ {_display_home(g['home'])}",
                "fair_away": zb.get("fair_away", ""),
                "fair_home": zb.get("fair_home", ""),
                "handicap_line": zb.get("handicap_line")
                or g.get("home_handicap")
                or g.get("away_handicap")
                or "",
                "flow_away": ratio_display(g.get("away_ratio")),
                "flow_home": ratio_display(zb.get("flow_home") or g.get("home_ratio")),
                "reasonableness_src": zb.get("reasonableness", ""),
                "handicap_side_src": zb.get("handicap_side", ""),
                "receiver_side_src": zb.get("receiver_side", ""),
                "pass_label_src": zb.get("pass_label", ""),
                "jiujiu": jj,
                "result": res,
                "score_display": format_score_display(g, result),
                "settlement": post_settlement_cells(res),
                "daily_ox": calc_daily_ox(g, result),
            }
        )
    return merged


def check_file_unlocked(path: str) -> None:
    try:
        with open(path, "r+b"):
            pass
    except PermissionError as e:
        raise SystemExit(f"無法開啟（請先關閉 Excel）：{path}") from e
    except FileNotFoundError:
        return
